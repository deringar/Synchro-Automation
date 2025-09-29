# -*- coding: utf-8 -*-
"""
Created on Fri Sep 27 2024
Last modified on Thurs Oct 3 2024

@authors: philip.gotthelf, alex.dering - Colliers Engineering & Design
"""

# main_window.py

import tkinter as tk  # Import the Tkinter module for GUI development.
# import tkinter.ttk as ttk
# from difflib import SequenceMatcher
from tkinter import filedialog
import csv  # Module to handle CSV file operations.
# import openpyxl as xl  # Used for working with Excel files (.xlsx format).
import os
import re  # Regular expression module for pattern matching in strings.
# import time  # Module for time-related functions.
# import json  # JSON module to parse and manipulate JSON data.
# from collections import OrderedDict
# from shutil import copy  # Used to copy files or directories.
from openpyxl import load_workbook, Workbook
import pandas as pd
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent
DEFAULT_TEMP_DIR = PROJECT_ROOT / "temp"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "out"

_DEBUG = False

def _debug_print(*args, **kwargs):
    if _DEBUG:
        print(*args, **kwargs)


PLACEHOLDER_METRIC_VALUES = {'', '-', 'na', 'n/a', 'nan', 'none'}

def _normalize_metric_value(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _has_metric_data(*values):
    return any(_normalize_metric_value(value).lower() not in PLACEHOLDER_METRIC_VALUES for value in values)


def _filter_rows_with_metric_data(rows, metric_slice=slice(3, 6)):
    filtered_rows = []
    for row in rows:
        if len(row) >= metric_slice.stop:
            metrics = row[metric_slice]
            if not _has_metric_data(*metrics):
                continue
        filtered_rows.append(row)
    return filtered_rows



def _freeze_nested(value):
    if isinstance(value, dict):
        return tuple(sorted((key, _freeze_nested(val)) for key, val in value.items()))
    if isinstance(value, list):
        return tuple(_freeze_nested(item) for item in value)
    if isinstance(value, set):
        return tuple(sorted(_freeze_nested(item) for item in value))
    return value



_LOS_PRIORITY = {'A': 0, 'B': 1, 'C': 2, 'D': 3, 'E': 4, 'F': 5, 'N': 0}


def _aggregate_los(values):
    worst_value = '-'
    worst_rank = -1
    for value in values:
        normalized = _normalize_metric_value(value).upper()
        if normalized in _LOS_PRIORITY:
            rank = _LOS_PRIORITY[normalized]
            if rank > worst_rank:
                worst_rank = rank
                worst_value = normalized
    return worst_value if worst_rank >= 0 else '-'


def _aggregate_delay(values):
    numeric_values = []
    for value in values:
        normalized = _normalize_metric_value(value)
        try:
            numeric_values.append(float(normalized))
        except (TypeError, ValueError):
            continue
    if not numeric_values:
        return '-'
    average_delay = sum(numeric_values) / len(numeric_values)
    return f"{average_delay:.1f}"




def select_reports_and_extract(output_directory=None):
    """Launch a file dialog so the user can pick Synchro text reports and export them to CSV."""
    root = tk.Tk()
    root.withdraw()
    try:
        selected_files = filedialog.askopenfilenames(
            title="Select Synchro report text files",
            filetypes=(('Text files', '*.txt'), ('All files', '*.*')),
        )
    finally:
        try:
            root.destroy()
        except Exception:
            pass

    if not selected_files:
        print("No files selected; nothing to extract.")
        return []

    temp_dir = DEFAULT_TEMP_DIR
    temp_dir.mkdir(parents=True, exist_ok=True)

    if output_directory:
        candidate_dir = Path(output_directory)
        if not candidate_dir.is_absolute():
            candidate_dir = PROJECT_ROOT / candidate_dir
        final_output_dir = candidate_dir
    else:
        final_output_dir = DEFAULT_OUTPUT_DIR
    final_output_dir.mkdir(parents=True, exist_ok=True)

    exported_files = []
    for selected in selected_files:
        source_path = Path(selected)
        helper_csv_path = temp_dir / f"{source_path.stem}.csv"

        try:
            final_csv_path = extract_data_to_csv(
                source_path,
                helper_csv_path,
                output_dir=final_output_dir,
            )
            exported_files.append(str(final_csv_path))
            print(f"Exported {source_path} -> {final_csv_path}")
        except Exception as exc:
            print(f"Failed to export {source_path}: {exc}")

    return exported_files

def _sanitize_lane_data(lane_data):
    if not isinstance(lane_data, tuple):
        lane_data = ('-', '-', '-', '-')
    vc_value, los_value, delay_value, capacity_value = lane_data
    vc_value = _normalize_metric_value(vc_value)
    los_value = _normalize_metric_value(los_value)
    delay_value = _normalize_metric_value(delay_value)
    capacity_value = _normalize_metric_value(capacity_value)
    has_data = _has_metric_data(vc_value, los_value, delay_value)
    return vc_value, los_value, delay_value, capacity_value, has_data

""" Part 1 """

def write_to_excel(file_path, lane_groups, delay, vc_ratio, los):
    # Helper function to transform each sublist into a dictionary with lane directions
    def separate_characters(sublist):
        result_dict = {}
        for item in sublist:
            # Extract the prefix and 'L', 'T', 'R' characters
            rest_chars = ''.join([char for char in item if char not in 'LTR'])
            separated_chars = [char for char in item if char in 'LTR']

            if rest_chars in result_dict:
                # Append characters if the key exists
                result_dict[rest_chars].extend(separated_chars)
            else:
                # Create new entry for the key
                result_dict[rest_chars] = separated_chars

        return result_dict

    # Function to enumerate the lane groups and create a dictionary
    def enumerate_result_list(result):
        enumerated_dict = {}
        for index, item in enumerate(result, start=1):
            if isinstance(item, list):  # Ensure each item is processed as a list
                enumerated_dict[index] = separate_characters(
                    item)  # Transform each list into a dictionary
            else:
                # Handle already existing dictionaries
                enumerated_dict[index] = item
        return enumerated_dict

    # Helper function to write headers
    def write_headers(ws, start_col='C'):
        headers = ['V/c', 'LOS', 'Delay']
        for idx, header in enumerate(headers):
            col = chr(ord(start_col) + idx)  # Dynamic column calculation
            ws[f'{col}2'] = header  # Write the header to row 2

    # Get the file name without extension
    file_with_ext = os.path.basename(file_path)
    file_name = os.path.splitext(file_with_ext)[0]

    # Enumerate the lane groups
    intersection_data = enumerate_result_list(lane_groups)

    # Create a new Excel workbook and add a sheet
    wb = Workbook()
    ws = wb.active  # Get the active worksheet

    # Write the file name in cell A1
    ws['A1'] = file_name

    # Write the headers for V/C, LOS, and Delay in row 2 (starting at column C)
    write_headers(ws, 'C')
    write_headers(ws, 'F')
    write_headers(ws, 'I')

    # Define the order of keys to process
    key_order = ['EB', 'WB', 'NB', 'SB', 'NE', 'NW', 'SE', 'SW']

    # Keep track of the last used row
    current_row = 2  # Starting at row 2 (A2 will be filled first)

    # Iterate through the enumerated intersection data
    for intersection_id, data in intersection_data.items():
        # Write the intersection ID in column A
        ws[f'A{current_row}'] = intersection_id

        # Move to the next row after writing the intersection ID
        current_row += 1

        # Ensure that 'data' is a dictionary before attempting to access its keys
        if isinstance(data, dict):
            # For each key in the specified order:
            for key in key_order:
                # Write the key in column A
                ws[f'A{current_row}'] = key

                # Check if there's a corresponding value in the lane data
                if key in data and data[key]:  # If values exist
                    # Write the corresponding values in column B, starting from the current row
                    for idx, item in enumerate(data[key]):
                        # Write each value downwards in column B
                        ws[f'B{current_row + idx}'] = item
                    # Update the current_row to the next empty row after writing all values
                    current_row += len(data[key])
                else:
                    # If no values exist, write an empty cell in column B
                    # Explicitly write an empty string
                    ws[f'B{current_row}'] = ''
                    current_row += 1  # Move to the next row for the next key

    # Write V/C, LOS, and Delay data into their respective columns, ensuring empty entries are included
    # Start from row 3 (the row after the headers)

    # Write V/C Ratio
    current_row_vc = 3  # Start writing V/C values
    for idx, vc_list in enumerate(vc_ratio):
        if idx < len(intersection_data):  # Ensure we don't exceed the intersection data length
            for item in vc_list:
                # Write each value downwards in column C
                ws[f'C{current_row_vc}'] = item
                current_row_vc += 1  # Move to the next row
            # Fill empty cells if the length of vc_list is shorter than the max
            while current_row_vc < (3 + len(vc_ratio)):
                ws[f'C{current_row_vc}'] = ''  # Write an empty string
                current_row_vc += 1

    # Write LOS values
    current_row_los = 3  # Start writing LOS values
    for idx, los_list in enumerate(los):
        if idx < len(intersection_data):  # Ensure we don't exceed the intersection data length
            for item in los_list:
                # Write each value downwards in column D
                ws[f'D{current_row_los}'] = item
                current_row_los += 1  # Move to the next row
            # Fill empty cells if the length of los_list is shorter than the max
            while current_row_los < (3 + len(los)):
                ws[f'D{current_row_los}'] = ''  # Write an empty string
                current_row_los += 1

    # Write Delay values
    current_row_delay = 3  # Start writing Delay values
    for idx, delay_list in enumerate(delay):
        if idx < len(intersection_data):  # Ensure we don't exceed the intersection data length
            for item in delay_list:
                # Write each value downwards in column E
                ws[f'E{current_row_delay}'] = item
                current_row_delay += 1  # Move to the next row
            # Fill empty cells if the length of delay_list is shorter than the max
            while current_row_delay < (3 + len(delay)):
                ws[f'E{current_row_delay}'] = ''  # Write an empty string
                current_row_delay += 1

    # Save the workbook
    excel_file_path = f"{file_name}_results.xlsx"
    wb.save(excel_file_path)
    print(f"Intersection data written to {excel_file_path}")


def separate_characters(result):
    # Initialize a list to hold the dictionaries
    transformed_results = []

    # Iterate through each sublist in the result
    for sublist in result:
        # Initialize a dictionary for this sublist
        result_dict = {}

        # Process each string in the sublist
        for item in sublist:
            # Extract the characters that are not 'L', 'T', or 'R' for the prefix
            # Characters other than L, T, R
            rest_chars = ''.join([char for char in item if char not in 'LTR'])
            # Characters that are L, T, or R
            separated_chars = [char for char in item if char in 'LTR']

            # Use the prefix as the key in the dictionary
            if rest_chars in result_dict:
                # Append to the existing entry if the key already exists
                result_dict[rest_chars].extend(separated_chars)
            else:
                # Create a new entry if the key does not exist
                result_dict[rest_chars] = separated_chars

        # Append the dictionary to the list of transformed results
        transformed_results.append(result_dict)

    return transformed_results


def save_as_csv(excel_file_path, csv_file_path):
    workbook = load_workbook(filename=excel_file_path)
    sheet = workbook.active

    with open(csv_file_path, mode='w', newline="") as file:
        writer = csv.writer(file)

        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)


def write_direction_data_to_files(sheet, matched_results, relevant_columns, headers, output_start_row=4):
    """
    Writes Volume, PHF, and HeavyVehicles data for each intersection and direction-turn
    from the specified column ranges in relevant_columns, and saves the results to
    separate files named based on the header in row 1 of each column range.

    Args:
    - sheet: The active sheet from which data is being read.
    - matched_results: A dictionary containing intersections and their corresponding turn data.
    - relevant_columns: A list of starting columns (e.g., [6, 9, 12] for 'F', 'I', 'L')
      from which Volume, PHF, and HeavyVehicles are read.
    - output_start_row: The row in the output sheet to start writing the data (default is 4).
    """
    for start_column in relevant_columns:
        # Define column positions relative to the starting column
        volume_col = start_column       # Volume is in start_column (e.g., F)
        phf_col = start_column + 1      # PHF is in start_column + 1 (e.g., G)
        # HeavyVehicles is in start_column + 2 (e.g., H)
        heavy_vehicles_col = start_column + 2

        # Get the header name from row 1 of the start_column (e.g., F1, I1, etc.)
        file_name_header = sheet.cell(row=1, column=start_column).value
        if not file_name_header:
            print(
                f"Skipping columns starting at {start_column} as no header was found in row 1.")
            continue

        # Create a new workbook for this specific column set
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "Results"
        output_sheet["A1"] = "[Lanes]"
        output_sheet["A2"] = "Lane Group Data"

        # Label cells with corresponding headers (A3-P3)
        for col, header in enumerate(headers, start=1):
            output_sheet.cell(row=3, column=col).value = header

        # Reset the output start row for each file
        output_start_row = 4

        # Iterate over each intersection and its direction-turn results
        for intersection_id, turns in matched_results.items():
            # Write Intersection ID and Labels in the output sheet
            output_sheet.cell(row=output_start_row, column=1).value = "Volume"
            output_sheet.cell(row=output_start_row + 1, column=1).value = "PHF"
            output_sheet.cell(row=output_start_row + 2,
                              column=1).value = "HeavyVehicles"
            output_sheet.cell(row=output_start_row,
                              column=2).value = intersection_id
            output_sheet.cell(row=output_start_row + 1,
                              column=2).value = intersection_id
            output_sheet.cell(row=output_start_row + 2,
                              column=2).value = intersection_id

            # Process each direction-turn within the intersection
            for direction_turn, info in turns.items():
                row_found = info['row']

                # Read data from the specified columns for the current row
                volume = sheet.cell(row=row_found, column=volume_col).value
                phf = sheet.cell(row=row_found, column=phf_col).value
                heavy_vehicles = sheet.cell(
                    row=row_found, column=heavy_vehicles_col).value

                # Write the data into the output sheet under the correct direction-turn column
                header_column = info['header_column']
                output_sheet.cell(row=output_start_row,
                                  column=header_column).value = volume
                output_sheet.cell(row=output_start_row + 1,
                                  column=header_column).value = phf
                output_sheet.cell(row=output_start_row + 2,
                                  column=header_column).value = heavy_vehicles

                # Debugging output
                print(f"Wrote to Results for intersection {intersection_id}, direction {direction_turn}: "
                      f"Volume: {volume}, PHF: {phf}, HeavyVehicles: {heavy_vehicles}")

            # Move to the next output row for the next intersection
            output_start_row += 3  # 3 rows for data + 1 row for separation

        # Save the output workbook to a file named by the header in row 1 of the start column
        output_file_path = f"{file_name_header}.xlsx"
        output_workbook.save(output_file_path)
        save_as_csv(output_file_path, f"{file_name_header}.csv")
        os.remove(f"{file_name_header}.xlsx")
        print(f"Output file saved as {file_name_header}.csv")

    return


""" Part 2 """

def read_input_file(file_path):
    # Load the input workbook and select the active sheet
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    # Define headers for the output sheet
    headers = [
        "RECORDNAME", "INTID", "NBL", "NBT", "NBR",
        "SBL", "SBT", "SBR", "EBL", "EBT", "EBR",
        "WBL", "WBT", "WBR", "NWR", "NWL", "NWT", "NEL", "NET", "NER",
        "SEL", "SER", "SET", "SWL", "SWR", "SWT", "PED", "HOLD"
    ]

    consecutive_empty_cells = 0
    intersections = {}

    # First pass: Find all intersection IDs and their corresponding row numbers
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value is None:
            consecutive_empty_cells += 1
            if consecutive_empty_cells >= 25:
                break
        else:
            consecutive_empty_cells = 0
            if isinstance(cell_value, int):
                intersections[cell_value] = row

    print(f"Found intersections: {intersections}")

    directions = ["EB", "WB", "NB", "SB", "NE", "NW", "SE", "SW"]
    # output_start_row = 4  # Start writing from row 4

    # Dictionary to store results for each intersection
    intersection_results = {}

    # Second pass: Process each intersection ID and search for directions
    for intersection_id, row_with_int in intersections.items():
        found_directions = {}

        # Search column C for directions starting from the intersection row
        for search_row in range(row_with_int, sheet.max_row + 1):
            direction_value = sheet.cell(search_row, column=3).value
            if direction_value in directions and direction_value not in found_directions:
                found_directions[direction_value] = search_row
                if len(found_directions) == len(directions):
                    break

        # Dictionary to store combined direction-turn keys (e.g., EBL, WBT)
        direction_turn_results = {}

        # For each found direction, search column D for 'L', 'T', 'R'
        for direction, found_row in found_directions.items():
            # Default is None (not found)
            turn_values = {"L": None, "T": None, "R": None}
            for search_row in range(found_row, sheet.max_row + 1):
                turn_value = sheet.cell(search_row, column=4).value
                if turn_value in ["L", "T", "R"]:
                    # Store the row number for each turn type found
                    turn_values[turn_value] = search_row
                # Break when all turn values have been found
                if all(turn_values.values()):
                    break

            # Combine direction and turn type to form keys like "EBL", "NBT", etc.
            for turn_type, row_found in turn_values.items():
                if row_found is not None:  # Only store if the turn was found
                    combined_key = f"{direction}{turn_type}"
                    direction_turn_results[combined_key] = row_found

        # Store the results for the current intersection
        intersection_results[intersection_id] = direction_turn_results

        # Display the results for debugging
        print(
            f"Direction-turn results for intersection {intersection_id}: {direction_turn_results}")

    # Match direction-turn results with corresponding headers
    header_mapping = {header: idx + 1 for idx, header in enumerate(headers)}

    matched_results = {}

    for intersection_id, turn_results in intersection_results.items():
        matched_results[intersection_id] = {}
        for direction_turn, row in turn_results.items():
            if direction_turn in header_mapping:
                matched_results[intersection_id][direction_turn] = {
                    "row": row,
                    "header_column": header_mapping[direction_turn]
                }

    relevant_columns = [6, 9, 12, 15]  # F-H, I-K, L-N

    write_direction_data_to_files(
        sheet, matched_results, relevant_columns, headers=headers, output_start_row=4)

    # Return intersection results if needed elsewhere
    return intersection_results


"""
Part 3: Data extraction

    * parse_minor_lane_mvmt(lines, start_line, end_line)
    * process_directions(twsc_summary_results)
    * parse_overall_data_v2(file_path)
    * parse_twsc_approach(df)
    * extract_data_to_csv(file_path, output_file, output_dir=None)
    * parse_lane_configs(int_lane_groups, intersection_ids)
"""


def parse_minor_lane_mvmt(lines, start_line, end_line):
    """
        Parse the "Minor Lane/Major Mvmt" data between the start and end lines.
        This function extracts the delay, V/C ratio, and LOS from lines containing these terms.
        Helper function to the parse_overall_data function.
    """

    search_phrase = "Minor Lane/Major Mvmt"
    search_terms = [r'\bControl Delay\b', r'\bCtrl Dly\b',
                    r'\bV/C Ratio\b', r'\bLOS\b', r'\bCapacity\b']

    # Initialize lists for each search term
    result = []
    delay_results = []
    vc_ratio_results = []
    los_results = []
    capacity_results = []

    # Search for "Minor Lane/Major Mvmt" in the provided line range
    for line_number in range(start_line, end_line):
        line = lines[line_number]
        if search_phrase in line:
            # Process the "Minor Lane/Major Mvmt" line to get the directions
            after_phrase = line.split(search_phrase)[1].strip()
            cleaned_value = ' '.join(after_phrase.split()).strip()
            result = (cleaned_value.split())
            # print(f"Result: {result}")
            # Now search for Delay, V/C Ratio, and LOS in lines below the current line
            for term in search_terms:
                term_results = []  # Temporary list to hold results for the current term
                for search_line_number in range(line_number + 1, end_line):
                    # Remove whitespace for accurate matching
                    line = lines[search_line_number].strip()
                    if re.search(term, line, re.IGNORECASE):
                        # Extract values after the term
                        # print(f"Found term: {term} in line: {line}")

                        if 'control delay' in term.lower() or 'ctrl dly' in term.lower() or 'v/c ratio' in term.lower() or 'capacity' in term.lower():
                            # For control delay or V/C ratio, we extract numbers (floats or '-')
                            numbers = re.findall(r'(\d+\.\d+|\d+|-)', line)
                            term_results.extend(
                                [float(num) if num != '-' else num for num in numbers])

                        elif 'los' in term.lower():
                            # For LOS, we extract single capital letters (A-F) or '-'
                            capital_letters = re.findall(r'\b[A-F]\b|-', line)
                            term_results.extend(capital_letters)

                # Add the term results to the corresponding results list
                if 'control delay' in term.lower() or 'ctrl dly' in term.lower():
                    if term_results:
                        delay_results.append(term_results)
                elif 'v/c ratio' in term.lower():
                    if term_results:
                        vc_ratio_results.append(term_results)
                elif 'los' in term.lower():
                    if term_results:
                        los_results.append(term_results)
                elif 'capacity' or r'^cap' in term.lower():
                    if term_results:
                        capacity_results.append(term_results)

    # Combine the results into tuples for easier reading
    merged_results = []

    # Iterate over the outer lists
    for i in range(len(vc_ratio_results)):  # Number of sublists
        # print(f"Index {i}:")
        # print(f"V/C Ratio: {vc_ratio_results[i]}")
        # print(f"LOS: {los_results[i]}")
        # print(f"Delay: {delay_results[i]}")  # Should now match other lists
        # print(f"Capacity: {capacity_results[i]}")
        # print("-" * 40)
        for j in range(len(vc_ratio_results[i])):  # Number of elements in each sublist
            merged_results.append((
                vc_ratio_results[i][j],  # Extract single element
                los_results[i][j],       # Extract single element
                delay_results[i][j],     # Extract single element
                capacity_results[i][j]   # Extract single element
            ))

    # Return the parsed results for integration with other parsing logic
    return result, merged_results


def integrate_awsc_data(awsc_data, combined_data):
    """
    Merges the AWSC lane data into the existing data handling structure,
    ensuring each intersection is collected correctly.
    """
    # print(f"\nAWSC Data: {awsc_data}")
    # print(f"\nCombined Data: {combined_data}")
    for awsc_entry in awsc_data:
        intersection_id = awsc_entry.get("ID")
        if not intersection_id:
            continue

        formatted_entry = {"ID": intersection_id}

        for lane, values in awsc_entry.items():
            if lane == "ID":
                continue
            # Directly unpack the tuple
            v_c_ratio, los, delay, cap = values
            formatted_entry[lane] = (v_c_ratio, los, delay, cap)

        combined_data.append(formatted_entry)

    return combined_data


def process_directions_sc(lane_data, lane_configs):
    """
    Processes lane names using process_directions to get actual movement labels.
    """
    processed_data = []

    # ✅ Process all lane data at once
    processed_list, _, _ = process_directions(lane_data, lane_configs)

    for original_entry, processed_entry in zip(lane_data, processed_list):
        processed_result = {"ID": original_entry["ID"]}

        # Map processed lane names back to their values
        for direction, suffixes in processed_entry.items():
            if direction == "ID":
                continue
            if isinstance(suffixes, list):
                for suffix in suffixes:
                    key = direction + suffix
                    if key in original_entry:
                        processed_result[key] = original_entry[key]
            else:
                key = direction + suffixes
                if key in original_entry:
                    processed_result[key] = original_entry[key]

        processed_data.append(processed_result)

    return processed_data


def parse_awsc_data(df):
    """
    Parses AWSC data from the dataframe and extracts lane group details,
    v/c ratio, LOS, delay, and capacity.
    """
    awsc_data = []
    intersection_id = None

    for index, row in df.iterrows():
        line = str(row[0]).strip()

        # Detect new intersection
        if line.isdigit():
            intersection_id = int(line)
            # print(f"Found intersection ID: {intersection_id} at row {index}")
            continue

        if line.lower() == "lane":
            lane_data = {"ID": str(intersection_id)}
            lane_columns = {}

            # Step 1: Identify all lanes (e.g., NBLn1, EBLn1)
            for col_index, cell in enumerate(row.values):
                if re.fullmatch(r'(EB|WB|NB|SB|NE|NW|SE|SW)(Ln\d+)?', str(cell)):
                    lane_columns[str(cell)] = col_index
                    # print(f"Found lane '{cell}' in column {col_index}")

            # Step 2: Collect V/C Ratio, LOS, Delay, Capacity
            next_row_index = index + 1

            # ✅ Initialize temp_data fresh for each intersection
            temp_data = {lane: ("-", "-", "-", "-") for lane in lane_columns.keys()}

            while next_row_index < len(df):
                next_row = df.iloc[next_row_index]
                row_label = str(next_row.iloc[0]).strip()

                # ✅ Stop when a new intersection starts
                if re.match(r'^\d+$', str(next_row[0])):
                    # print(f"Encountered new intersection at row {next_row_index}, stopping data collection.")
                    break

                # Process relevant rows only
                if any(keyword in row_label for keyword in ["V/C Ratio", "LOS", "Delay", "Cap"]):
                    # print(f"Processing data row '{row_label}' at index {next_row_index}")

                    for lane, col in lane_columns.items():
                        v_c_ratio, los, delay, cap = temp_data[lane]

                        # ✅ Only update if the new data is valid (non-empty and not '-')
                        if "V/C Ratio" in row_label:
                            new_value = next_row[col] if pd.notna(next_row[col]) else '-'
                            v_c_ratio = new_value if new_value != '-' else v_c_ratio
                            # print(f"{lane} V/C Ratio: {v_c_ratio}")

                        elif "LOS" in row_label:
                            new_value = next_row[col] if pd.notna(next_row[col]) else '-'
                            los = new_value if new_value != '-' else los
                            # print(f"{lane} LOS: {los}")

                        elif "Delay" in row_label:
                            new_value = next_row[col] if pd.notna(next_row[col]) else '-'
                            delay = new_value if new_value != '-' else delay
                            # print(f"{lane} Delay: {delay}")

                        elif "Cap" in row_label:
                            new_value = next_row[col] if pd.notna(next_row[col]) else '-'
                            cap = new_value if new_value != '-' else cap
                            # print(f"{lane} Capacity: {cap}")

                        # Update lane data
                        temp_data[lane] = (v_c_ratio, los, delay, cap)

                next_row_index += 1

            # ✅ Add lane data specific to this intersection
            lane_data.update(temp_data)
            awsc_data.append(lane_data)
            # print(f"\nCompleted data collection for intersection ID: {intersection_id} -> {lane_data}")

    # print(f"\nAWSC data: \n{awsc_data}")
    return awsc_data


def process_directions(twsc_summary_results, lane_configs):
    processed_list = []         # Parsed movement data -> [dict]
    original_key_list = []      # Original movement names -> [dict]
    combined_mvmt_names = []    # Translated movement names -> [dict]

    lane_split_flag = False     # Flag to signal lane splitting -> bool
    next_suffix = None          # Used for splitting 4 lane scenarios -> str, None

    # ✅ Flatten lane_configs in case it contains lists inside
    flattened_lane_configs = []
    for entry in lane_configs:
        if isinstance(entry, list):  # If it's a list of dictionaries, extend
            flattened_lane_configs.extend(entry)
        else:
            flattened_lane_configs.append(entry)  # Otherwise, add normally

    for entry in twsc_summary_results:
        processed_dict = {"ID": entry["ID"]}
        original_key_dict = {"ID": entry["ID"]}
        intersection_id = int(entry["ID"])

        # ✅ Ensure lane_configs is accessed correctly
        lane_config = next(
            (config for config in flattened_lane_configs if isinstance(config, dict) and str(config.get("Intersection ID")) == str(intersection_id)),
            None
        )

        if lane_config is None:
            print(f"⚠️ Warning: No lane configuration found for Intersection {intersection_id}, skipping.")
            continue

        combined_mvmt = []

        for key, value in entry.items():
            if key == "ID":
                continue  # Skip the ID key

            direction = key[:2]
            suffix = key[2:]

            print(f"\nProcessing movement: {key} -> Direction: {direction}, Suffix: {suffix}")

            # Get the number of configured lanes for the direction
            # Add suffixes to the original key dict
            original_key_dict.setdefault(direction, []).append(suffix)

            # 🔹 Convert "LnX" to the correct movement (L, T, R) from lane_config
            if suffix.startswith("Ln") and lane_config and direction in lane_config:
                try:
                    lane_index = int(suffix[2:]) - 1  # Convert "Ln1" to 0-based index
                    print(f"  Found lane index {lane_index} in {direction}")

                    if 0 <= lane_index < len(lane_config[direction]):
                        print(f"  Replacing '{suffix}' with '{lane_config[direction][lane_index]}'")
                        if lane_split_flag and next_suffix:
                            lane_split_flag = False
                            suffix = next_suffix
                        else:
                            suffix = lane_config[direction][lane_index]  # Replace with correct movement
                            if len(suffix) == 4:        # eg. suffix = 'LTTR', WB: [Ln1, Ln2]; (WB: <2>)
                                suffix = suffix[:2]
                                next_suffix = suffix[2:]
                                lane_split_flag = True

                                print(f"\n[*] Found four lane approach! (Intersection #{intersection_id}, suffix='{key}' -> suffix='{suffix}')")
                                print("[*] ... ")

                    else:
                        print(f"  Lane index {lane_index} out of range for {direction}, setting to '-'")
                        suffix = "-"  # Set to '-' if index is invalid
                except (ValueError, IndexError):
                    print(f"  Failed to process {suffix}, setting to '-'")
                    suffix = "-"  # Set to '-' if parsing fails

            # 🔹 Ensure suffix is stored correctly
            existing_value = processed_dict.get(direction)

            if existing_value is None:
                processed_dict[direction] = suffix  # First occurrence, store directly
            else:
                if not isinstance(existing_value, list):
                    existing_value = [existing_value]
                if isinstance(suffix, list):
                    existing_value.extend(suffix)
                else:
                    existing_value.append(suffix)
                processed_dict[direction] = existing_value

        # 🔹 Format the final combined movement names
        for direction, value in processed_dict.items():
            if direction != "ID":
                if isinstance(value, list):
                    combined_mvmt.append(direction + ''.join(value))
                else:
                    combined_mvmt.append(direction + value)

        combined_mvmt_names.append(combined_mvmt)

        # Append to results
        processed_list.append(processed_dict)
        original_key_list.append(original_key_dict)

        print(f"\nProcessed Data for Intersection {intersection_id}: {processed_dict}")
        print(f"\nOriginal Key Data: {original_key_dict}")
        # print(f"\nCombined Movement Names: {combined_mvmt_names}")

    return processed_list, original_key_list, combined_mvmt_names


def parse_lane_configs(int_lane_groups, intersection_ids):
    parsed_list = []  # This will store the parsed dictionaries for each group
    raw_data_list = []
    intersection_ids = sorted(intersection_ids, key=int)
    # print(f"Length of matching_configs: {len(int_lane_groups)}")
    # print(f"Length of intersection_ids: {len(intersection_ids)}")

    for idx, lane_dict in enumerate(int_lane_groups):

        intersection_id = intersection_ids[idx]

        # Skip if the intersection ID is already in parsed_list
        if any(parsed_dict.get("Intersection ID") == intersection_id for parsed_dict in parsed_list):
            continue

        parsed_dict = {
            "Intersection ID": intersection_id,
            # Initialize with three None values for L, T, R
            'EB': [None, None, None],
            'WB': [None, None, None],
            'NB': [None, None, None],
            'SB': [None, None, None],
            'NE': [None, None, None],
            'NW': [None, None, None],
            'SE': [None, None, None],
            'SW': [None, None, None]
        }

        # Initialize the raw data dictionary
        raw_data_dict = {
            "Intersection ID": intersection_id,
            'EB': [None, None, None],
            'WB': [None, None, None],
            'NB': [None, None, None],
            'SB': [None, None, None],
            'NE': [None, None, None],
            'NW': [None, None, None],
            'SE': [None, None, None],
            'SW': [None, None, None]
        }

        for direction, value in lane_dict.items():
            if value is None or value == '':
                value = '-'
                continue

            # Process each direction and suffix (L, T, R)
            suffixes = {
                'L': 0,  # Index 0 for Left
                'T': 1,  # Index 1 for Through
                'R': 2   # Index 2 for Right
            }

            for suffix, idx in suffixes.items():
                # Parse the value for numbers and special characters < and >
                if direction.endswith(suffix):
                    # Store the raw value directly in raw_data_dict in the correct position
                    direction_prefix = direction[:-1]
                    if direction_prefix in raw_data_dict:
                        # Store unparsed raw value
                        raw_data_dict[direction_prefix][idx] = value

                    parsed_value = ''

                    if '<' in value:
                        parsed_value += 'L'  # Leading left if < is present
                    number_part = ''.join(
                        filter(str.isdigit, value))  # Get the number
                    if number_part:
                        # Repeat based on the number
                        parsed_value += suffix * int(number_part)

                    else:
                        parsed_value += suffix

                    if '>' in value:
                        parsed_value += 'R'  # Trailing right if > is present

                    # Store the parsed value in the respective direction and suffix position
                    # Get the prefix like EB, WB, etc.
                    direction_prefix = direction[:-1]
                    if direction_prefix in parsed_dict:
                        parsed_dict[direction_prefix][idx] = parsed_value or None

        # Remove None values from each list in the parsed_dict
        for key in list(parsed_dict.keys()):
            if key != "Intersection ID":  # Don't touch the Intersection ID key
                parsed_dict[key] = [
                    value for value in parsed_dict[key] if value is not None]
                # If the list is empty (no valid values), set it to '-'
                if not parsed_dict[key]:
                    parsed_dict[key] = '-'

        # Clean up the raw_data_dict in the same way
        for key in raw_data_dict:
            if key != "Intersection ID":
                raw_data_dict[key] = [value for value in raw_data_dict[key]]
                if not raw_data_dict[key]:
                    raw_data_dict[key] = '-'

        # Debugging output
        # print(f"\nParsed Lane Config (Intersection #{intersection_id}):\n{parsed_dict} \nRaw Lane Config (Intersection #{intersection_id}):\n{raw_data_dict}")

        # Append the parsed_dict for this lane group to the final list
        parsed_list.append(parsed_dict)
        raw_data_list.append(raw_data_dict)

    return parsed_list, raw_data_list


def parse_overall_data_v2(file_path, df):
    """
        Function to handle the parsing of the summary data
    """

    int_regex = r'^\d+:'  # Regex to match lines that start with an integer followed by a colon

    search_phrases = ["Minor Lane/Major Mvmt",
                      "Intersection Summary"]

    # Lists to store results
    synchro_results = []
    hcm_results = []
    twsc_results = []
    awsc_results = []

    # List to store matching line numbers
    matching_lines = []

    # Open and read the file into memory
    with open(file_path, 'r') as file:
        lines = file.readlines()  # Read all lines

    # Find the line numbers containing the integer followed by a colon
    for line_number, line in enumerate(lines, start=1):
        if re.match(int_regex, line):  # Use re.match to check if the line starts with the regex
            matching_lines.append(line_number)

    intersection_index = 0

    # Now search from each line in matching_lines until the next line matching the regex
    for start_line in matching_lines:
        # Get the ID from the corresponding line in 'lines'
        # Accessing lines using start_line - 1
        id_match = re.match(int_regex, lines[start_line - 1])
        id_value = id_match.group(0).strip(
            ':') if id_match else None  # Get the ID before the colon

        # print(f"\n\t---\nIntersection {id_value}\n\t---")

        # Process Synchro Results first
        found_phrase = False
        for line_number in range(start_line, len(lines)):
            line = lines[line_number]

            # Check for a new ID match before processing further
            new_id_match = re.match(int_regex, line)
            if new_id_match:
                new_id_value = new_id_match.group(0).strip(':')
                # print(f"Found new ID {new_id_value} at line {line_number}")
                id_value = new_id_value  # Update the ID to the new one found

            if "Intersection Summary" in line:
                # print(f"Found 'Intersection Summary' at line {line_number}")

                found_phrase = True  # Mark that we found the phrase

                # Set the end line to the next empty line starting from this line
                end_line = line_number + 1
                while end_line < len(lines) and lines[end_line].strip() != '':
                    if "HCM" in lines[end_line]:
                        break
                    end_line += 1  # Continue until we find an empty line

                if "HCM" in lines[end_line]:
                    # print(f"Found 'HCM' at line {end_line}, skipping Synchro block for this ID\n")
                    continue

                # print(f"Synchro block ends at line {end_line}\n")

                # Initialize values to None
                vc_ratio_value = None
                los_value = None
                delay_value = None

                # Now process the following lines for the search terms until the next blank line
                for search_line_number in range(line_number + 1, end_line):
                    line = lines[search_line_number]

                    # print(f"Processing Synchro data at line {search_line_number}: {line.strip()}")

                    # Check for 'v/c ratio' and extract the next float
                    # if re.search(r'v/c ratio', line, re.IGNORECASE):
                    #     float_match = re.search(r'(\d+\.\d+|\d+)', line)
                    #     if float_match:
                    #         vc_ratio_value = float(float_match.group(0))
                    #         # print(f"Extracted v/c ratio: {vc_ratio_value}")

                    # Check for 'delay' and extract the next float
                    if re.search(r'delay', line, re.IGNORECASE):
                        float_match = re.search(r'(\d+\.\d+|\d+)', line)
                        if float_match:
                            delay_value = float(float_match.group(0))
                            # print(f"Extracted delay: {delay_value}")

                    # Check for 'LOS' and extract the next capital letter (A-F)
                    if re.search(r'LOS', line, re.IGNORECASE):
                        capital_match = re.search(r'\b[A-F]\b', line)
                        if capital_match:
                            los_value = capital_match.group(0)
                            # print(f"Extracted LOS: {los_value}")

                # print(f"Final Synchro values for ID {id_value}: v/c ratio={vc_ratio_value}, delay={delay_value}, LOS={los_value}")

                # Store Synchro results only if ID is not already present
                if not (los_value is None and delay_value is None) and found_phrase:
                    if not any(result['ID'] == id_value for result in synchro_results):
                        synchro_results.append({
                            'ID': id_value,
                            'v/c ratio': vc_ratio_value if vc_ratio_value is not None else '-',
                            'los': los_value if los_value is not None else '-',
                            'delay': delay_value if delay_value is not None else '-',
                            'index': intersection_index
                        })
                        intersection_index += 1

                # Stop further processing of Synchro block and move on to HCM
                break  # Exit after processing this block for Synchro

            # Skip lines between the ID and the next search phrase
            if found_phrase:
                break  # Stop looking at this block and continue with HCM

        # Now search for HCM Results
        found_phrase = False  # Reset for HCM block
        for line_number in range(start_line, len(lines)):
            line = lines[line_number]

            # Check if a new ID appears before parsing the HCM block
            new_id_match = re.match(int_regex, lines[line_number - 1])
            if new_id_match:
                id_value = new_id_match.group(0).strip(':')
                # print(f"Updated ID to {id_value} for HCM parsing at line {line_number}")

            # Process only HCM-related lines
            for search_phrase in search_phrases:
                if search_phrase in line:
                    found_phrase = True
                    phrase_found = search_phrase  # Store which search phrase was found

                    # Set the end line to the next empty line starting from this line
                    end_line = line_number + 1
                    while end_line < len(lines) and lines[end_line].strip() != '':
                        end_line += 1  # Continue until we find an empty line

                    # Initialize values to None
                    vc_ratio_value = '-'
                    los_value = None
                    delay_value = None
                    found_hcm = False  # Flag to track if we found HCM lines

                    # Process HCM lines
                    for search_line_number in range(line_number + 1, end_line):
                        line = lines[search_line_number]

                        # Check for 'v/c ratio' and extract the next float
                        if re.search(r'v/c ratio', line, re.IGNORECASE):
                            float_match = re.search(r'(\d+\.\d+|\d+)', line)
                            if float_match:
                                vc_ratio_value = float(float_match.group(0))

                        # Check for 'delay' and extract the next float
                        if re.search(r'delay', line, re.IGNORECASE):
                            # Find the position of 'delay' in the line
                            delay_pos = line.lower().find('delay')
                            if delay_pos != -1:  # If 'delay' is found
                                # Get everything after 'delay'
                                after_delay = line[delay_pos +
                                                   len('delay'):].strip()
                                # Search for float in the remaining substring
                                float_match = re.search(
                                    r'(\d+\.\d+|\d+)', after_delay)
                                if float_match:
                                    delay_value = float(float_match.group(0))

                        # Check for 'LOS' and extract the next capital letter (A-F)
                        if re.search(r'LOS', line, re.IGNORECASE):
                            capital_match = re.search(r'\b[A-F]\b', line)
                            if capital_match:
                                los_value = capital_match.group(0)

                        if re.search(r'capacity', line, re.IGNORECASE):
                            float_match = re.search(r'(\d+\.\d+|\d+)', line)
                            if float_match:
                                capacity_value = float(float_match.group(0))

                        # Mark the line as HCM only if it starts with "HCM"
                        if line.startswith("HCM"):
                            found_hcm = True

                    # Store HCM results only if we processed HCM lines
                    if found_hcm:
                        # Check which search phrase was found and structure the result accordingly
                        if phrase_found == "Intersection Summary":
                            hcm_results.append({
                                'ID': id_value,
                                'v/c ratio': vc_ratio_value if vc_ratio_value is not None else '-',
                                'los': los_value if los_value is not None else '-',
                                'delay': delay_value if delay_value is not None else '-',
                                'capacity': capacity_value if capacity_value is not None else '-',
                                'index': intersection_index
                            })
                            intersection_index += 1
                        elif phrase_found == "Minor Lane/Major Mvmt":
                            movement_results, merged_results = parse_minor_lane_mvmt(
                                lines, line_number, end_line)

                            # Create a dictionary where keys are from the movement results
                            hcm_entry = {'ID': id_value}

                            for i in range(len(movement_results)):
                                # print(i)
                                # Using the movement results as keys
                                hcm_entry[movement_results[i]] = merged_results[i]
                            twsc_results.append(hcm_entry)

                            # print(f"\nMerged entry: {hcm_entry}")

                        # Stop collecting on a blank line
                        if line.strip() == "":
                            print(f"Blank line encountered at line {line_number}")  # Debugging output
                            break

                        break  # Exit after processing the HCM block

                    # Skip lines between the ID and the next search phrase
                    if found_phrase:
                        # print(f"Found phrase at line {line_number}: {phrase_found}")  # Debugging output
                        break  # Stop looking at this block and move on to the next intersection

    # Print for debugging
    # print("\nSynchro Signalized Summary Results (Intersection Summary):", synchro_results)
    # print("\nHCM Signalized Summary Results (Intersection Summary):", hcm_results)
    # print("\nTWSC Summary Results (Minor Lane/...):", twsc_results)
    # print("\nAWSC Summary Results (Lane):", awsc_results, '\n')

    return twsc_results, synchro_results, hcm_results, awsc_results


def parse_twsc_approach(df):
    """
        Parses the approach data for each direction of a TWSC intersection
        Returns a list of dictionaries, one for each TWSC intersection found in the dataframe
    """
    approach_data = []  # List to hold all parsed data
    intersection_id = None  # Store the ID of the intersection we are currently looking at

    # Iterate through each row in the DataFrame
    for index, row in df.iterrows():
        line = str(row[0]).strip()  # Consider column 1 as the line to process
        # print(f"\nProcessing line {index}: {line}")

        # Check if the row in column 1 contains an integer (this is the intersection ID)
        if line.isdigit():
            intersection_id = int(line)
            # print(f"Found Intersection ID {intersection_id} at line {index}")
            continue  # Move to the next row

        # Check if the line starts with "approach" and contains at least one direction
        if line.lower() == "approach":
            # print(f"Found 'Approach' at line {index}: {line}")

            # Check if any of the specified directions are present in the line after "approach"
            present_directions = {direction: direction in row.values for direction in [
                "EB", "WB", "NB", "SB", 'NE', 'NW', 'SE', 'SW']}
            # print(f"Present directions: {present_directions}")

            # If no directions are found after "approach", skip this line
            if not any(present_directions.values()):
                print("No valid directions found, skipping line.")
                continue  # Skip the line if no directions are present

            approach_dict = {
                "Intersection ID": intersection_id,
                "EB": {"Approach Delay": None, "Approach LOS": '-'},
                "WB": {"Approach Delay": None, "Approach LOS": '-'},
                "NB": {"Approach Delay": None, "Approach LOS": '-'},
                "SB": {"Approach Delay": None, "Approach LOS": '-'},
                "NE": {"Approach Delay": None, "Approach LOS": '-'},
                "NW": {"Approach Delay": None, "Approach LOS": '-'},
                "SE": {"Approach Delay": None, "Approach LOS": '-'},
                "SW": {"Approach Delay": None, "Approach LOS": '-'},
            }

            # pd.set_option('display.max_columns', 100)  # Show all columns

            # Step 1: Record the positions (columns) of directions
            direction_columns = {}
            for direction in ["EB", "WB", "NB", "SB", 'NE', 'NW', 'SE', 'SW']:
                if present_directions[direction]:
                    # Find the column where the direction was found
                    direction_columns[direction] = row[row ==
                                                       direction].index[0]
                    # print(f"Direction {direction} found") # in column {direction_columns[direction]}.")

            # Now check subsequent rows for "HCM Control Delay" and "HCM LOS"
            next_row_index = index + 1
            while next_row_index < len(df):
                next_row = df.iloc[next_row_index]  # Get the next row

                # Check for "HCM Control Delay"
                if "hcm control delay" in str(next_row.iloc[0]).lower():
                    # print(f"Found 'HCM Control Delay' at row {next_row_index}.")

                    # Assign the delay values from the columns where directions were found
                    for direction, col in direction_columns.items():
                        delay_value = next_row[col]
                        # Check for numeric values
                        if pd.notna(delay_value) and re.match(r'\b\d+\.\d+|\b\d+', str(delay_value)):
                            approach_dict[direction]["Approach Delay"] = delay_value
                            # print(f"Setting {direction} Approach Delay: {delay_value}")
                        else:
                            # Store '-' if no valid value
                            approach_dict[direction]["Approach Delay"] = '-'
                            # print(f"No valid delay value for {direction}, setting to '-'.")

                # Check for "HCM LOS"
                elif "hcm los" in str(next_row.iloc[0]).lower():
                    # print(f"Found 'HCM LOS' at row {next_row_index}.")

                    # Assign the LOS values (A-F) from the columns where directions were found
                    for direction, col in direction_columns.items():
                        los_value = str(next_row[col]).strip().upper()
                        # Check if the value is a valid LOS (A-F)
                        if los_value in 'ABCDEF' and los_value != '':
                            approach_dict[direction]["Approach LOS"] = los_value
                            # print(f"Setting {direction} Approach LOS: {los_value}")
                        else:
                            # Store '-' if no valid LOS value
                            approach_dict[direction]["Approach LOS"] = '-'
                            # print(f"No LOS value for {direction}, setting to '-'.")

                # Move to the next row
                next_row_index += 1

                # Exit the loop if an empty row is found
                if next_row.isna().all():  # Check if the row is entirely empty or NaN
                    break

            # Step 3: Remove directions with no valid data
            approach_dict = {
                k: v for k, v in approach_dict.items() if k == "Intersection ID" or (isinstance(v, dict) and (v["Approach Delay"] is not None or v["Approach LOS"] != '-'))
            }

            # If there's any valid data, add it to approach_data
            if approach_dict:
                approach_data.append(approach_dict)
                # print(f"Added approach data: {approach_dict}")

    # print(f"\nFinal approach data (TWSC Intersections):\n{approach_data}")
    return approach_data


def extract_data_to_csv(file_path, output_file, output_dir=None):
    source_path = Path(file_path)
    file_path = str(source_path)

    if output_dir is None:
        final_output_dir = DEFAULT_OUTPUT_DIR
    else:
        final_output_dir = Path(output_dir)
        if not final_output_dir.is_absolute():
            final_output_dir = PROJECT_ROOT / final_output_dir
    final_output_dir.mkdir(parents=True, exist_ok=True)

    if output_file:
        candidate_helper = Path(output_file)
        helper_csv_path = candidate_helper if candidate_helper.is_absolute() else DEFAULT_TEMP_DIR / candidate_helper.name
    else:
        helper_csv_path = DEFAULT_TEMP_DIR / f"{source_path.stem}.csv"
    helper_csv_path.parent.mkdir(parents=True, exist_ok=True)

    final_output_path = final_output_dir / f"{source_path.stem}-filtered.csv"

    data = []  # To store the final data for CSV
    # skip_lines = 0  # Counter to track skipped lines
    intersection_count = 0
    intersection_ids = []
    collecting = False  # Flag to indicate if we're collecting data
    # Flag to track if we're collecting after "Minor Lane/Major Mvmt"
    collecting_minor_lane = False
    lane_groups = []
    """
        Parse and extract relevant intersection data from the text file
        into a Dataframe and generate a CSV file
    """
    with source_path.open('r') as file:
        for line in file:
            stripped_line = line.rstrip('\n')  # Remove the newline character

            # Step 1: Look for a line starting with a digit and a colon
            if re.match(r'^\d+:', stripped_line):
                # Extract the intersection count from the beginning of the line
                intersection_count = int(
                    re.match(r'^(\d+):', stripped_line).group(1))
                data.append([intersection_count])
                intersection_ids.append(intersection_count)

                collecting = True  # Start collecting data
                collecting_minor_lane = False
                continue  # Skip the current line

            # Step 2: Start collecting after finding "Lane Group", "Movement", or "Intersection"
            if collecting or collecting_minor_lane:
                # If the line is empty, stop collecting (except when we're in the middle of collecting after "Minor Lane/Major Mvmt")
                if stripped_line == "":
                    collecting = False
                    collecting_minor_lane = False
                    continue  # Skip the empty line

                # Split the line based on double tabs or multiple spaces
                new_row = re.split(
                    r'\t\t|\s{2}\t|\s\t|\t', stripped_line.strip())
                # Remove empty cells
                new_row = [cell for cell in new_row if cell]
                data.append(new_row)  # Append the new row to data

            # Look for "Lane Group", "Movement", or "Intersection" after finding the intersection line
            if not collecting and not collecting_minor_lane:
                if re.match(r'\s*Lane Group', stripped_line) or re.match(r'\s*Movement', stripped_line):
                    collecting = True  # Start collecting data when either is found
                    # Split the line based on double tabs or multiple spaces
                    new_row = re.split(
                        r'\t\t|\s{2}\t|\s\t|\t', stripped_line.strip())
                    # Remove empty cells after stripping
                    new_row = [cell.strip() for cell in new_row if cell]
                    if new_row:  # Check if new_row is not empty after cleaning
                        data.append(new_row)  # Append the new row to data
                        # Append cleaned values to lane_groups
                        lane_groups.append(new_row[1:])
                        # print(new_row[1:])  # Print the cleaned values
                    # print(f"Collecting data for {intersection_count} under {stripped_line.strip()}")
                # If we find "Intersection" before "Lane Group" or "Movement"
                elif re.match(r'^\s*Intersection', stripped_line):
                    collecting = True  # Start collecting, ignoring blank lines
                    # Split the line based on double tabs or multiple spaces
                    new_row = re.split(
                        r'\t\t|\s{2}\t|\s\t|\t', stripped_line.strip())
                    # Remove empty cells
                    new_row = [cell for cell in new_row if cell]
                    data.append(new_row)  # Append the new row to data
                    # print(f"Collecting data for {intersection_count} under Intersection")
                # If we find "Minor Lane/Major Mvmt", start collecting until the next blank line
                elif re.match(r'^\s*Minor Lane/Major Mvmt', stripped_line) or re.match(r'^Approach', stripped_line) or re.match(r'^Lane\s[2]*', stripped_line):
                    collecting_minor_lane = True  # Start collecting after "Minor Lane/Major Mvmt"
                    # Split the line based on double tabs or multiple spaces
                    new_row = re.split(
                        r'\t\t|\s{2}\t|\s\t|\t', stripped_line.strip())
                    # Remove empty cells
                    new_row = [cell for cell in new_row if cell]
                    data.append(new_row)  # Append the new row to data
                    # print(f"Started collecting for {intersection_count} after Minor Lane/Major Mvmt")

    # print(f"Intersection ID's stored (length = {len(intersection_ids)}): \n{intersection_ids}\n")

    # Step 5: Create a DataFrame and save to CSV
    df = pd.DataFrame(data)
    df.to_csv(helper_csv_path, index=False)
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)

    # Define the terms to search for
    terms_to_match = [
        "V/c ratio(x)",
        "LnGrp Delay(d), s/veh",
        "LnGrp LOS",
        "V/c ratio",
        "Control delay (s/veh)",
        "LOS",
        "Approach Delay (s/veh)",
        "Approach Delay, s/veh",
        "Approach LOS",
        # "Sign control"
    ]

    # Initialize an empty dictionary to store row indices
    signalized_int_data = []
    all_intersection_configs = []
    row_indices = {}
    group_config_data = {}

    current_id = None
    j = 0

    """
        Storing and grouping intersection data (signalized/unsignalized)
    """
    # Iterate through DataFrame rows
    for index, row in df.iterrows():
        # Check if the first column contains a new intersection ID that is in the intersection_ids list
        first_column_value = str(row[0]).strip()
        if first_column_value.isdigit():
            potential_id = int(first_column_value)
            if potential_id in intersection_ids:
                # If we have an existing intersection ID, save the collected data for it
                if current_id is not None and (group_config_data or row_indices):
                    # Check if row_indices has at least one term from terms_to_match
                    if any(term in row_indices.values() for term in terms_to_match):
                        intersection_data = {
                            "Intersection ID": current_id,
                            "Configurations": group_config_data,
                            **{term: line_num for line_num, term in row_indices.items()}
                        }
                        signalized_int_data.append(intersection_data)

                    # Reset data structures for the next intersection
                    row_indices = {}
                    group_config_data = {}

                # Set the new intersection ID
                current_id = potential_id
                continue  # Move to the next iteration as we've identified a new intersection

        # Detect rows containing "Lane Group" or "Movement"
        if "Lane Configurations" in row.values:
            # Create an empty dictionary to hold the configurations
            config_dict = {}
            # print(f"1^ Intersetion {current_id}\n{row}")
            # Iterate over lane_groups[j] and row values simultaneously, skipping empty values
            for i, key in enumerate(lane_groups[j]):
                if i + 1 < len(row):  # Ensure there's a corresponding value in the row
                    # Get and clean the value in the row
                    value = str(row[i + 1]).strip()
                    if value != 'None' and value != '':  # Only add the key-value pair if the value is non-empty
                        config_dict[key] = value

            # Append the config_dict to the group_config_data list if it contains data
            if config_dict:
                config_dict['ID'] = f'{current_id}'
                group_config_data = config_dict
                all_intersection_configs.append(config_dict)

            # Move to the next set of lane groups
            j += 1
        # Check for the presence of "LOS" first, as it is case-sensitive
        if "LOS" in row.values:
            row_indices[index] = "LOS"
            continue
        # For other terms, check in a case-insensitive manner
        for term in terms_to_match:
            if any(str(cell).lower() == term.lower() for cell in row if term.lower() != "los"):
                row_indices[index] = term
                break  # Exit the inner loop if a term is found

    # Final addition for the last intersection data
    if current_id is not None and (group_config_data or row_indices):
        # Ensure the dictionary contains at least one of the terms_to_match
        if any(term in row_indices.values() for term in terms_to_match):
            intersection_data = {
                "Intersection ID": current_id,
                "Configurations": group_config_data,
                **{term: line_num for line_num, term in row_indices.items()}
            }
            signalized_int_data.append(intersection_data)


    lane_configurations, raw_lane_configs = parse_lane_configs(
        all_intersection_configs, intersection_ids)
    # print(f"\nLane configs: {all_intersection_configs}")

    # Initialize an empty list to store the combined dictionaries
    combined_list = []

    """
        Iterate through the list of signalized intersections
    """
    i = 0
    for intersection in signalized_int_data:
        intersection_id = intersection.get("Intersection ID")
        combined_dict = {"Intersection ID": intersection_id}

        # Initialize approach data for all possible directions
        approach_data = {
            direction: {"Approach Delay": None, "Approach LOS": None}
            for direction in ['EB', 'WB', 'NB', 'SB', 'NE', 'NW', 'SE', 'SW']
        }

        # Access lane configurations for the current intersection
        lane_config = lane_configurations[i]
        raw_lane_config = raw_lane_configs[i]

        i += 1

        # Initialize dictionaries to store v/c ratio, LOS, and delay values
        directions = ['EB', 'WB', 'NB', 'SB', 'NE', 'NW', 'SE', 'SW']
        vc_ratio_values = {direction: [None, None, None]
                           for direction in directions}
        los_values = {direction: [None, None, None]
                      for direction in directions}
        delay_values = {direction: [None, None, None]
                        for direction in directions}

        # Flags to ensure each type of data is added once
        vc_ratio_added = False
        ln_grp_los_added = False
        los_added = False

        """
            Search and process search terms in the current intersection
        """
        for term, row_index in intersection.items():
            if term not in terms_to_match:
                continue  # Skip terms not in the list

            term_lower = term.lower()
            row_data = df.iloc[row_index].replace(
                "", "-").fillna("-").tolist()[1:]  # Exclude first column value

            # Process v/c ratio values
            if "v/c ratio(x)" == term_lower or "v/c ratio" == term_lower:
                for idx, direction in enumerate(directions):
                    start_index = idx * 3  # Adjust based on 3 values per direction
                    if start_index + 2 < len(row_data):
                        vc_ratio_values[direction] = row_data[start_index:start_index + 3]
                    else:
                        vc_ratio_values[direction] = ["-", "-", "-"]
                if not vc_ratio_added:
                    combined_dict[term] = [
                        value for value in row_data if value != '-']
                    vc_ratio_added = True

            # Process LOS values
            if "lngrp los" == term_lower or "los" == term_lower:
                for idx, direction in enumerate(directions):
                    start_index = idx * 3
                    if start_index + 2 < len(row_data):
                        los_values[direction] = row_data[start_index:start_index + 3]
                    else:
                        los_values[direction] = ["-", "-", "-"]
                if "lngrp los" == term_lower and not ln_grp_los_added:
                    combined_dict[term] = [
                        value for value in row_data if value != '-']
                    ln_grp_los_added = True
                elif "los" == term_lower and not los_added and not ln_grp_los_added:
                    combined_dict[term] = [
                        value for value in row_data if value != '-']
                    los_added = True

            # Process delay values
            if "control delay (s/veh)" == term_lower or "lngrp delay(d), s/veh" == term_lower:
                for idx, direction in enumerate(directions):
                    start_index = idx * 3
                    if start_index + 2 < len(row_data):
                        delay_values[direction] = row_data[start_index:start_index + 3]
                    else:
                        delay_values[direction] = ["-", "-", "-"]

            # Process approach delay
            if "approach delay" in term_lower:
                filtered_row_data = [
                    value for value in row_data if value != '-']
                for idx, direction in enumerate(directions):
                    if lane_config.get(direction) == '-':
                        filtered_row_data.insert(idx, '-')
                for idx, direction in enumerate(directions[:len(filtered_row_data)]):
                    approach_data[direction]["Approach Delay"] = filtered_row_data[idx]

            # Process approach LOS
            if term_lower == "approach los":
                filtered_row_data = [
                    value for value in row_data if value != '-']
                for idx, direction in enumerate(directions):
                    if lane_config.get(direction) == '-':
                        filtered_row_data.insert(idx, '-')
                for idx, direction in enumerate(directions[:len(filtered_row_data)]):
                    approach_data[direction]["Approach LOS"] = filtered_row_data[idx]

            # Add other terms to combined_dict
            if term_lower not in ["v/c ratio", "los", "v/c ratio(x)", "lngrp los"]:
                combined_dict[term] = [
                    value for value in row_data if value != '-']

        # Finalize combined_dict
        value_set = [vc_ratio_values, delay_values, los_values]

        """
            Preference algorithm
        """
        # Process each type of value set: vc_ratio, los, and delay
        for values, term in zip(value_set, list(combined_dict.keys())[1:]):
            for direction in directions:
                # Identify non-zero values based on the type of data
                if values == los_values:
                    # For LOS values, consider grades A-F (skipping '-' or None)
                    non_zero_values = [value if value not in [
                        "-", None] else "Z" for value in values[direction]]
                    max_non_zero_value = max(non_zero_values)
                else:
                    # For vc_ratio and delay values, treat them as floats
                    non_zero_values = [float(value) if value not in [
                        "-", "0", None] else float('inf') for value in values[direction]]
                    non_inf_values = [
                        v for v in non_zero_values if v != float('inf')]
                    min_non_zero_value = min(non_zero_values)
                    max_non_zero_value = max(
                        non_inf_values) if non_inf_values else float('-inf')

                # Determine if we need to update based on raw lane configuration
                needs_update = any(
                    value not in [
                        "-", "0", None] and lane_config in ["0", None]
                    for value, lane_config in zip(values[direction], raw_lane_config[direction])
                )

                # Skip updating if all configurations are valid
                if not needs_update:
                    # print(f"Skipping updates for {direction} in {values} as all configurations are valid.")
                    continue

                # Apply the update logic based on the type of data
                for i, (value, lane_config) in enumerate(zip(values[direction], raw_lane_config[direction])):
                    if value in ["-", "0", None] and lane_config not in ["0", None]:
                        if values == los_values:
                            # Higher letter for LOS
                            values[direction][i] = max_non_zero_value
                        else:
                            values[direction][i] = str(max_non_zero_value)

                # Replace the lowest non-zero value with '-'
                if min_non_zero_value != float('inf') and values != los_values:
                    if len(non_inf_values) > 1:
                        min_index = non_zero_values.index(min_non_zero_value)
                        values[direction][min_index] = '-'

            # Filter out '-' values and add to combined_dict
            filtered_values = {direction: [value for value in values[direction]
                                           if value != '-' and value != 'Z'] for direction in directions}
            combined_dict[term] = [
                value for sublist in filtered_values.values() for value in sublist]

        # Merge approach data into combined_dict and append to final lists
        combined_dict.update(approach_data)
        combined_list.append(combined_dict)

        # print(f"\nCombined data for signalized Intersection {intersection_id}: \n{combined_dict}")

    # Dump empty lane configurations
    for direction in ['EB', 'WB', 'NB', 'SB', 'NE', 'NW', 'SE', 'SW']:
        for config in lane_configurations:
            if config.get(direction) == '-':
                del config[direction]


    awsc_overall = parse_awsc_data(df)
    awsc_combined_data = integrate_awsc_data(awsc_overall, [])
    awsc_dir_formatted, _, _ = process_directions(awsc_overall, lane_configurations)

    # print(f"\nCombined data: {awsc_combined_data}")
    # print(f"\nAWSC Directions: {awsc_dir_formatted}")

    twsc_overall, synchro_overall, hcm_overall, _ = parse_overall_data_v2(
        file_path, df)
    twsc_intersections = parse_twsc_approach(df)

    # Extract all unique 'ID' values
    unique_ids = list(set(entry['ID'] for entry in twsc_overall))
    # print(f"\nUnique IDs: {unique_ids}")

    # ✅ Fixed version: Filter and ensure unique entries
    seen_ids = set()
    matching_configs = [
        config for config in all_intersection_configs
        if config.get('ID') in unique_ids and config.get('ID') not in seen_ids and not seen_ids.add(config.get('ID'))
    ]
    # print(f"$$${matching_configs}")
    twsc_parsed_configs, _ = parse_lane_configs(matching_configs, unique_ids)
    # print("****", twsc_parsed_configs)
    twsc_intersection_directions, _, _ = process_directions(
        twsc_overall, twsc_parsed_configs)

    combined_list.extend(twsc_intersections)

    # print(f"\nTWSC Directions: {twsc_intersection_directions}")
    # print(f"\nTWSC Intersections:\n{twsc_intersections}")
    # print(f"\nCombined list: \n{combined_list}")
    # print(f"\ntwsc_overall:\n{twsc_overall}\n")

    seen_signatures = set()
    deduped_list = []
    for item in combined_list:
        identifier = _normalize_metric_value(item.get('Intersection ID'))
        if identifier:
            signature = (identifier, _freeze_nested(item))
            if signature in seen_signatures:
                continue
            seen_signatures.add(signature)
        deduped_list.append(item)
    combined_list = deduped_list

    # Create an empty DataFrame to hold all intersections' data
    final_df = pd.DataFrame()

    general_terms = {
        'v/c': ['V/c ratio', 'V/c ratio(x)', 'LnGrp v/c'],
        'delay': ['Control delay (s/veh)', 'LnGrp Delay(d), s/veh'],
        'los': ['LOS', 'LnGrp LOS']
    }


    # Sort combined_list by Intersection ID for ordered processing
    combined_list_sorted = sorted(
        combined_list, key=lambda x: int(x.get("Intersection ID", "ID")))

    # Combine both lists and sort by the "index" key
    combined_overall_data = sorted(
        synchro_overall + hcm_overall + twsc_overall + awsc_combined_data, key=lambda x: x.get('index', 0))
    overall_idx = 0

    # print(f"\nSorted combined list: \n{combined_list_sorted}")

    # Process each intersection in the sorted list
    for data_dict in combined_list_sorted:

        # Control printing of Intersection ID only once per direction set
        intersection_id = data_dict.get("Intersection ID")
        intersection_id_printed = False

        """ Find the matching lane configuration for this intersection """
        lane_config = next((config for config in lane_configurations if config.get(
            "Intersection ID") == intersection_id), None)

        """ Check if the intersection has TWSC data """
        twsc_summary_result = next(
            (twsc for twsc in twsc_overall if twsc.get("ID") == str(intersection_id)), None)
        twsc_summary_directions = next((twsc_dir for twsc_dir in twsc_intersection_directions if twsc_dir.get(
            "ID") == str(intersection_id)), None)

        """ Check if the intersection has AWSC data """
        awsc_summary_result = next(
            (awsc for awsc in awsc_overall if awsc.get("ID") == str(intersection_id)), None)
        awsc_summary_directions = next(
            (awsc_dir for awsc_dir in awsc_dir_formatted if awsc_dir.get("ID") == str(intersection_id)), None)


        # Prefer AWSC then TWSC summary if available, otherwise use lane_config
        if awsc_summary_result and awsc_summary_directions:
            lane_config = None
        elif twsc_summary_result and twsc_summary_directions:
            lane_config = None
            awsc_summary_result = None
        else:
            awsc_summary_result = None
            twsc_summary_result = None

        # Skip intersections without lane config, TWSC, or AWSC summary
        if not lane_config and not twsc_summary_result and not awsc_summary_result:
            print(f"No lane configuration, TWSC, or AWSC summary found for Intersection ID: {intersection_id}")
            continue

        # Prepare data for the intersection's DataFrame
        intersection_data = []

        # Separate indexing for v/c, LOS, and Delay values
        j = 0

        # Retrieve all entries from combined_overall_data for this intersection
        overall_data = [item for item in combined_overall_data if item['ID'] == str(intersection_id)]

        # Default values for the overall data row (to be updated if data is found)
        overall_los = '-'
        overall_delay = '-'

        """ Processing if lane configuration is available"""
        if lane_config:
            for direction, lanes in lane_config.items():
                if direction == "Intersection ID":
                    continue

                approach_delay = data_dict.get(direction, {}).get("Approach Delay", '-')
                approach_los = data_dict.get(direction, {}).get("Approach LOS", '-')

                if approach_delay is None:
                    approach_delay = '-'
                if approach_los is None:
                    approach_los = '-'

                approach_delay = _normalize_metric_value(approach_delay)
                approach_los = _normalize_metric_value(approach_los)

                direction_rows_added = False
                direction_los_values = []
                direction_delay_values = []

                for i, lane in enumerate(lanes):
                    intersection_id_str = str(intersection_id) if not intersection_id_printed else ''
                    intersection_id_printed = True
                    direction_value = direction if i == 0 else ''

                    vc_value = los_value = delay_value = '-'

                    for term in general_terms['v/c']:
                        matching_key = next((key for key in data_dict if key.lower() == term.lower()), None)
                        if matching_key:
                            vc_value = data_dict[matching_key][j] if j < len(data_dict[matching_key]) else '-'
                            break

                    for term in general_terms['los']:
                        if term in data_dict:
                            los_value = data_dict[term][j] if j < len(data_dict[term]) else '-'
                            break

                    for term in general_terms['delay']:
                        if term in data_dict:
                            delay_value = data_dict[term][j] if j < len(data_dict[term]) else '-'
                            break

                    vc_value = _normalize_metric_value(vc_value)
                    los_value = _normalize_metric_value(los_value)
                    delay_value = _normalize_metric_value(delay_value)

                    if _has_metric_data(vc_value, los_value, delay_value):
                        intersection_data.append([intersection_id_str, direction_value, lane, vc_value, los_value, delay_value])
                        direction_rows_added = True
                        if los_value.lower() not in PLACEHOLDER_METRIC_VALUES:
                            direction_los_values.append(los_value)
                        if delay_value.lower() not in PLACEHOLDER_METRIC_VALUES:
                            direction_delay_values.append(delay_value)

                    j += 1

                if direction_rows_added and _has_metric_data('-', approach_los, approach_delay):
                    intersection_data.append(['', f"{direction} Overall", '', '-', approach_los, approach_delay])

            if overall_data:
                index = overall_idx if overall_idx < len(overall_data) else 0
                overall_entry = overall_data[index]
                overall_los = overall_entry.get('los', '-')
                overall_delay = overall_entry.get('delay', '-')
                overall_idx = (index + 1) % max(len(overall_data), 1)
            else:
                overall_los = '-'
                overall_delay = '-'

            if _has_metric_data('-', overall_los, overall_delay):
                overall_row = ['', 'Overall', '', '-', overall_los, overall_delay]
                if overall_row not in intersection_data:
                    intersection_data.append(overall_row)

        # """ Processing TWSC summary data"""



        elif twsc_summary_result:
            ORDERED_DIRECTIONS = ['EB', 'WB', 'NB', 'SB', 'NE', 'NW', 'SE', 'SW']
            processed_directions = set()
            for base_direction in ORDERED_DIRECTIONS:
                movements = twsc_summary_directions.get(base_direction)
                if not movements:
                    continue

                if base_direction in processed_directions:
                    continue

                base_direction_printed = False

                if isinstance(movements, list):
                    idx = 0
                    for mv in movements:
                        idx += 1
                        dir_val = mv

                        key = f"{base_direction}Ln{idx}"
                        if key not in twsc_summary_result:
                            key = f"{base_direction}{dir_val}"

                        lane_data = twsc_summary_result.get(key, ('-', '-', '-', '-'))
                        vc_value, los_value, delay_value, capacity_value, has_data = _sanitize_lane_data(lane_data)
                        if not has_data:
                            continue
                        id_cell = str(intersection_id) if not intersection_id_printed else ''
                        intersection_data.append([
                            id_cell,
                            base_direction if not base_direction_printed else '',
                            dir_val,
                            vc_value,
                            los_value,
                            delay_value,
                        ])
                        intersection_id_printed = True
                        base_direction_printed = True
                else:
                    dir_val = movements
                    key = f"{base_direction}Ln1"
                    if key not in twsc_summary_result:
                        key = f"{base_direction}{dir_val}"
                    lane_data = twsc_summary_result.get(key, ('-', '-', '-', '-'))
                    vc_value, los_value, delay_value, capacity_value, has_data = _sanitize_lane_data(lane_data)
                    if not has_data:
                        continue
                    id_cell = str(intersection_id) if not intersection_id_printed else ''
                    intersection_data.append([
                        id_cell,
                        base_direction,
                        dir_val,
                        vc_value,
                        los_value,
                        delay_value,
                    ])
                    intersection_id_printed = True

                processed_directions.add(base_direction)

        elif awsc_summary_result:
            ORDERED_DIRECTIONS = ['EB', 'WB', 'NB', 'SB', 'NE', 'NW', 'SE', 'SW']
            for base_direction in ORDERED_DIRECTIONS:
                movements = awsc_summary_directions.get(base_direction)
                if movements is None:
                    continue
                if base_direction == 'ID':
                    continue

                base_direction_printed = False
                lane_keys = sorted(
                    [key for key in awsc_summary_result.keys() if key.startswith(base_direction) and 'Ln' in key]
                )
                intersection_id_str = str(intersection_id) if not intersection_id_printed else ''

                if isinstance(movements, str):
                    if len(lane_keys) == 1:
                        lane_key = lane_keys[0]
                        lane_data = awsc_summary_result.get(lane_key, ('-', '-', '-', '-'))
                        vc_value, los_value, delay_value, capacity_value, has_data = _sanitize_lane_data(lane_data)
                        if not has_data:
                            continue
                        intersection_data.append([intersection_id_str, base_direction, movements, vc_value, los_value, delay_value])
                        intersection_id_printed = True
                    elif len(lane_keys) == 2 and len(movements) == 4:
                        movement_parts = [movements[:2], movements[2:]]
                        for lane_key, movement in zip(lane_keys, movement_parts):
                            lane_data = awsc_summary_result.get(lane_key, ('-', '-', '-', '-'))
                            vc_value, los_value, delay_value, capacity_value, has_data = _sanitize_lane_data(lane_data)
                            if not has_data:
                                continue
                            id_cell = str(intersection_id) if not intersection_id_printed else ''
                            intersection_data.append([id_cell, base_direction, movement, vc_value, los_value, delay_value])
                            intersection_id_printed = True
                elif isinstance(movements, list):
                    for lane_key, movement in zip(lane_keys, movements):
                        base_direction_str = base_direction if not base_direction_printed else ''
                        lane_data = awsc_summary_result.get(lane_key, ('-', '-', '-', '-'))
                        vc_value, los_value, delay_value, capacity_value, has_data = _sanitize_lane_data(lane_data)
                        if not has_data:
                            continue
                        id_cell = str(intersection_id) if not intersection_id_printed else ''
                        intersection_data.append([id_cell, base_direction_str, movement, vc_value, los_value, delay_value])
                        intersection_id_printed = True
                        base_direction_printed = True
                    base_direction_printed = False

        """ Add final data if it exists"""
        intersection_data = _filter_rows_with_metric_data(intersection_data)

        if intersection_data:
            print(f"\nIntersection {intersection_id} data: \n{intersection_data}")

            intersection_data.append([''] * 6)
            intersection_df = pd.DataFrame(intersection_data, columns=['Intersection ID', 'Direction', 'Lane', 'V/c', 'LOS', 'Delay'])
            final_df = pd.concat([final_df, intersection_df], ignore_index=True)

            # Write the final DataFrame to a CSV file
            final_df.to_csv(final_output_path, index=False)

    """
        Output finalized data for debugging and testing

    i = 0
    # Initialize the intersection ID from id_combined_list
    for item in combined_list_sorted:

        # Determine the intersection ID and the data dictionary based on whether the item is a tuple or dictionary

        intersection_id = item.get("Intersection ID")
        data_dict = item

        # Print each term and its data in a readable format, excluding direction data (EB, WB, NB, SB)
        print(f"Intersection {intersection_id}:")
        for term, data in data_dict.items():
            if term not in ['EB', 'WB', 'NB', 'SB', 'NE', 'NW', 'SE', 'SW']:  # Only print non-directional data here
                # Join data if it's a list, otherwise convert it to a string
                if isinstance(data, list):
                    data_str = ", ".join(map(str, data))
                else:
                    data_str = str(data)
                print(f"  {term}: {data_str}")
        print()
        # Print Approach Delay and LOS for each direction (EB, WB, NB, SB)
        for direction in ['EB', 'WB', 'NB', 'SB', 'NE', 'NW', 'SE', 'SW']:
            # Retrieve approach delay and LOS for the current direction
            # print(f"Getting approach data for '{direction}'...")
            # print(data_dict.get(direction, {}))
            approach_delay = data_dict.get(direction, {}).get("Approach Delay", '-')
            approach_los = data_dict.get(direction, {}).get("Approach LOS", '-')
            # Only delete if the direction exists in data_dict and conditions are met
            print(f"  {direction}: Approach Delay = {approach_delay}, Approach LOS = {approach_los}")
        print()
        # Find the matching lane configuration for this intersection ID in group_config_data
        lane_config = next((config for config in lane_configurations if config.get("Intersection ID") == intersection_id), None)
        raw_config = next((raw for raw in raw_lane_configs if raw.get("Intersection ID") == intersection_id), None)

        # Print lane configurations for the current intersection if available
        if lane_config:
            lane_config_str = ", ".join(
                f"{key}: {value}" for key, value in lane_config.items() if key != "Intersection ID" and value != '-'
            )
            print(f"  Lane Configurations: {lane_config_str}")
        else:
            print(f"  No lane configurations found for Intersection ID: {intersection_id}")

        # Print raw direction configurations for the current intersection if available
        if raw_config:
            raw_config_str = ", ".join(
                f"{key}: {value}" for key, value in raw_config.items() if key != "Intersection ID" and value != [None, None, None]
            )
            print(f"  Raw Direction Configurations: {raw_config_str}")
        else:
            print(f"  No raw direction configurations found for Intersection ID: {intersection_id}")

        i+=1
        # Add a blank line for readability between intersections
        print("\n" + "_" * 40 + "\n")

    print(f"Total number of useable datasets found: {len(combined_list_sorted)}")
    print("_" * 40 + "\n")
    return str(final_output_path)
    """

if __name__ == "__main__":
    # Route all printed output to a local log file (overwrite each run)
    class _Tee:
        def __init__(self, *streams):
            self.streams = streams
        def write(self, data):
            for s in self.streams:
                try:
                    s.write(data)
                except Exception:
                    pass
            return len(data)
        def flush(self):
            for s in self.streams:
                try:
                    s.flush()
                except Exception:
                    pass

    try:
        _log_file = open('log.txt', 'w', encoding='utf-8')
        sys.stdout = _Tee(sys.stdout, _log_file)
        sys.stderr = _Tee(sys.stderr, _log_file)
    except Exception as _e:
        # If logging setup fails, continue without file logging
        print(f"Warning: failed to initialize log file: {_e}")


    if '--select' in sys.argv:
        output_dir = None
        try:
            select_index = sys.argv.index('--output-dir')
            if select_index + 1 < len(sys.argv):
                output_dir = sys.argv[select_index + 1]
        except ValueError:
            pass
        select_reports_and_extract(output_dir)
        sys.exit(0)
    
    # Test input/ouput files
    test_report_1 = "test/Test Report 1.txt"
    test_report_2 = "test/Test Report 2.txt"
    test_report_3 = "test/Test Report 3.txt"
    test_report_4 = "test/Test Report 4.txt"
    test_twsc = "test/TEST TWSC.txt"
    test_awsc = "test/TEST AWSC.txt"

    test_report_1_csv = "test-report-1.csv"
    test_report_2_csv = "test-report-2.csv"
    test_report_3_csv = "test-report-3.csv"
    test_report_4_csv = "test-report-4.csv"
    test_twsc_csv = "test-twsc.csv"
    test_awsc_csv = "test-awsc.csv"

    test_report_alpha = "test/Test Report Alpha.txt"
    test_alpha_csv = "test-alpha.csv"
    helper_dir = DEFAULT_TEMP_DIR
    final_output_dir_path = DEFAULT_OUTPUT_DIR
    helper_dir.mkdir(parents=True, exist_ok=True)
    final_output_dir_path.mkdir(parents=True, exist_ok=True)

    # parse_overall_data_v2(file)  # Gets the data for overall

    # Testing with Test Report 1.txt
    print('\n' + "*"*35 + "\n| Results for 'Test Report 1.txt' |\n" + "*"*35 + '\n')
    extract_data_to_csv(test_report_1, helper_dir / test_report_1_csv, output_dir=final_output_dir_path)

    # Testing with Test Report 2.
    print('\n' + "*"*35 + "\n| Results for 'Test Report 2.txt' |\n" + "*"*35 + '\n')
    extract_data_to_csv(test_report_2, helper_dir / test_report_2_csv, output_dir=final_output_dir_path)

    # Testing with Test Report 3.txt
    print('\n' + "*"*35 + "\n| Results for 'Test Report 3.txt' |\n" + "*"*35 + '\n')
    extract_data_to_csv(test_report_3, helper_dir / test_report_3_csv, output_dir=final_output_dir_path)

    # Testing with Test Report 3.txt
    print('\n' + "*"*35 + "\n| Results for 'Test Report 4.txt' |\n" + "*"*35 + '\n')
    extract_data_to_csv(test_report_4, helper_dir / test_report_4_csv, output_dir=final_output_dir_path)

    print('\n' + "*"*35 + "\n| Results for 'TEST TWSC.txt' |\n" + "*"*35 +'\n')
    extract_data_to_csv(test_twsc, helper_dir / test_twsc_csv, output_dir=final_output_dir_path)

    print('\n' + "*"*35 + "\n| Results for 'TEST AWSC.txt' |\n" + "*"*35 + '\n')
    extract_data_to_csv(test_awsc, helper_dir / test_awsc_csv, output_dir=final_output_dir_path)

    print('\n' + "*"*35 + "\n| Results for 'Test Report Alpha.txt' |\n" + "*"*35 + '\n')
    extract_data_to_csv(test_report_alpha, helper_dir / test_alpha_csv, output_dir=final_output_dir_path)

    # lane_groups = separate_characters(movement)
    # print(f"\nLane groups:\n{lane_groups}")
    # write_to_excel(file, movement, delay, vc, los)
