# -*- coding: utf-8 -*-
"""
Created on Fri Sep 27 11:19:28 2024

@author: alex.dering
"""

# main_window.py

import tkinter as tk  # Import the Tkinter module for GUI development.
import tkinter.ttk as ttk  # Import themed widgets from Tkinter for better styling.
from difflib import SequenceMatcher  # Used for comparing sequences and finding similarities.
from tkinter import messagebox, filedialog  # Import specific Tkinter features for message boxes and file dialogs.
import csv  # Module to handle CSV file operations.
import openpyxl as xl  # Used for working with Excel files (.xlsx format).
import os  # OS module for interacting with the operating system (file paths, etc.).
import pyautogui  # Provides functions to control the mouse and keyboard.
import re  # Regular expression module for pattern matching in strings.
import time  # Module for time-related functions.
import json  # JSON module to parse and manipulate JSON data.
from collections import OrderedDict  # Import ordered dictionary to maintain the order of keys.
from shutil import copy  # Used to copy files or directories.
from openpyxl import Workbook
"""
______________________ HELPER FUNCTIONS ______________________
"""
def is_empty(target):
    if target is None:
        return True
    if type(target) == str:
        if target.strip():
            return False
        else:
            return True


def identify_type(record_name):
    if record_name == 'Arrive On Green':
        control_type = 'hcm signalized'

    elif record_name == 'Opposing Approach':
        control_type = 'hcm all way stop'

    elif record_name == 'Int Delay, s/veh':
        control_type = 'hcm two way stop'

    elif record_name == 'Conflicting Circle Lanes':
        control_type = 'hcm roundabout'

    elif record_name == 'Right Turn on Red':
        control_type = 'synchro signalized'

    elif record_name == 'Degree Utilization, x':
        control_type = 'synchro all way stop'

    elif record_name == 'cSH':
        control_type = 'synchro two way stop'

    elif record_name == 'Control Type: Roundabout':
        control_type = 'synchro roundabout'

    else:
        control_type = None

    return control_type


def get_bounds(file):
    pattern = re.compile('([0-9]+):\w*')
    bounds = list()
    intersections = list()
    data = dict()

    with open(file) as f:
        reader = csv.reader(f, delimiter='\t')
        file_data = list(reader)
    for index, line in enumerate(file_data):
        if line:
            record_name = line[0].strip()
            header_match = pattern.match(record_name)
            if header_match:
                bounds.append(index)
                intersection = int(header_match.groups()[0])
                intersections.append(intersection)
    bounds.append(index)

    for index, inter in enumerate(intersections):
        if inter not in data.keys():
            data[inter] = dict()
        data[inter]['bounds'] = bounds[index:index + 2]
        start, end = data[inter]['bounds']
        for line in file_data[start: end]:
            if line:
                record_name = line[0].strip()
                record_type = identify_type(record_name)
                if record_type:
                    data[inter]['type'] = record_type
                    break
                else:
                    data[inter]['type'] = None
    return data


def find_line(data, search, give_index=False):
    for index, line in enumerate(data):
        if line:
            record_name = line[0].strip()
            if record_name == search:
                if give_index:
                    return line, index
                else:
                    return line
    return None


def get_overall(data_list, control_type):
    # returns overall values in form: [delay, LOS]

    if control_type == 'hcm signalized':
        keys = ['HCM 6th Ctrl Delay', 'HCM 6th LOS']

    elif control_type == 'hcm all way stop':
        keys = ['Intersection Delay, s/veh', 'Intersection LOS']

    elif control_type == 'hcm two way stop':
        keys = ['Int Delay, s/veh']

    elif control_type == 'hcm roundabout':
        keys = ['Intersection Delay, s/veh', 'Intersection LOS']

    elif control_type == 'synchro signalized':
        pass

    elif control_type == 'synchro all way stop':
        # assumes hcm 2000
        keys = ['Delay', 'Level of Service']

    elif control_type == 'synchro two way stop':
        # assumes hcm 2000
        keys = ['Average Delay']

    elif control_type == 'synchro roundabout':
        # todo add synchro roundabout support
        return [None, None]

    else:
        return [None, None]

    if control_type == 'synchro signalized':
        for row in data_list:
            if row:
                if 'Intersection Signal Delay: ' in row[0]:
                    delay = row[0][27:].strip()
                    los = row[5][-1]
                    return [delay, los]

    else:
        output = [None, None]
        for index, key in enumerate(keys):
            row = find_line(data_list, key)
            for entry in row[2:]:
                if entry:
                    output[index] = entry
                    break

        return output


def standardize(results_file):
    
    with open(results_file) as f:
        reader = csv.reader(f, delimiter='\t')
        file_content = list(reader)
    database = dict()
    parsed = get_bounds(results_file)
    for intersection in parsed:

        db = parsed[intersection]
        start = min(db['bounds'])
        end = max(db['bounds'])
        subset = file_content[start:end]
        control_type = db['type']
        database[intersection] = OrderedDict()
        database[intersection]['overall'] = dict()
        delay, los = get_overall(subset, control_type)
        database[intersection]['overall']['delay'] = delay
        database[intersection]['overall']['los'] = los

        # initialize storage variables
        header_by_int = OrderedDict()
        secondary_key = OrderedDict()
        second_info = list()
        header_by_int_alt = dict()
        roundabout_lanes = list()

        # declare search parameters
        if control_type == 'hcm signalized':
            header_key = 'Movement'

            lookup = {'V/C Ratio(X)': 'vc_ratio',
                      'LnGrp Delay(d),s/veh': 'ln_delay',
                      'LnGrp LOS': 'ln_los',
                      'Approach Delay, s/veh': 'app_delay',
                      'Approach LOS': 'app_los'}

        elif control_type == 'hcm all way stop':

            header_key = 'Movement'
            secondary_header_key = 'Lane'

            lookup = {'HCM Control Delay': 'app_delay',
                      'HCM LOS': 'app_los'}

            lookup_2 = {'HCM Lane V/C Ratio': 'vc_ratio',
                        'HCM Control Delay': 'ln_delay',
                        'HCM Lane LOS': 'ln_los'}

        elif control_type == 'hcm two way stop':

            header_key = 'Movement'
            secondary_header_key = 'Minor Lane/Major Mvmt'
            lookup = {'HCM Control Delay, s': 'app_delay',
                      'HCM LOS': 'app_los'}

            lookup_2 = {'HCM Lane V/C Ratio': 'vc_ratio',
                        'HCM Control Delay (s)': 'ln_delay',
                        'HCM Lane LOS': 'ln_los'}

        elif control_type == 'hcm roundabout':
            header_key = 'Approach'
            lookup = {'Approach Delay, s/veh': 'app_delay',
                      'Approach LOS': 'app_los'}

            lookup_2 = {'V/C Ratio': 'vc_ratio',
                        'Control Delay, s/veh': 'ln_delay',
                        'LOS': 'ln_los'}

        if control_type == 'synchro signalized':
            header_key = 'Lane Group'
            lookup = {'v/c Ratio': 'vc_ratio',
                      'Control Delay': 'ln_delay',
                      'LOS': 'ln_los',
                      'Approach Delay': 'app_delay',
                      'Approach LOS': 'app_los'}

        elif control_type == 'synchro all way stop':
            header_key = 'Movement'
            lookup = {'Degree Utilization, x': 'vc_ratio',
                      'Control Delay (s)': 'ln_delay',
                      'LnGrp LOS': 'ln_los',
                      'Approach Delay (s)': 'app_delay',
                      'Approach LOS': 'app_los'}

        elif control_type == 'synchro two way stop':
            header_key = 'Movement'
            lookup = {'Volume to Capacity': 'vc_ratio',
                      'Control Delay (s)': 'ln_delay',
                      'Lane LOS': 'ln_los',
                      'Approach Delay (s)': 'app_delay',
                      'Approach LOS': 'app_los'}

        elif control_type == 'synchro roundabout':
            header_key = 'Movement'
            lookup = {'Volume to Capacity': 'vc_ratio',
                      'Control Delay (s)': 'ln_delay',
                      'Lane LOS': 'ln_los',
                      'Approach Delay (s)': 'app_delay',
                      'Approach LOS': 'app_los'}

        # main data collection
        if control_type == 'synchro roundabout':
            pass

        elif control_type == 'hcm roundabout':

            movement_headers = find_line(subset, header_key)
            for index, content in enumerate(movement_headers[2:]):
                index += 2
                if content:
                    header_by_int[index] = content
                    header_by_int_alt[index - 1] = content

            lanes = find_line(subset, 'Entry Lanes')
            for index, lane in enumerate(lanes[2:]):
                index += 2
                if lane:
                    for num in range(int(lane)):
                        roundabout_lanes.append(header_by_int[index])

            configurations = find_line(subset, 'Designated Moves')
            for index, content in enumerate(configurations[2:]):
                index += 2
                if content:
                    direction = roundabout_lanes[0]
                    roundabout_lanes.pop(0)
                    if len(content) == 1:
                        move = content
                    elif len(content) == 2:
                        if 'T' in content:
                            move = 'T'
                        else:
                            move = 'L'
                    else:
                        move = 'T'

                    database[intersection][direction + move] = dict()
                    config = str()
                    if 'L' in content:
                        config += '<'
                    if 'T' in content:
                        config += '1'
                    if 'R' in content:
                        config += '>'
                    database[intersection][direction + move]['config'] = config

                    for lookup_value, data_tag in lookup_2.items():
                        line = find_line(subset, lookup_value)
                        value = line[index]
                        database[intersection][direction + move][data_tag] = value
            # todo revisit for multilane roundabout support

            for lookup_value, data_tag in lookup.items():
                line = find_line(subset, lookup_value)
                for index, item in enumerate(line[2:]):
                    index += 2
                    if item:
                        direction = header_by_int[index]
                        for record, dictionary in database[intersection].items():
                            if record[:2] == direction:
                                dictionary[data_tag] = item

        elif control_type in ['hcm signalized', 'synchro signalized']:
            movement_headers = find_line(subset, header_key)
            for index, content in enumerate(movement_headers[2:]):
                index += 2
                if content:
                    database[intersection][content] = dict()
                    header_by_int[index] = content

            configurations = find_line(subset, 'Lane Configurations')
            for index, content in enumerate(configurations[2:]):
                index += 2
                if content:
                    key = header_by_int[index]
                    database[intersection][key]['config'] = content

            for line in subset:
                if line:
                    record_name = line[0].strip()
                    for lookup_value, data_tag in lookup.items():
                        if record_name == lookup_value:
                            database_field = data_tag
                            for column_num, value in enumerate(line):
                                if column_num > 1 and column_num in header_by_int.keys():
                                    movement = header_by_int[column_num]
                                    if movement in database[intersection]:
                                        database[intersection][movement][database_field] = value
                            # exit loop since each line can only be one record
                            break

        elif control_type in ['hcm all way stop', 'hcm two way stop']:
            movement_headers = find_line(subset, header_key)
            alternate_header_line, second_index = find_line(subset, secondary_header_key, give_index=True)
            for index, content in enumerate(movement_headers[2:]):
                index += 2
                if content:
                    database[intersection][content] = dict()
                    header_by_int[index] = content

            for index, header in enumerate(alternate_header_line[2:]):
                index += 2
                if header:
                    second_info.append((header[:2], index))
                    secondary_key[header] = index

            configurations = find_line(subset, 'Lane Configurations')
            for index, content in enumerate(configurations[2:]):
                index += 2
                if content:
                    key = header_by_int[index]
                    database[intersection][key]['config'] = content

            for movement in database[intersection]:
                if 'config' in database[intersection][movement].keys():
                    config = database[intersection][movement]['config']
                else:
                    continue
                if config != '0':
                    for index, pair in enumerate(second_info):
                        if movement[:2] == pair[0]:
                            header_by_int_alt[pair[1]] = movement
                            second_info.pop(index)
                            break

            for line in subset[:second_index]:
                if line:
                    record_name = line[0].strip()
                    for lookup_value, data_tag in lookup.items():
                        if record_name == lookup_value:
                            database_field = data_tag
                            for column_num, value in enumerate(line):
                                if column_num > 1 and column_num in header_by_int.keys():
                                    movement = header_by_int[column_num]
                                    if movement in database[intersection]:
                                        database[intersection][movement][database_field] = value
                            # exit loop since each line can only be one record
                            break

            for line in subset[second_index:]:
                if line:
                    record_name = line[0].strip()
                    for lookup_value, data_tag in lookup_2.items():
                        if record_name == lookup_value:
                            database_field = data_tag
                            for column_num, value in enumerate(line):
                                if column_num > 1 and column_num in header_by_int_alt.keys():
                                    movement = header_by_int_alt[column_num]
                                    if movement in database[intersection]:
                                        database[intersection][movement][database_field] = value
                            # exit loop since each line can only be one record
                            break
    return database


def order(txt):
    output = str()
    if txt.find('L') != -1:
        output += 'L'
    if txt.find('T') != -1:
        output += 'T'
    if txt.find('R') != -1:
        output += 'R'
    return output


def label(field, config):
    output = str()
    if not field:
        return output
    if len(field) == 2:
        return field
    if field.find('Ln') != -1:
        return None
    direction = field[2]
    if config.find('<') == -1 and config.find('>') == -1:
        if config.find('0') == -1:
            return direction
    if config.find('<') != -1:
        output += 'L'
    for num in range(1, 9):
        if config.find(str(num)) != -1:
            output += direction
    if config.find('>') != -1:
        output += 'R'
    return order(output)


def similar(str1, str2):
    return SequenceMatcher(None, str1, str2).ratio()


def set_default():
    defaults = {'synchro_exe': 'C:\\Program Files (x86)\\Trafficware\\Version10\\Synchro10.exe',
                'synchro_dir': '',
                'model_path': '',
                'rows': 1000,
                'columns': 30,
                'update_los': 1}

    with open('settings.json', 'w') as file:
        json.dump(defaults, file)


def load_settings():
    try:
        with open('settings.json', 'r') as file:
            defaults = json.load(file)

    except FileNotFoundError:
        set_default()
        with open('settings.json', 'r') as file:
            defaults = json.load(file)
    return defaults


def center_window(x, y, master):
    screen_width, screen_height = master.winfo_screenwidth(), master.winfo_screenheight()
    x_coord, y_coord = int((screen_width - x) / 2), int((screen_height - y) / 2)
    if x == 0 and y == 0:
        size = str()
    else:
        size = f'{x}x{y}'
    position = f'+{x_coord}+{y_coord}'
    return size + position


def replace_slash(string):
    return string.replace('/', '\\')


def get_row(worksheet, intersection):
    for row in range(1, worksheet.max_row + 1):
        cell_value = worksheet.cell(row, 1).value
        if cell_value is None:
            return row, 'direct'
        elif cell_value == intersection:
            return row, 'direct'
        elif cell_value > intersection:
            method = 'insert'
            return row, method
        elif cell_value < intersection:
            for i in range(row, worksheet.max_row + 1):
                if worksheet.cell(i, 1).value > intersection:
                    return i, 'insert'
                elif i == worksheet.max_row:
                    return i, 'append'


def get_sheet(wb, name):
    for sheet in wb.sheetnames:
        if sheet == name:
            return wb[sheet]
        wb.create_sheet(title=name)
    return wb[name]


def give_notice(master):
    notice = tk.Toplevel()
    notice.geometry(center_window(600, 200, master))
    notice.wm_attributes('-topmost', 1)
    text = ttk.Label(notice)
    text.config(text='Synchronizer is beginning control of Synchro.'
                     '\nPlease do not touch the mouse until the program has finished.',
                justify='center',
                font=('open sans', 15))
    text.pack()
    notice.after(5000, notice.destroy)

"""
______________________ CLASSES ______________________
"""

class Scenario:
    def __init__(self, name):
        self.name = name
        self.hour = None
        self.year = None
        self.condition = None
        self.syn_file = None
        self.volumes = None
        self.los_data = None
        self.los_results = None
        self.model_data_column = None

    def parse(self):
        for hour in ['AM', 'PM', 'SAT']:
            if self.name.find(hour) != -1:
                self.name.replace(hour, '')
                self.hour = hour
                break


class Settings:
    def __init__(self, master=None):
        self.master = master
        defaults = load_settings()
        # build ui
        self.settings_window = tk.Toplevel(master)
        self.main_frame = ttk.Frame(self.settings_window)
        self.notebook_1 = ttk.Notebook(self.main_frame)

        self.model_outer = ttk.Frame(self.notebook_1)
        self.model_outer.columnconfigure(0, weight=1)

        self.search_bounds = ttk.Labelframe(self.model_outer)
        self.search_bounds.columnconfigure(1, weight=1)

        self.row_label = ttk.Label(self.search_bounds)
        self.row_label.config(text='Rows:')
        self.row_label.grid(sticky='w')
        self.row_entry = ttk.Entry(self.search_bounds)
        _text_ = defaults['rows']
        self.row_entry.delete('0', 'end')
        self.row_entry.insert('0', _text_)
        self.row_entry.grid(column='1', row='0', sticky='nsew', padx=10)

        self.col_label = ttk.Label(self.search_bounds)
        self.col_label.config(text='Columns:')
        self.col_label.grid(sticky='w')
        self.col_entry = ttk.Entry(self.search_bounds)
        _text_ = defaults['columns']
        self.col_entry.delete('0', 'end')
        self.col_entry.insert('0', _text_)
        self.col_entry.grid(column='1', row='1', sticky='nsew', padx=10)

        self.search_bounds.config(height='200', text='Boundaries', width='200')
        self.search_bounds.rowconfigure(0, weight=1)
        self.search_bounds.grid(padx='0', pady='0', sticky='we')

        self.model_path_frame = ttk.Labelframe(self.model_outer)
        self.model_path_label = ttk.Label(self.model_path_frame)
        self.model_path_label.config(text='Default Path:')
        self.model_path_label.grid(sticky='w')
        self.model_path_frame.columnconfigure(1, weight=1)

        self.model_path_entry = ttk.Entry(self.model_path_frame)
        self.model_path_entry.config(text='Default Path:')
        _text_ = defaults['model_path']
        self.model_path_entry.delete('0', 'end')
        self.model_path_entry.insert('0', _text_)
        self.model_path_entry.grid(column='1', row='0', sticky='nsew', padx=10)

        self.model_browse = ttk.Button(self.model_path_frame)
        self.model_browse.config(text='Browse', command=self.model_browse_func)
        self.model_browse.grid(column='2', row='0')

        self.model_path_frame.config(height='200', text='Default Model Path', width='200')
        self.model_path_frame.grid(column='0', row='1', sticky='nsew')

        self.model_outer.config(height='200', width='200')
        self.model_outer.grid()
        self.notebook_1.add(self.model_outer, text='Model')

        self.syn_frame = ttk.Labelframe(self.notebook_1)
        self.syn_frame.columnconfigure(1, weight=1)

        self.syn_app_label = ttk.Label(self.syn_frame)
        self.syn_app_label.config(text='Synchro app location:')
        self.syn_app_label.grid()

        self.syn_app_entry = ttk.Entry(self.syn_frame)
        self.syn_app_entry.config(cursor='arrow')
        _text_ = defaults['synchro_exe']
        self.syn_app_entry.delete('0', 'end')
        self.syn_app_entry.insert('0', _text_)
        self.syn_app_entry.grid(column='1', row='0', sticky='nsew', padx=10)

        self.syn_dir_label = ttk.Label(self.syn_frame)
        self.syn_dir_label.config(text='Default Synchro folder:')
        self.syn_dir_label.grid(column='0', row='1')

        self.syn_dir_entry = ttk.Entry(self.syn_frame)
        self.syn_dir_entry.config(cursor='arrow')
        _text_ = defaults['synchro_dir']
        self.syn_dir_entry.delete('0', 'end')
        self.syn_dir_entry.insert('0', _text_)
        self.syn_dir_entry.grid(column='1', row='1', sticky='nsew', padx=10)

        self.syn_browse = ttk.Button(self.syn_frame)
        self.syn_browse.config(text='Browse', command=self.syn_browse_func)
        self.syn_browse.grid(column='2', row='0')

        self.syn_dir_browse = ttk.Button(self.syn_frame)
        self.syn_dir_browse.config(text='Browse', command=self.syn_dir_browse_func)
        self.syn_dir_browse.grid(column='2', row='1')

        self.syn_frame.config(height='200', text='Synchro')
        self.syn_frame.grid()
        self.notebook_1.add(self.syn_frame, text='Synchro')

        self.gen_tab_frame = ttk.Labelframe(self.notebook_1)
        self.gen_tab_frame.config(height='200', text='General', width='200')
        self.gen_tab_frame.pack(side='top')

        self.gen_label = ttk.Label(self.gen_tab_frame)
        self.gen_label.config(text='Update LOS by Default:')
        self.gen_label.grid()

        self.update_los_yes = ttk.Radiobutton(self.gen_tab_frame, text='Yes')
        self.update_los_yes.config(variable=self.master.update_los, value=1)
        self.update_los_yes.grid(row=0, column=1)

        self.update_los_no = ttk.Radiobutton(self.gen_tab_frame, variable=self.master.update_los, text='No')
        self.update_los_no.config(variable=self.master.update_los, value=0)
        self.update_los_no.grid(row=0, column=2)

        self.notebook_1.add(self.gen_tab_frame, text='General')
        self.notebook_1.config(height='200', width='200')
        self.notebook_1.pack(fill='both', side='top')
        self.button_frame = ttk.Frame(self.main_frame)
        self.button_frame.columnconfigure((0, 1), weight=1)

        self.done_button = ttk.Button(self.button_frame)
        self.done_button.config(text='Done', command=self.done)
        self.done_button.grid(sticky='e')

        self.cancel_button = ttk.Button(self.button_frame)
        self.cancel_button.config(text='Cancel', command=self.settings_window.destroy)
        self.cancel_button.grid(column='1', row='0', sticky='w')

        self.button_frame.config()
        self.button_frame.pack(fill='both', side='top')

        self.main_frame.config(height='200', width='200')
        self.main_frame.pack(fill='both', side='top')

        self.settings_window.config(height='200', width='200')
        self.settings_window.geometry('480x320')
        self.settings_window.title('Settings')

        # Main widget
        self.mainwindow = self.settings_window

    def model_browse_func(self):
        path = filedialog.askopenfile()
        if path:
            path = replace_slash(path.name)
            self.model_path_entry.delete('0', 'end')
            self.model_path_entry.insert('0', path)

    def syn_browse_func(self):
        path = filedialog.askopenfile()
        if path:
            path = replace_slash(path.name)
            self.syn_app_entry.delete('0', 'end')
            self.syn_app_entry.insert('0', path)

    def syn_dir_browse_func(self):
        path = replace_slash(filedialog.askdirectory())
        if path:
            self.syn_dir_entry.delete('0', 'end')
            self.syn_dir_entry.insert('0', path)

    def done(self):
        rows = int(self.row_entry.get())
        columns = int(self.col_entry.get())
        model_path = replace_slash(self.model_path_entry.get())
        syn_exe = replace_slash(self.syn_app_entry.get())
        update_los = self.master.update_los
        syn_dir = replace_slash(self.syn_dir_entry.get())
        defaults = {'synchro_exe': syn_exe,
                    'synchro_dir': syn_dir,
                    'model_path': model_path,
                    'rows': rows,
                    'columns': columns,
                    'update_los': update_los}

        self.master.synchro_app_path = defaults['synchro_exe']
        self.master.synchro_dir = defaults['synchro_dir']
        self.master.model_path = defaults['model_path']
        self.master.default_rows = defaults['rows']
        self.master.default_columns = defaults['columns']
        self.master.update_los = defaults['update_los']

        self.master.main_win.model_entry.delete('0', 'end')
        self.master.main_win.model_entry.insert('0', self.master.model_path)
        self.master.main_win.syn_entry.delete('0', 'end')
        self.master.main_win.syn_entry.insert('0', self.master.synchro_dir)

        with open('settings.json', 'w') as file:
            json.dump(defaults, file)

        self.settings_window.destroy()


class MainWindow:
    def __init__(self, master=None):
        self.master = master
        self.data = None

        # UI Setup
        self.setup_ui()

    def setup_ui(self):
        self.master.columnconfigure(0, weight=1)
        self.master.rowconfigure(0, weight=1)

        self.frame_1 = ttk.Frame(self.master)
        self.frame_1.columnconfigure(1, weight=1)
        self.frame_1.rowconfigure((0, 1, 2), weight=1)

        # Create label and entry for Model file location
        self.model_entry = self.create_label_entry("Model file location:", self.master.model_path, 0, self.model_browse_func)

        # Create label and entry for Synchro file folder
        self.syn_entry = self.create_label_entry("Synchro file folder:", self.master.synchro_dir, 1, self.syn_browse_func)

        self.los_button = ttk.Checkbutton(self.frame_1, variable=self.master.update_los, text='Update LOS Table')
        self.los_button.grid(column=0, row=2)

        self.create_button("Start", self.launch_file_match, 2, 2, sticky='e')
        self.create_button("Settings", self.launch_settings, 0, 3, sticky='w')

        self.utilities = ttk.Labelframe(self.frame_1, text='Other Functions')
        self.utilities.grid(row=3, column=1)

        self.create_button("Copy Files", self.copy, 0, 0, parent=self.utilities, side='left')
        self.create_button("LOS Only", None, 0, 1, parent=self.utilities, side='left')  # Placeholder for LOS Only button

        notes = '''Instructions:\n
                1. Please ensure the Synchro files you wish to update are not open on any computer.\n
                2. Check that the model file is in our standard format.'''
        self.note_label = ttk.Label(self.frame_1, text=notes)
        self.note_label.grid(row=4, columnspan=3)

        self.frame_1.grid(sticky='nsew')

        # Main widget
        self.mainwindow = self.frame_1

    def create_label_entry(self, label_text, default_value, row, browse_command):
        label = ttk.Label(self.frame_1, text=label_text)
        label.grid(column=0, row=row)

        entry = ttk.Entry(self.frame_1)
        entry.insert(0, default_value)
        entry.grid(column=1, row=row, sticky='nsew')

        browse_button = ttk.Button(self.frame_1, text='Browse', command=browse_command)
        browse_button.grid(row=row, column=2)

        return entry  # Return the entry widget to assign it to an instance variable

    def create_button(self, text, command, col, row, parent=None, sticky=None, side=None):
        button = ttk.Button(parent if parent else self.frame_1, text=text, command=command)
        if side:
            button.pack(side=side)
        else:
            button.grid(column=col, row=row, sticky=sticky)

    def model_browse_func(self):
        path = filedialog.askopenfile(filetypes=[('Excel Files', '*.xlsx')])
        if path:
            self.update_entry(path.name, self.model_entry)

    def syn_browse_func(self):
        path = filedialog.askdirectory()
        if path:
            self.update_entry(path, self.syn_entry)

    def update_entry(self, path, entry):
        path = replace_slash(path)
        entry.delete(0, 'end')
        entry.insert(0, path)

    def launch_settings(self):
        Settings(self.master)

    def launch_file_match(self):
        self.master.synchro_dir = self.syn_entry.get()
        self.master.storage_dir = os.path.join(self.master.synchro_dir, 'temp')
        os.makedirs(self.master.storage_dir, exist_ok=True)
        self.master.model_path = replace_slash(self.model_entry.get())
        self.master.find_volume_data(self.master.model_path)
        FileMatchApp(self.master)

    def copy(self):
        Copier(self.master)

    def run(self):
        self.mainwindow.mainloop()
        

class Base(tk.Tk):
    DEFAULT_ROWS = 1000
    DEFAULT_COLUMNS = 30
    VALID_SCENARIOS = ['EXISTING', 'NO BUILD', 'BUILD']
    SCENARIO_CONDITIONS = {'EXISTING': ['EX'],
                           'NO BUILD': ['NB'],
                           'BUILD': ['B', 'BD'],
                           'IMPROVEMENT': ['IMP']}
    
    def __init__(self):
        super().__init__()
        self.title('Synchronizer')
        self.geometry(center_window(500, 200, self))
        self.resizable(width=0, height=0)
        self.wm_attributes('-topmost', 0)
        
        self.windows = {}
        self.main_win = None
        self.storage_dir = None
        self.model_sheet_name = str()
        self.model_data = {}
        self.scenario_list = []
        self.scenario_data = {}
        self.selected_scenarios = []
        self.scenarios = []  # hold Scenario objects
        self.ws = None
        self.data_columns = []
        
        # Load settings
        defaults = {
            'synchro_exe': 'C:\\Program Files (x86)\\Trafficware\\Version10\\Synchro10.exe',
            'synchro_dir': '',
            'model_path': '',
            'rows': self.DEFAULT_ROWS,
            'columns': self.DEFAULT_COLUMNS,
            'update_los': 1
        }
        saved_settings = load_settings()

        self.synchro_app_path = saved_settings.get('synchro_exe', defaults['synchro_exe'])
        self.synchro_dir = saved_settings.get('synchro_dir', defaults['synchro_dir'])
        self.model_path = saved_settings.get('model_path', defaults['model_path'])
        self.default_rows = saved_settings.get('rows', defaults['rows'])
        self.default_columns = saved_settings.get('columns', defaults['columns'])
        self.update_los = saved_settings.get('update_los', defaults['update_los'])

    def match_ws_name(self, workbook_path, title):
        wb = xl.load_workbook(filename=workbook_path, data_only=True)
        match = max(wb.sheetnames, key=lambda sheet: similar(sheet, title), default=None)
        return wb[match] if match else None

    def find_volume_data(self, extra_scenario=None):
        valid_scenarios = [extra_scenario] if extra_scenario else self.VALID_SCENARIOS
        output = {}
        
        wb = xl.load_workbook(filename=self.model_path, data_only=True)
        self.model_sheet_name = max(wb.sheetnames, key=lambda sheet: similar(sheet, 'Model'), default=None)
        self.ws = wb[self.model_sheet_name]

        for row in range(1, self.ws.max_row):
            if self.ws.cell(row, 1).value == 1:
                year, scenario = None, None
                for column in range(1, self.ws.max_column):
                    year_cell = self.ws.cell(row - 4, column).value
                    scenario_cell = self.ws.cell(row - 3, column).value
                    hour_cell = self.ws.cell(row - 2, column).value
                    
                    if year_cell is not None:
                        year = str(year_cell)
                    if scenario_cell is not None:
                        scenario = str(scenario_cell)
                    if hour_cell in ['AM', 'PM', 'SAT'] and scenario in valid_scenarios:
                        name = f"{year} {scenario} {hour_cell}"
                        if not any(found_scenario.name == name for found_scenario in self.scenarios):
                            sc = Scenario(name)
                            sc.hour = hour_cell
                            sc.year = year
                            sc.condition = scenario
                            sc.model_data_column = column
                            self.match_syn_file(sc, self.synchro_dir)
                            self.scenarios.append(sc)
                        else:
                            messagebox.showwarning('Duplicate', 'One or more scenarios were duplicated and not added.')

        self.scenario_data = output
        return output.keys()

    def match_syn_file(self, scenario, dir):
        key = self.SCENARIO_CONDITIONS.get(scenario.condition, [scenario.condition])
        match = max(
            (file for file in os.scandir(dir) if file.path.endswith('.syn')),
            key=lambda file: max(similar(file.name, scenario.hour + acronym) for acronym in key),
            default=None
        )
        scenario.syn_file = match.path if match else str()

    # keyboard
    # Convert model volumes to Synchro UTDF
    def convert_utdf(self, scenario='test_write', column=5):
        # Open model to copy data
        # wb = xl.load_workbook(filename=model)
        # active = wb.active
        ws = self.ws  # need to make sure sheet is titled "Model"
        startColumn = 'C'  # get direction column from user or default
        dataColumns = ['F', 'G', 'H']  # from scenarios to update

        volume_data = dict()
        movement_list = ['RECORDNAME', 'INTID']

        for row in range(15, self.default_rows):
            intersection = None
            direction = None

            cell = ws.cell(row, 1).value
            if type(cell) in [int, float]:
                intersection = int(cell)
                volume_data[intersection] = dict()
            cell = ws.cell(row, 3).value
            if type(cell) == str:
                direction = cell
                volume_data[intersection][direction] = dict()
            turn = ws.cell(row, 4).value
            if intersection and direction and turn:
                volume = ws.cell(row, column).value
                if volume is None:
                    volume = 0
                else:
                    volume = int(volume)
                volume_data[intersection][direction + turn] = volume
                if direction + turn not in movement_list:
                    movement_list.append(direction + turn)

        # Dictionary Format:
        # {intersection:{direction1:{L:0, T:0, R:0
        #                                        },
        #                            direction2:{},
        #                            direction3:{}}}

        # print(volume_data)
        file = self.storage_dir + '\\' + scenario + '.csv'
        with open(file, 'w', newline='') as f:
            f.write('[Lanes]\nLane Group Data\n')
            writer = csv.DictWriter(f, fieldnames=movement_list)
            writer.writeheader()
            for intid in volume_data:
                payload = volume_data[intid]
                payload['RECORDNAME'] = 'Volume'
                payload['INTID'] = intid
                writer.writerow(payload)
        return file

    def click_button(self, image, x_off=0, y_off=0):
        self.manage_error()
        for attempt in range(2):  # Try to find button again if not found
            time.sleep(2)
            result = pyautogui.locateCenterOnScreen(image, confidence=0.9)
            if result is not None:
                x_coord, y_coord = result
                x_coord += x_off
                y_coord += y_off
                pyautogui.click(x_coord, y_coord)
                print(image, x_coord, y_coord)
                return x_coord, y_coord
        print(image, None)
        return None

    def startup(self):
        start = os.system('start "" "' + self.synchro_app_path + '"')
        if start == 0:
            self.click_button('License.png', 0, 75)
            self.click_button('Update.png')
            self.click_button('Maximize.png')
        return start

    def manage_error(self):
        windows = [{'name': 'Error Symbol.png',
                    'x': 288,
                    'y': 88},
                   {'name': 'Unexpected Error.png',
                    'x': 240,
                    'y': -78},
                   {'name': 'Activity Log.png',
                    'x': 224,
                    'y': -119},
                   {'name': 'Read Only.png',
                    'x': -42,
                    'y': 30}
                   ]
        for window in windows:
            name = window['name']
            x = window['x']
            y = window['y']
            result = pyautogui.locateCenterOnScreen(name, confidence=0.9)
            if result is not None:
                pyautogui.click(x, y)
                break

    # update volumes in synchro
    def import_to_synchro(self, syn_file, data_file):
        # Open synchro file
        print(data_file)
        self.click_button('Open File.png')
        self.click_button('Open Window.png', 185, 498)
        pyautogui.write(syn_file)
        self.click_button('Open.png')

        # Click transfer tab
        self.click_button('Transfer Tab.png')

        # Click Merge File
        self.click_button('Merge File.png')

        # Select input volumes file
        # self.click_button('Merge File Open Logo.png', 306, 501)
        pyautogui.press('delete')
        pyautogui.write(data_file)
        self.click_button('Merge File Open Logo.png', 652, 537)
        self.click_button('Confirm Merge.png', -42, 29)
        self.click_button('Save File.png')
        # return True
        # pyautogui.getWindowsWithTitle("Photos")[0].maximize()
        # time.sleep(2)
        # win32api.SetCursorPos((x, y))

    # retrieve LOS data from synchro
    def export_from_synchro(self, scenario):
        file = self.storage_dir + '\\' + scenario + '.txt'
        # Click report button
        self.click_button('Reports.png')

        # Click save as text
        self.click_button('Save Text.png')

        # self.click_button('Open Window.png', 121, 462)

        # Write file name
        pyautogui.press('delete')
        pyautogui.write(file)

        # Save text file
        self.click_button('Save.png')
        self.click_button('Confirm Save As.png', 190, 93)

        return file

    def update_report(self, scenarios, report_table=None):
        if report_table is None:
            report_table = 'synchronizer results.xlsx'
        report_table = self.storage_dir + '\\' + report_table
        wb = xl.Workbook()
        ws = wb.active
        ws.title = 'AM'

        for scenario in scenarios:
            data = scenario.los_data
            hour = scenario.hour
            sheet = get_sheet(wb, hour)
            condition = scenario.condition
            if condition == 'EXISTING':
                column = 5
            elif condition == 'NO-BUILD':
                column = 8
            elif condition == 'BUILD':
                column = 11
            else:
                column = sheet.max_column

            for intersection in data:
                row, method = get_row(sheet, intersection)
                ov_los = None
                ov_delay = None
                for turn_move, values in data[intersection].items():
                    if turn_move == 'overall':
                        ov_delay = values['delay']
                        ov_los = values['los']
                        continue

                    movement_name = label(turn_move, values.get('config', ''))
                    if movement_name:
                        vc_ratios = list()
                        los_values = list()
                        delays = list()
                        app_los_values = list()
                        app_delays = list()
                        last_move = turn_move[:2]

                        for direction in movement_name:
                            search = turn_move[:2] + direction
                            if search not in data[intersection].keys():
                                continue
                            vc_ratios.append(data[intersection][search].get('vc_ratio', ''))
                            los_values.append(data[intersection][search].get('ln_los', ''))
                            delays.append(data[intersection][search].get('ln_delay', ''))
                            app_los_values.append(data[intersection][search].get('app_los', ''))
                            app_delays.append(data[intersection][search].get('app_delay', ''))

                        vc = max(vc_ratios)
                        los = max(los_values)
                        delay = max(delays)
                        app_los = max(app_los_values)
                        app_delay = max(app_delays)

                        if vc == '' and los == '' and delay == '':
                            continue

                        if method == 'direct':
                            sheet.cell(row, 1).value = intersection
                            sheet.cell(row, 3).value = turn_move[:2]
                            sheet.cell(row, 4).value = movement_name
                            sheet.cell(row, column).value = vc
                            sheet.cell(row, column + 1).value = los
                            sheet.cell(row, column + 2).value = delay
                            row += 1

                        elif method == 'insert':
                            sheet.insert_rows(row)
                            sheet.cell(row, 1).value = intersection
                            sheet.cell(row, 3).value = turn_move[:2]
                            sheet.cell(row, 4).value = movement_name
                            sheet.cell(row, column).value = vc
                            sheet.cell(row, column + 1).value = los
                            sheet.cell(row, column + 2).value = delay
                            row += 1

                        elif method == 'append':
                            row += 1
                            sheet.cell(row, 1).value = intersection
                            sheet.cell(row, 3).value = turn_move[:2]
                            sheet.cell(row, 4).value = movement_name
                            sheet.cell(row, column).value = vc
                            sheet.cell(row, column + 1).value = los
                            sheet.cell(row, column + 2).value = delay

                        if app_delay == '' and app_los == '':
                            continue

                        if last_move and turn_move != last_move:
                            if method == 'direct':
                                sheet.cell(row, 1).value = intersection
                                sheet.cell(row, 3).value = 'Approach'
                                sheet.cell(row, column + 1).value = app_los
                                sheet.cell(row, column + 2).value = app_delay
                                row += 1

                            elif method == 'insert':
                                sheet.insert_rows(row)
                                sheet.cell(row, 1).value = intersection
                                sheet.cell(row, 3).value = 'Approach'
                                sheet.cell(row, column + 1).value = app_los
                                sheet.cell(row, column + 2).value = app_delay
                                row += 1

                            elif method == 'append':
                                row += 1
                                sheet.cell(row, 1).value = intersection
                                sheet.cell(row, 3).value = 'Approach'
                                sheet.cell(row, column + 1).value = app_los
                                sheet.cell(row, column + 2).value = app_delay

                if ov_los and ov_delay:
                    if method == 'direct':
                        sheet.cell(row, 1).value = intersection
                        sheet.cell(row, 3).value = 'Overall'
                        sheet.cell(row, column + 1).value = ov_los
                        sheet.cell(row, column + 2).value = ov_delay
                        row += 1

                    elif method == 'insert':
                        sheet.insert_rows(row)
                        sheet.cell(row, 1).value = intersection
                        sheet.cell(row, 3).value = 'Overall'
                        sheet.cell(row, column + 1).value = ov_los
                        sheet.cell(row, column + 2).value = ov_delay
                        row += 1

                    elif method == 'append':
                        row += 1
                        sheet.cell(row, 1).value = intersection
                        sheet.cell(row, 3).value = 'Overall'
                        sheet.cell(row, column + 1).value = ov_los
                        sheet.cell(row, column + 2).value = ov_delay

        wb.save(report_table)
        return report_table


class ProgressWindow:
    def __init__(self, master=None):
        self.master = master
        # build ui
        self.progress_window = tk.Toplevel(self.master)
        self.progress_window.geometry(center_window(400, 400, self.master))
        self.progress_window.columnconfigure(0, weight=1)
        self.progress_frame = ttk.Frame(self.progress_window)
        self.progress_frame.columnconfigure(0, weight=1)
        self.progress_frame.columnconfigure(1, weight=0)
        self.status_text_box = tk.Text(self.progress_frame)
        self.status_text_box.config(autoseparators='false')
        self.status_text_box.grid(column='0', row='0', sticky='nsew')
        self.scrollbar_3 = ttk.Scrollbar(self.progress_frame, command=self.status_text_box.yview)
        self.scrollbar_3.grid(column='1', row='0', sticky='nsew')
        self.status_text_box.configure(yscrollcommand=self.scrollbar_3.set)
        # self.button_3 = ttk.Button(self.progress_frame)
        # self.button_3.config(text='button_3')
        # self.button_3.grid(column='0', row='1', sticky='s')
        # self.progress_frame.config(height='200', width='200')
        self.progress_frame.grid(padx=10, pady=10, sticky='nsew')
        # self.progress_window.config(height='200', width='200')
        self.progress_window.title('Program Status')
        give_notice(self.master)
        self.progress_window.after(6000, self.run)

    def run(self):
        time.sleep(2)
        success = self.master.startup()
        if success != 0:
            self.status_text_box.insert('end', 'Failed to start Synchro\n')
            # return
        for scenario_obj in self.master.selected_scenarios:
            scenario = scenario_obj.name
            filename = scenario_obj.syn_file
            column = scenario_obj.model_data_column
            process_update = 'Processing: ' + scenario + '\n'
            self.status_text_box.insert('end', process_update)
            utdf_volumes = self.master.convert_utdf(scenario=scenario, column=column)
            self.status_text_box.insert('end', 'Importing volumes to Synchro... \n')
            self.master.import_to_synchro(filename, utdf_volumes)
            self.status_text_box.insert('end', 'Import complete\n')

            if self.master.update_los:
                self.status_text_box.insert('end', 'Exporting LOS data from Synchro...\n')
                scenario_obj.los_results = self.master.export_from_synchro(scenario)
                time.sleep(5)
                self.status_text_box.insert('end', 'Export complete\n')
                scenario_obj.los_data = standardize(scenario_obj.los_results)

        if self.master.update_los:
            self.status_text_box.insert('end', 'Writing LOS data to excel file\n')
            output_results = self.master.update_report(self.master.selected_scenarios)
            self.status_text_box.insert('end', 'Write complete\nThe program has finished\n')
            self.status_text_box.insert('end', f'LOS results are saved at: {output_results}')


class Copier:
    def __init__(self, master=None):
        self.window = tk.Toplevel(master)
        self.window.columnconfigure(1, weight=1)

        self.old_dir_label = ttk.Label(self.window, text='Copy from:')
        self.old_dir_label.grid(row=0, column=0, sticky='e', padx=10)

        self.old_dir_entry = ttk.Entry(self.window)
        self.old_dir_entry.grid(row=0, column=1, sticky='ew')

        self.old_dir_button = ttk.Button(self.window, text='Browse', command=self.browse)
        self.old_dir_button.bind('<Button 1>', self.browse)
        self.old_dir_button.grid(row=0, column=2)

        self.new_dir_label = ttk.Label(self.window, text='Copy to:')
        self.new_dir_label.grid(row=1, column=0, sticky='e', padx=10)

        self.new_dir_entry = ttk.Entry(self.window)
        self.new_dir_entry.grid(row=1, column=1, sticky='ew')

        self.new_dir_button = ttk.Button(self.window, text='Browse', command=self.browse)
        self.new_dir_button.bind('<Button 1>', self.browse)
        self.new_dir_button.grid(row=1, column=2)

        self.new_date_label = ttk.Label(self.window, text='New date:')
        self.new_date_label.grid(row=2, column=0, sticky='e', padx=10)

        self.new_date_entry = ttk.Entry(self.window)
        self.new_date_entry.grid(row=2, column=1, sticky='ew')

        self.start = ttk.Button(self.window, text='Start', command=self.copy_files)
        self.start.grid(row=3, columnspan=3)

        # self.check_syn = ttk.Checkbutton(self.window, text='Synchro')
        # self.check_syn.grid(row=3)
        #
        # self.check_pdf = ttk.Checkbutton(self.window, text='Synchro')
        # self.check_pdf.grid(row=4)

    def browse(self, event):
        file = filedialog.askdirectory()
        if file is None:
            return
        row = event.widget.grid_info()['row']
        if row == 0:
            entry = self.old_dir_entry
        else:
            entry = self.new_dir_entry

        entry.delete(0, 'end')
        entry.insert('end', file)

    def copy_files(self):
        pattern = re.compile('[0-9]*')
        old_dir = self.old_dir_entry.get()

        for file in os.scandir(old_dir):
            print(file)
            if not file.name.endswith('syn'):
                continue
            new_date = self.new_date_entry.get()
            old_date = re.match(pattern, file.name).group(0)
            if old_date == '':
                new_file_name = new_date + file.name
            else:
                new_file_name = file.name.replace(old_date, new_date)
            new_path = self.new_dir_entry.get() + '\\' + new_file_name
            copy(file.path, new_path)
        self.window.destroy()


class FileMatchApp:
    def __init__(self, master=None):
        self.master = master
        self.file_window = tk.Toplevel(self.master)
        self.file_window.geometry(center_window(500, 400, self.master))  # 500 400
        self.file_window.minsize(width=500, height=100)
        self.file_window.wm_attributes('-topmost', 1)
        self.file_window.wm_attributes('-topmost', 0)

        self.file_window.columnconfigure(0, weight=1)
        # self.file_window.columnconfigure(1, weight=0)
        # self.file_window.rowconfigure(0, weight=1)
        self.entry_data = list()
        self.entry_dict = dict()
        # self.base = Base()
        # self.base.find_columns()
        self.file_dict = dict()
        scenarios = self.master.scenario_data.keys()
        # build ui
        self.frame_1 = ttk.Labelframe(self.file_window)
        self.frame_1.config(text='Add Scenario: ')
        self.frame_1.columnconfigure(0, weight=1)
        self.frame_1.grid(sticky='nsew', padx=10, pady=10)

        self.frame_2 = ttk.Frame(self.file_window)
        self.frame_2.columnconfigure((0, 1, 2, 4), weight=0)
        self.frame_2.columnconfigure(3, weight=1)
        self.frame_2.grid(sticky='nsew')

        self.frame_3 = ttk.Frame(self.file_window)
        self.frame_3.grid()

        self.search_bar = ttk.Entry(self.frame_1)
        self.search_bar.grid(row=0, column=0, sticky='ew')

        search_button = ttk.Button(self.frame_1)
        search_button.config(text='Search')
        search_button.config(command=self.add_scenario)
        search_button.grid(row=0, column=1)

        done = ttk.Button(self.frame_3)
        done.config(text='Done')
        done.config(command=self.decode)
        done.grid()

        clear = ttk.Button(self.frame_3)
        clear.config(text='Clear Blanks')
        clear.config(command=self.clear)
        clear.grid(row=0, column=1)

        # Main widget
        # self.mainwindow = self.frame_2

        for scenario in master.scenarios:
            name = scenario.name
            file = scenario.syn_file
            self.add_row(name, file, scenario)

    def get_path(self, event):
        x, y = event.widget.winfo_pointerxy()
        row = event.widget.winfo_containing(x, y).grid_info()['row']
        path = filedialog.askopenfilename(filetypes=[('Synchro Files', '*.syn')])
        if path:
            path = replace_slash(path)
        entry = self.frame_2.grid_slaves(row=row, column=3)[0]
        entry.delete('0', 'end')
        entry.insert('0', path)
        entry.after_idle(entry.xview_moveto, 1)
        return path

    def add_scenario(self):
        scenario = self.search_bar.get()
        results = self.master.find_volume_data(extra_scenario=scenario)
        if len(results) == 0:
            messagebox.showwarning('Scenario Not Found', 'The entered scenario was not found.')
        for result in results:
            self.add_row(result)

    def add_row(self, name='', file='', obj=None):
        row = self.frame_2.grid_size()[1]

        new_plus_button = ttk.Button(self.frame_2)
        new_plus_button.config(text='+')
        new_plus_button.config(command=self.add_row)
        new_plus_button.grid(row=str(row), sticky='w')

        new_minus_button = ttk.Button(self.frame_2)
        new_minus_button.config(text='-')
        new_minus_button.bind('<Button-1>', self.delete_row)
        new_minus_button.grid(column=1, row=str(row), sticky='w')

        scenario_entry = ttk.Entry(self.frame_2)
        scenario_entry.delete('0', 'end')
        scenario_entry.insert('0', name)
        scenario_entry.after(50, scenario_entry.xview_moveto, 1)
        scenario_entry.grid(column='2', row=str(row), sticky='w')

        file_path_entry = ttk.Entry(self.frame_2)
        file_path_entry.delete('0', 'end')
        file_path_entry.insert('0', file)
        file_path_entry.after(500, file_path_entry.xview_moveto, 1)
        file_path_entry.grid(column='3', row=str(row), sticky='ew')

        browse = ttk.Button(self.frame_2)
        browse.config(text='Browse')
        browse.bind('<Button-1>', self.get_path)
        browse.grid(column='4', row=str(row), sticky='e')

        self.entry_dict[scenario_entry] = dict()
        self.entry_dict[scenario_entry]['file'] = file_path_entry
        self.entry_dict[scenario_entry]['obj'] = obj

    def delete_row(self, event):

        plus_buttons_left = 0
        for widget in self.frame_2.winfo_children():
            if widget.cget('text') == '+':
                plus_buttons_left += 1

        if plus_buttons_left > 1:
            x, y = event.widget.winfo_pointerxy()
            row = event.widget.winfo_containing(x, y).grid_info()['row']
            for widget in self.frame_2.winfo_children():
                if widget.grid_info()['row'] == row:
                    self.entry_dict.pop(widget, None)
                    widget.destroy()

    def clear(self):
        removed_widgets = list()
        for sc_widget in self.entry_dict.keys():
            row = sc_widget.grid_info()['row']
            sc_content = sc_widget.get()
            file_content = self.entry_dict[sc_widget]['file'].get()
            if is_empty(sc_content) or is_empty(file_content):
                for widget in self.frame_2.winfo_children():
                    if widget.grid_info()['row'] == row:
                        removed_widgets.append(widget)
                        widget.destroy()
        for item in removed_widgets:
            self.entry_dict.pop(item, None)

    def decode(self):
        for sc_widget in self.entry_dict.keys():
            sc_content = sc_widget.get()
            file_content = self.entry_dict[sc_widget]['file'].get()
            if is_empty(sc_content) and is_empty(file_content):
                continue
            elif is_empty(sc_content) or is_empty(file_content):
                messagebox.showwarning('Empty Input',
                                       'A row is missing a scenario name or file path. Please add the data or delete '
                                       'the row')
                return
            else:
                obj = self.entry_dict[sc_widget]['obj']
                obj.name = sc_content
                obj.syn_file = replace_slash(file_content)
                self.master.selected_scenarios.append(obj)

        self.file_window.destroy()
        give_notice(self.master)
        ProgressWindow(self.master)

    def old_decode(self):
        for widget in self.frame_2.winfo_children():
            if type(widget) == ttk.Entry:
                contents = widget.get()
                # Check to make sure entry box is not empty
                if contents is not None and contents != '':
                    # Scenarios stored in column 2, file names in column 3
                    if widget.grid_info()['column'] == 2:
                        key = 'scenario'
                    else:
                        key = 'filename'
                    row = widget.grid_info()['row']
                    if row not in self.file_dict.keys():
                        self.file_dict[row] = dict()
                    self.master.file_dict[row][key] = widget.get()

        self.file_window.destroy()
        ProgressWindow(self.master)









