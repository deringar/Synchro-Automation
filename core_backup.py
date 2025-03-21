# -*- coding: utf-8 -*-
"""
Created on Fri Sep 27 2024
Last modified on Thurs Oct 3 2024

@authors: philip.gotthelf, alex.dering - Colliers Engineering & Design
"""

# main_window.py

import tkinter as tk  # Import the Tkinter module for GUI development.
# Import themed widgets from Tkinter for better styling.
import tkinter.ttk as ttk
# Used for comparing sequences and finding similarities.
from difflib import SequenceMatcher
# Import specific Tkinter features for message boxes and file dialogs.
from tkinter import messagebox, filedialog
import csv  # Module to handle CSV file operations.
import openpyxl as xl  # Used for working with Excel files (.xlsx format).
# OS module for interacting with the operating system (file paths, etc.).
import os
import re  # Regular expression module for pattern matching in strings.
import time  # Module for time-related functions.
import json  # JSON module to parse and manipulate JSON data.
# Import ordered dictionary to maintain the order of keys.
from collections import OrderedDict
from shutil import copy  # Used to copy files or directories.
from openpyxl import load_workbook, Workbook
import pandas as pd

"""
____________________________ AD _____________________________

write_headers(ws, 'C')
write_headers(ws, 'F')
write_headers(ws, 'I')
"""





"""
______________________ HELPER FUNCTIONS ______________________
"""

# Check if the target is empty


def is_empty(target):
    # Check if the target is None
    if target is None:
        return True
    # If the target is a string, check if it is empty or contains only whitespace
    if type(target) == str:
        if target.strip():
            return False  # The string is not empty
        else:
            return True  # The string is empty
    return False  # The target is not empty (not None or empty string)


"""
STEP 2
"""

# Identify the type of control based on the record name


def identify_type(record_name):
    # Map record names to control types
    if record_name == 'Arrive On Green':
        control_type = 'hcm signalized'
    elif record_name == 'Opposing Approach':
        control_type = 'hcm all way stop'
    elif record_name == 'Int Delay, s/veh':
        control_type = 'hcm two way stop'
    elif record_name == 'Conflicting Circle Lanes':
        control_type = 'hcm roundabout'
    elif record_name == 'Right Turn on Red':
        control_type = 'synchro signalized'
    elif record_name == 'Degree Utilization, x':
        control_type = 'synchro all way stop'
    elif record_name == 'cSH':
        control_type = 'synchro two way stop'
    elif record_name == 'Control Type: Roundabout':
        control_type = 'synchro roundabout'
    else:
        control_type = None  # Unrecognized record name

    return control_type

# Get bounds of intersections in the file


def get_bounds(file):
    # Regex pattern to match intersection records
    pattern = re.compile('([0-9]+):\w*')
    bounds = list()  # To store the bounds of intersections
    intersections = list()  # To store intersection IDs
    data = dict()  # To store intersection data

    # Read the file content
    with open(file) as f:
        reader = csv.reader(f, delimiter='\t')
        file_data = list(reader)

    # Iterate through the file data to find intersection bounds
    for index, line in enumerate(file_data):
        if line:  # Skip empty lines
            record_name = line[0].strip()  # Get the first element of the line
            # Match the record name with the pattern
            header_match = pattern.match(record_name)
            if header_match:  # If there's a match, it's an intersection record
                bounds.append(index)  # Store the index of the bound
                # Get the intersection ID
                intersection = int(header_match.groups()[0])
                intersections.append(intersection)  # Store the intersection ID

    bounds.append(index)  # Append the last index for bounds

    # Process the intersections to gather data
    for index, inter in enumerate(intersections):
        if inter not in data.keys():
            # Initialize a dictionary for each intersection
            data[inter] = dict()
        # Set bounds for the intersection
        data[inter]['bounds'] = bounds[index:index + 2]
        start, end = data[inter]['bounds']

        # Loop through the lines within the bounds
        for line in file_data[start:end]:
            if line:  # Skip empty lines
                record_name = line[0].strip()  # Get the record name
                # Identify the control type
                record_type = identify_type(record_name)
                if record_type:  # If a control type is found
                    data[inter]['type'] = record_type  # Set the control type
                    break  # Exit loop once the type is found
                else:
                    data[inter]['type'] = None  # No control type found
    return data  # Return the constructed data dictionary

# Find a line in data matching the search term


def find_line(data, search, give_index=False):
    # Loop through each line of data
    for index, line in enumerate(data):
        if line:  # Skip empty lines
            record_name = line[0].strip()  # Get the first element of the line
            if record_name == search:  # Check if it matches the search term
                if give_index:
                    return line, index  # Return line and index if requested
                else:
                    return line  # Return the line only
    return None  # Return None if no match is found

# Get overall values like delay and LOS based on control type


def get_overall(data_list, control_type):
    # returns overall values in the form: [delay, LOS]

    # Determine the keys based on the control type
    if control_type == 'hcm signalized':
        keys = ['HCM 6th Ctrl Delay', 'HCM 6th LOS']
    elif control_type == 'hcm all way stop':
        keys = ['Intersection Delay, s/veh', 'Intersection LOS']
    elif control_type == 'hcm two way stop':
        keys = ['Int Delay, s/veh']
    elif control_type == 'hcm roundabout':
        keys = ['Intersection Delay, s/veh', 'Intersection LOS']
    elif control_type == 'synchro signalized':
        pass  # To be implemented for synchro signalized
    elif control_type == 'synchro all way stop':
        keys = ['Delay', 'Level of Service']  # Assumes HCM 2000
    elif control_type == 'synchro two way stop':
        keys = ['Average Delay']  # Assumes HCM 2000
    elif control_type == 'synchro roundabout':
        return [None, None]  # To be implemented for synchro roundabouts
    else:
        return [None, None]  # Return None for unrecognized control types

    # Handle data extraction for 'synchro signalized' control type
    if control_type == 'synchro signalized':
        for row in data_list:
            if row:
                # Look for specific record
                if 'Intersection Signal Delay: ' in row[0]:
                    delay = row[0][27:].strip()  # Extract delay
                    los = row[5][-1]  # Extract level of service
                    return [delay, los]  # Return extracted values

    # If not 'synchro signalized', extract data using keys
    output = [None, None]
    for index, key in enumerate(keys):
        row = find_line(data_list, key)  # Find the row for each key
        if row is None:
            print(
                f"Warning: Key '{key}' not found in data_list for control type '{control_type}'.")
            continue  # Skip this key if not found
        for entry in row[2:]:  # Skip the first two columns
            if entry:  # Get the first non-empty entry
                output[index] = entry
                break

    return output  # Return the overall values

# Standardize the results from the file


def standardize(results_file):
    # Read the content of the results file
    with open(results_file) as f:
        reader = csv.reader(f, delimiter='\t')
        file_content = list(reader)  # Store the file content as a list
    database = dict()  # To store the standardized data
    parsed = get_bounds(results_file)  # Get intersection bounds and types

    # Iterate through parsed intersections to build the database
    for intersection in parsed:
        db = parsed[intersection]  # Get data for the intersection
        start = min(db['bounds'])  # Get the starting index for bounds
        end = max(db['bounds'])  # Get the ending index for bounds
        subset = file_content[start:end]  # Get the relevant data subset
        control_type = db['type']  # Get the control type
        # Initialize an ordered dictionary for intersection
        database[intersection] = OrderedDict()
        # Initialize overall data dictionary
        database[intersection]['overall'] = dict()
        delay, los = get_overall(subset, control_type)  # Get delay and LOS
        database[intersection]['overall']['delay'] = delay  # Store delay
        database[intersection]['overall']['los'] = los  # Store LOS

        # Initialize storage variables for detailed data
        header_by_int = OrderedDict()  # Movement headers by intersection
        secondary_key = OrderedDict()  # Secondary keys for alternate headers
        second_info = list()  # List to store additional information
        header_by_int_alt = dict()  # Alternate movement headers
        roundabout_lanes = list()  # To store roundabout lane information

        # Declare search parameters based on control type
        if control_type == 'hcm signalized':
            header_key = 'Movement'

            lookup = {'V/C Ratio(X)': 'vc_ratio',
                      'LnGrp Delay(d),s/veh': 'ln_delay',
                      'LnGrp LOS': 'ln_los',
                      'Approach Delay, s/veh': 'app_delay',
                      'Approach LOS': 'app_los'}

        elif control_type == 'hcm all way stop':

            header_key = 'Movement'
            secondary_header_key = 'Lane'

            lookup = {'HCM Control Delay': 'app_delay',
                      'HCM LOS': 'app_los'}

            lookup_2 = {'HCM Lane V/C Ratio': 'vc_ratio',
                        'HCM Control Delay': 'ln_delay',
                        'HCM Lane LOS': 'ln_los'}

        elif control_type == 'hcm two way stop':

            header_key = 'Movement'
            secondary_header_key = 'Minor Lane/Major Mvmt'
            lookup = {'HCM Control Delay, s': 'app_delay',
                      'HCM LOS': 'app_los'}

            lookup_2 = {'HCM Lane V/C Ratio': 'vc_ratio',
                        'HCM Control Delay (s)': 'ln_delay',
                        'HCM Lane LOS': 'ln_los'}

        elif control_type == 'hcm roundabout':
            header_key = 'Approach'
            lookup = {'Approach Delay, s/veh': 'app_delay',
                      'Approach LOS': 'app_los'}

            lookup_2 = {'V/C Ratio': 'vc_ratio',
                        'Control Delay, s/veh': 'ln_delay',
                        'LOS': 'ln_los'}

        if control_type == 'synchro signalized':
            header_key = 'Lane Group'
            lookup = {'v/c Ratio': 'vc_ratio',
                      'Control Delay': 'ln_delay',
                      'LOS': 'ln_los',
                      'Approach Delay': 'app_delay',
                      'Approach LOS': 'app_los'}

        elif control_type == 'synchro all way stop':
            header_key = 'Movement'
            lookup = {'Degree Utilization, x': 'vc_ratio',
                      'Control Delay (s)': 'ln_delay',
                      'LnGrp LOS': 'ln_los',
                      'Approach Delay (s)': 'app_delay',
                      'Approach LOS': 'app_los'}

        elif control_type == 'synchro two way stop':
            header_key = 'Movement'
            lookup = {'Volume to Capacity': 'vc_ratio',
                      'Control Delay (s)': 'ln_delay',
                      'Lane LOS': 'ln_los',
                      'Approach Delay (s)': 'app_delay',
                      'Approach LOS': 'app_los'}

        elif control_type == 'synchro roundabout':
            header_key = 'Movement'
            lookup = {'Volume to Capacity': 'vc_ratio',
                      'Control Delay (s)': 'ln_delay',
                      'Lane LOS': 'ln_los',
                      'Approach Delay (s)': 'app_delay',
                      'Approach LOS': 'app_los'}

        # main data collection
        if control_type == 'synchro roundabout':
            pass

        elif control_type == 'hcm roundabout':

            movement_headers = find_line(subset, header_key)
            for index, content in enumerate(movement_headers[2:]):
                index += 2
                if content:
                    header_by_int[index] = content
                    header_by_int_alt[index - 1] = content

            lanes = find_line(subset, 'Entry Lanes')
            for index, lane in enumerate(lanes[2:]):
                index += 2
                if lane:
                    for num in range(int(lane)):
                        roundabout_lanes.append(header_by_int[index])

            configurations = find_line(subset, 'Designated Moves')
            for index, content in enumerate(configurations[2:]):
                index += 2
                if content:
                    direction = roundabout_lanes[0]
                    roundabout_lanes.pop(0)
                    if len(content) == 1:
                        move = content
                    elif len(content) == 2:
                        if 'T' in content:
                            move = 'T'
                        else:
                            move = 'L'
                    else:
                        move = 'T'

                    database[intersection][direction + move] = dict()
                    config = str()
                    if 'L' in content:
                        config += '<'
                    if 'T' in content:
                        config += '1'
                    if 'R' in content:
                        config += '>'
                    database[intersection][direction + move]['config'] = config

                    for lookup_value, data_tag in lookup_2.items():
                        line = find_line(subset, lookup_value)
                        value = line[index]
                        database[intersection][direction +
                                               move][data_tag] = value
            # todo revisit for multilane roundabout support

            for lookup_value, data_tag in lookup.items():
                line = find_line(subset, lookup_value)
                for index, item in enumerate(line[2:]):
                    index += 2
                    if item:
                        direction = header_by_int[index]
                        for record, dictionary in database[intersection].items():
                            if record[:2] == direction:
                                dictionary[data_tag] = item

        elif control_type in ['hcm signalized', 'synchro signalized']:
            movement_headers = find_line(subset, header_key)
            for index, content in enumerate(movement_headers[2:]):
                index += 2
                if content:
                    database[intersection][content] = dict()
                    header_by_int[index] = content

            configurations = find_line(subset, 'Lane Configurations')
            for index, content in enumerate(configurations[2:]):
                index += 2
                if content:
                    key = header_by_int[index]
                    database[intersection][key]['config'] = content

            for line in subset:
                if line:
                    record_name = line[0].strip()
                    for lookup_value, data_tag in lookup.items():
                        if record_name == lookup_value:
                            database_field = data_tag
                            for column_num, value in enumerate(line):
                                if column_num > 1 and column_num in header_by_int.keys():
                                    movement = header_by_int[column_num]
                                    if movement in database[intersection]:
                                        database[intersection][movement][database_field] = value
                            # exit loop since each line can only be one record
                            break

        elif control_type in ['hcm all way stop', 'hcm two way stop']:
            movement_headers = find_line(subset, header_key)
            alternate_header_line, second_index = find_line(
                subset, secondary_header_key, give_index=True)
            for index, content in enumerate(movement_headers[2:]):
                index += 2
                if content:
                    database[intersection][content] = dict()
                    header_by_int[index] = content

            for index, header in enumerate(alternate_header_line[2:]):
                index += 2
                if header:
                    second_info.append((header[:2], index))
                    secondary_key[header] = index

            configurations = find_line(subset, 'Lane Configurations')
            for index, content in enumerate(configurations[2:]):
                index += 2
                if content:
                    key = header_by_int[index]
                    database[intersection][key]['config'] = content

            for movement in database[intersection]:
                if 'config' in database[intersection][movement].keys():
                    config = database[intersection][movement]['config']
                else:
                    continue
                if config != '0':
                    for index, pair in enumerate(second_info):
                        if movement[:2] == pair[0]:
                            header_by_int_alt[pair[1]] = movement
                            second_info.pop(index)
                            break

            for line in subset[:second_index]:
                if line:
                    record_name = line[0].strip()
                    for lookup_value, data_tag in lookup.items():
                        if record_name == lookup_value:
                            database_field = data_tag
                            for column_num, value in enumerate(line):
                                if column_num > 1 and column_num in header_by_int.keys():
                                    movement = header_by_int[column_num]
                                    if movement in database[intersection]:
                                        database[intersection][movement][database_field] = value
                            # exit loop since each line can only be one record
                            break

            for line in subset[second_index:]:
                if line:
                    record_name = line[0].strip()
                    for lookup_value, data_tag in lookup_2.items():
                        if record_name == lookup_value:
                            database_field = data_tag
                            for column_num, value in enumerate(line):
                                if column_num > 1 and column_num in header_by_int_alt.keys():
                                    movement = header_by_int_alt[column_num]
                                    if movement in database[intersection]:
                                        database[intersection][movement][database_field] = value
                            # exit loop since each line can only be one record
                            break

    df = pd.DataFrame(database)
    output = "test.csv"
    df.to_csv(output, index=False)

    print(database)
    return database


def order(txt):
    """Returns a string that contains 'L', 'T', and 'R' based on their presence in the input text."""
    output = str()
    if txt.find('L') != -1:
        output += 'L'
    if txt.find('T') != -1:
        output += 'T'
    if txt.find('R') != -1:
        output += 'R'
    return output  # Return the constructed output string


def label(field, config):
    """Generates a label based on the field and configuration rules, returning a direction or None."""
    output = str()
    if not field:
        return output
    if len(field) == 2:
        return field
    if field.find('Ln') != -1:
        return None
    direction = field[2]  # Extract the direction from the field
    # If no special characters are found in config, return the direction if '0' is not present
    if config.find('<') == -1 and config.find('>') == -1:
        if config.find('0') == -1:
            return direction

    # Check if '<' is in the config, if so, add 'L' to output
    if config.find('<') != -1:
        output += 'L'

    # Loop through numbers 1 to 8 and check if they are in the config
    for num in range(1, 9):
        if config.find(str(num)) != -1:
            output += direction  # Append direction for each found number

    # Check if '>' is in the config, if so, add 'R' to output
    if config.find('>') != -1:
        output += 'R'

    return order(output)  # Call order to finalize the output


def similar(str1, str2):
    """Returns the similarity ratio between two strings using SequenceMatcher."""
    return SequenceMatcher(None, str1, str2).ratio()


def load_settings():
    """Loads settings from a JSON file, creating defaults if the file does not exist."""

    def set_default():
        """Sets default settings and saves them to a JSON file."""
        defaults = {
            'synchro_exe': 'C:\\Program Files (x86)\\Trafficware\\Version10\\Synchro10.exe',
            'synchro_dir': '',
            'model_path': '',
            'rows': 1000,
            'columns': 30,
            'update_los': 1
        }

        # Write the default settings to a JSON file
        with open('settings.json', 'w') as file:
            json.dump(defaults, file)

    try:
        with open('settings.json', 'r') as file:
            defaults = json.load(file)

    except FileNotFoundError:
        # If the file doesn't exist, create default settings and load them
        set_default()
        with open('settings.json', 'r') as file:
            defaults = json.load(file)

    return defaults  # Return the loaded or default settings


def center_window(x, y, master):
    """Calculates the position to center a window of size (x, y) on the screen."""
    screen_width, screen_height = master.winfo_screenwidth(), master.winfo_screenheight()
    x_coord, y_coord = int((screen_width - x) /
                           2), int((screen_height - y) / 2)

    # Prepare the size string for window geometry
    if x == 0 and y == 0:
        size = str()
    else:
        size = f'{x}x{y}'

    # Create the final position string for the window
    position = f'+{x_coord}+{y_coord}'
    return size + position  # Return the geometry string


def replace_slash(string):
    """Replaces forward slashes with backslashes in a given string."""
    return string.replace('/', '\\')


def get_row(worksheet, intersection):
    """Finds the appropriate row in the worksheet for a given intersection value."""
    for row in range(1, worksheet.max_row + 1):
        # Get the value in the first column of the row
        cell_value = worksheet.cell(row, 1).value

        # If the cell is empty, return the row and method 'direct'
        if cell_value is None:
            return row, 'direct'
        # If the cell value matches the intersection, return the row and method 'direct'
        elif cell_value == intersection:
            return row, 'direct'
        # If the cell value is greater than the intersection, return the row and method 'insert'
        elif cell_value > intersection:
            method = 'insert'
            return row, method
        # If the cell value is less than the intersection, continue searching
        elif cell_value < intersection:
            for i in range(row, worksheet.max_row + 1):
                # If a subsequent cell value is greater than the intersection, return the row and method 'insert'
                if worksheet.cell(i, 1).value > intersection:
                    return i, 'insert'
                # If we reach the last row without finding a greater value, return the last row and method 'append'
                elif i == worksheet.max_row:
                    return i, 'append'


def get_sheet(wb, name):
    """Retrieves a sheet by name from the workbook, creating it if it does not exist."""
    for sheet in wb.sheetnames:
        if sheet == name:
            return wb[sheet]  # Return the existing sheet
        # If not found, create a new sheet with the specified name
        wb.create_sheet(title=name)

    return wb[name]  # Return the newly created sheet


"""
______________________ CLASSES ______________________
"""
# Stores details about a traffic scenario such as its name, hour, year, condition, and various data files.


class Scenario:
    def __init__(self, name):
        self.name = name
        self.hour = None
        self.year = None
        self.condition = None
        self.syn_file = None
        self.volumes = None
        self.los_data = None
        self.los_results = None
        self.model_data_column = None

    # Processes the scenario name to extract the hour (e.g., AM, PM, SAT).
    def parse(self):
        for hour in ['AM', 'PM', 'SAT']:
            if self.name.find(hour) != -1:
                self.name.replace(hour, '')
                self.hour = hour
                break

# Manages the settings for the application, including loading and saving settings to a JSON file.
# Builds a user interface to allow the user to configure settings such as default paths and boundaries.


class Settings:
    def __init__(self, master=None):
        self.master = master
        defaults = load_settings()
        # build ui
        self.settings_window = tk.Toplevel(master)
        self.main_frame = ttk.Frame(self.settings_window)
        self.notebook_1 = ttk.Notebook(self.main_frame)

        self.model_outer = ttk.Frame(self.notebook_1)
        self.model_outer.columnconfigure(0, weight=1)

        self.search_bounds = ttk.Labelframe(self.model_outer)
        self.search_bounds.columnconfigure(1, weight=1)

        self.row_label = ttk.Label(self.search_bounds)
        self.row_label.config(text='Rows:')
        self.row_label.grid(sticky='w')
        self.row_entry = ttk.Entry(self.search_bounds)
        _text_ = defaults['rows']
        self.row_entry.delete('0', 'end')
        self.row_entry.insert('0', _text_)
        self.row_entry.grid(column='1', row='0', sticky='nsew', padx=10)

        self.col_label = ttk.Label(self.search_bounds)
        self.col_label.config(text='Columns:')
        self.col_label.grid(sticky='w')
        self.col_entry = ttk.Entry(self.search_bounds)
        _text_ = defaults['columns']
        self.col_entry.delete('0', 'end')
        self.col_entry.insert('0', _text_)
        self.col_entry.grid(column='1', row='1', sticky='nsew', padx=10)

        self.search_bounds.config(height='200', text='Boundaries', width='200')
        self.search_bounds.rowconfigure(0, weight=1)
        self.search_bounds.grid(padx='0', pady='0', sticky='we')

        self.model_path_frame = ttk.Labelframe(self.model_outer)
        self.model_path_label = ttk.Label(self.model_path_frame)
        self.model_path_label.config(text='Default Path:')
        self.model_path_label.grid(sticky='w')
        self.model_path_frame.columnconfigure(1, weight=1)

        self.model_path_entry = ttk.Entry(self.model_path_frame)
        self.model_path_entry.config(text='Default Path:')
        _text_ = defaults['model_path']
        self.model_path_entry.delete('0', 'end')
        self.model_path_entry.insert('0', _text_)
        self.model_path_entry.grid(column='1', row='0', sticky='nsew', padx=10)

        self.model_browse = ttk.Button(self.model_path_frame)
        self.model_browse.config(text='Browse', command=self.model_browse_func)
        self.model_browse.grid(column='2', row='0')

        self.model_path_frame.config(
            height='200', text='Default Model Path', width='200')
        self.model_path_frame.grid(column='0', row='1', sticky='nsew')

        self.model_outer.config(height='200', width='200')
        self.model_outer.grid()
        self.notebook_1.add(self.model_outer, text='Model')

        self.syn_frame = ttk.Labelframe(self.notebook_1)
        self.syn_frame.columnconfigure(1, weight=1)

        self.syn_app_label = ttk.Label(self.syn_frame)
        self.syn_app_label.config(text='Synchro app location:')
        self.syn_app_label.grid()

        self.syn_app_entry = ttk.Entry(self.syn_frame)
        self.syn_app_entry.config(cursor='arrow')
        _text_ = defaults['synchro_exe']
        self.syn_app_entry.delete('0', 'end')
        self.syn_app_entry.insert('0', _text_)
        self.syn_app_entry.grid(column='1', row='0', sticky='nsew', padx=10)

        self.syn_dir_label = ttk.Label(self.syn_frame)
        self.syn_dir_label.config(text='Default Synchro folder:')
        self.syn_dir_label.grid(column='0', row='1')

        self.syn_dir_entry = ttk.Entry(self.syn_frame)
        self.syn_dir_entry.config(cursor='arrow')
        _text_ = defaults['synchro_dir']
        self.syn_dir_entry.delete('0', 'end')
        self.syn_dir_entry.insert('0', _text_)
        self.syn_dir_entry.grid(column='1', row='1', sticky='nsew', padx=10)

        self.syn_browse = ttk.Button(self.syn_frame)
        self.syn_browse.config(text='Browse', command=self.syn_browse_func)
        self.syn_browse.grid(column='2', row='0')

        self.syn_dir_browse = ttk.Button(self.syn_frame)
        self.syn_dir_browse.config(
            text='Browse', command=self.syn_dir_browse_func)
        self.syn_dir_browse.grid(column='2', row='1')

        self.syn_frame.config(height='200', text='Synchro')
        self.syn_frame.grid()
        self.notebook_1.add(self.syn_frame, text='Synchro')

        self.gen_tab_frame = ttk.Labelframe(self.notebook_1)
        self.gen_tab_frame.config(height='200', text='General', width='200')
        self.gen_tab_frame.pack(side='top')

        self.gen_label = ttk.Label(self.gen_tab_frame)
        self.gen_label.config(text='Update LOS by Default:')
        self.gen_label.grid()

        self.update_los_yes = ttk.Radiobutton(self.gen_tab_frame, text='Yes')
        self.update_los_yes.config(variable=self.master.update_los, value=1)
        self.update_los_yes.grid(row=0, column=1)

        self.update_los_no = ttk.Radiobutton(
            self.gen_tab_frame, variable=self.master.update_los, text='No')
        self.update_los_no.config(variable=self.master.update_los, value=0)
        self.update_los_no.grid(row=0, column=2)

        self.notebook_1.add(self.gen_tab_frame, text='General')
        self.notebook_1.config(height='200', width='200')
        self.notebook_1.pack(fill='both', side='top')
        self.button_frame = ttk.Frame(self.main_frame)
        self.button_frame.columnconfigure((0, 1), weight=1)

        self.done_button = ttk.Button(self.button_frame)
        self.done_button.config(text='Done', command=self.done)
        self.done_button.grid(sticky='e')

        self.cancel_button = ttk.Button(self.button_frame)
        self.cancel_button.config(
            text='Cancel', command=self.settings_window.destroy)
        self.cancel_button.grid(column='1', row='0', sticky='w')

        self.button_frame.config()
        self.button_frame.pack(fill='both', side='top')

        self.main_frame.config(height='200', width='200')
        self.main_frame.pack(fill='both', side='top')

        self.settings_window.config(height='200', width='200')
        self.settings_window.geometry('480x320')
        self.settings_window.title('Settings')

        # Main widget
        self.mainwindow = self.settings_window

    def model_browse_func(self):
        path = filedialog.askopenfile()
        if path:
            path = replace_slash(path.name)
            self.model_path_entry.delete('0', 'end')
            self.model_path_entry.insert('0', path)

    def syn_browse_func(self):
        path = filedialog.askopenfile()
        if path:
            path = replace_slash(path.name)
            self.syn_app_entry.delete('0', 'end')
            self.syn_app_entry.insert('0', path)

    def syn_dir_browse_func(self):
        path = replace_slash(filedialog.askdirectory())
        if path:
            self.syn_dir_entry.delete('0', 'end')
            self.syn_dir_entry.insert('0', path)

    def done(self):
        rows = int(self.row_entry.get())
        columns = int(self.col_entry.get())
        model_path = replace_slash(self.model_path_entry.get())
        syn_exe = replace_slash(self.syn_app_entry.get())
        update_los = self.master.update_los
        syn_dir = replace_slash(self.syn_dir_entry.get())
        defaults = {'synchro_exe': syn_exe,
                    'synchro_dir': syn_dir,
                    'model_path': model_path,
                    'rows': rows,
                    'columns': columns,
                    'update_los': update_los}

        self.master.synchro_app_path = defaults['synchro_exe']
        self.master.synchro_dir = defaults['synchro_dir']
        self.master.model_path = defaults['model_path']
        self.master.default_rows = defaults['rows']
        self.master.default_columns = defaults['columns']
        self.master.update_los = defaults['update_los']

        self.master.main_win.model_entry.delete('0', 'end')
        self.master.main_win.model_entry.insert('0', self.master.model_path)
        self.master.main_win.syn_entry.delete('0', 'end')
        self.master.main_win.syn_entry.insert('0', self.master.synchro_dir)

        with open('settings.json', 'w') as file:
            json.dump(defaults, file)

        self.settings_window.destroy()

# Represents the main window of the application.
# Contains methods to set up the UI, create various UI elements (labels, buttons), and handle user interactions.
# Provides functionality to launch other components like settings and file matching tools.


class MainWindow:
    def __init__(self, master=None):
        self.master = master
        self.data = None

        # UI Setup
        self.setup_ui()

    def setup_ui(self):
        self.master.columnconfigure(0, weight=1)
        self.master.rowconfigure(0, weight=1)

        self.frame_1 = ttk.Frame(self.master)
        self.frame_1.columnconfigure(1, weight=1)
        self.frame_1.rowconfigure((0, 1, 2), weight=1)

        # Create label and entry for Model file location
        self.model_entry = self.create_label_entry(
            "Model file location:", self.master.model_path, 0, self.model_browse_func)

        # Create label and entry for Synchro file folder
        self.syn_entry = self.create_label_entry(
            "Synchro file folder:", self.master.synchro_dir, 1, self.syn_browse_func)

        self.los_button = ttk.Checkbutton(
            self.frame_1, variable=self.master.update_los, text='Update LOS Table')
        self.los_button.grid(column=0, row=2)

        self.utilities = ttk.Labelframe(self.frame_1, text='Other Functions')
        self.utilities.grid(row=3, column=1)

        self.create_button("Copy Files", self.copy, 0, 0,
                           parent=self.utilities, side='left')
        # Placeholder for LOS Only button
        self.create_button("LOS Only", None, 0, 1,
                           parent=self.utilities, side='left')

        notes = '''Instructions:\n
                1. Please ensure the Synchro files you wish to update are not open on any computer.\n
                2. Check that the model file is in our standard format.'''
        self.note_label = ttk.Label(self.frame_1, text=notes)
        self.note_label.grid(row=4, columnspan=3)

        self.frame_1.grid(sticky='nsew')

        # Main widget
        self.mainwindow = self.frame_1

    def create_label_entry(self, label_text, default_value, row, browse_command):
        label = ttk.Label(self.frame_1, text=label_text)
        label.grid(column=0, row=row)

        entry = ttk.Entry(self.frame_1)
        entry.insert(0, default_value)
        entry.grid(column=1, row=row, sticky='nsew')

        browse_button = ttk.Button(
            self.frame_1, text='Browse', command=browse_command)
        browse_button.grid(row=row, column=2)

        return entry  # Return the entry widget to assign it to an instance variable

    def create_button(self, text, command, col, row, parent=None, sticky=None, side=None):
        button = ttk.Button(parent if parent else self.frame_1,
                            text=text, command=command)
        if side:
            button.pack(side=side)
        else:
            button.grid(column=col, row=row, sticky=sticky)

    def model_browse_func(self):
        path = filedialog.askopenfile(filetypes=[('Excel Files', '*.xlsx')])
        if path:
            self.update_entry(path.name, self.model_entry)

    def syn_browse_func(self):
        path = filedialog.askdirectory()
        if path:
            self.update_entry(path, self.syn_entry)

    def update_entry(self, path, entry):
        path = replace_slash(path)
        entry.delete(0, 'end')
        entry.insert(0, path)

    def launch_settings(self):
        Settings(self.master)

    def copy(self):
        Copier(self.master)

    def run(self):
        self.mainwindow.mainloop()

# Extends tk.Tk to serve as the base window for the application.
# Contains several methods for interacting with Synchro files and managing the application's
# main operations, such as matching worksheet names, converting data, and handling errors.


class Base(tk.Tk):
    # Default values for the number of rows and columns in the application
    DEFAULT_ROWS = 1000
    DEFAULT_COLUMNS = 30

    # List of valid scenario types
    VALID_SCENARIOS = ['EXISTING', 'NO BUILD', 'BUILD']

    # Mapping of scenario conditions to their abbreviations
    SCENARIO_CONDITIONS = {
        'EXISTING': ['EX'],
        'NO BUILD': ['NB'],
        'BUILD': ['B', 'BD'],
        'IMPROVEMENT': ['IMP']
    }

    def __init__(self):
        # Initialize the Tkinter window
        super().__init__()
        self.title('Synchronizer')  # Set the title of the window
        # Center and size the window
        self.geometry(center_window(500, 200, self))
        # Allow the window to be behind others
        self.wm_attributes('-topmost', 0)

        # Initialize various attributes for managing the application state
        self.windows = {}  # Dictionary to hold child windows
        self.main_win = None  # Reference to the main window
        self.storage_dir = None  # Directory to store files
        self.model_sheet_name = str()  # Name of the current model sheet
        self.model_data = {}  # Dictionary to hold model data
        self.scenario_list = []  # List to store scenario objects
        self.scenario_data = {}  # Dictionary to hold scenario-related data
        self.selected_scenarios = []  # List of selected scenarios
        self.scenarios = []  # List to hold Scenario objects
        self.ws = None  # Reference to the currently active worksheet
        self.data_columns = []  # List of data columns

        # Load settings for the application
        defaults = {
            'synchro_exe': 'C:\\Program Files (x86)\\Trafficware\\Version10\\Synchro10.exe',
            'synchro_dir': '',
            'model_path': '',
            'rows': self.DEFAULT_ROWS,
            'columns': self.DEFAULT_COLUMNS,
            'update_los': 1
        }
        saved_settings = load_settings()  # Load previously saved settings

        # Set application paths and default settings, using saved values if available
        self.synchro_app_path = saved_settings.get(
            'synchro_exe', defaults['synchro_exe'])
        self.synchro_dir = saved_settings.get(
            'synchro_dir', defaults['synchro_dir'])
        self.model_path = saved_settings.get(
            'model_path', defaults['model_path'])
        self.default_rows = saved_settings.get('rows', defaults['rows'])
        self.default_columns = saved_settings.get(
            'columns', defaults['columns'])
        self.update_los = saved_settings.get(
            'update_los', defaults['update_los'])

    def find_volume_data(self, extra_scenario=None):
        """
        Load volume data from the model workbook based on specified scenarios.

        Args:
            extra_scenario (str, optional): An additional scenario to consider.

        Returns:
            output.keys(): Keys of the scenario data collected from the model.
        """
        valid_scenarios = [
            extra_scenario] if extra_scenario else self.VALID_SCENARIOS
        output = {}

        wb = xl.load_workbook(filename=self.model_path,
                              data_only=True)  # Load the model workbook
        self.model_sheet_name = max(
            wb.sheetnames, key=lambda sheet: similar(sheet, 'Model'), default=None)
        self.ws = wb[self.model_sheet_name]  # Set the active worksheet

        # Iterate through rows of the worksheet to find valid scenario data
        for row in range(1, self.ws.max_row):
            if self.ws.cell(row, 1).value == 1:  # Check if the row is valid
                year, scenario = None, None
                # Iterate through columns to extract year, scenario, and hour data
                for column in range(1, self.ws.max_column):
                    year_cell = self.ws.cell(row - 4, column).value
                    scenario_cell = self.ws.cell(row - 3, column).value
                    hour_cell = self.ws.cell(row - 2, column).value

                    if year_cell is not None:
                        year = str(year_cell)  # Convert year to string
                    if scenario_cell is not None:
                        # Convert scenario to string
                        scenario = str(scenario_cell)
                    if hour_cell in ['AM', 'PM', 'SAT'] and scenario in valid_scenarios:
                        # Create a scenario name and check for duplicates
                        name = f"{year} {scenario} {hour_cell}"
                        if not any(found_scenario.name == name for found_scenario in self.scenarios):
                            sc = Scenario(name)  # Create a new Scenario object
                            sc.hour = hour_cell
                            sc.year = year
                            sc.condition = scenario
                            sc.model_data_column = column  # Store column index for the model data
                            # Match the corresponding .syn file
                            self.match_syn_file(sc, self.synchro_dir)
                            # Add the scenario to the list
                            self.scenarios.append(sc)
                        else:
                            messagebox.showwarning(
                                'Duplicate', 'One or more scenarios were duplicated and not added.')

        self.scenario_data = output  # Update scenario data
        return output.keys()  # Return the keys of the collected scenario data

    # Convert model volumes to Synchro UTDF
    def convert_utdf(self, scenario='test_write', column=5):
        # Open model to copy data
        # wb = xl.load_workbook(filename=model)
        # active = wb.active
        ws = self.ws  # need to make sure sheet is titled "Model"
        startColumn = 'C'  # Get direction column from user or default
        dataColumns = ['F', 'G', 'H']  # From scenarios to update

        volume_data = dict()
        movement_list = ['RECORDNAME', 'INTID']

        for row in range(15, self.default_rows):
            intersection = None
            direction = None

            cell = ws.cell(row, 1).value
            if type(cell) in [int, float]:
                intersection = int(cell)
                volume_data[intersection] = dict()
            cell = ws.cell(row, 3).value
            if type(cell) == str:
                direction = cell
                volume_data[intersection][direction] = dict()
            turn = ws.cell(row, 4).value
            if intersection and direction and turn:
                volume = ws.cell(row, column).value
                if volume is None:
                    volume = 0
                else:
                    volume = int(volume)
                volume_data[intersection][direction + turn] = volume
                if direction + turn not in movement_list:
                    movement_list.append(direction + turn)

        # Dictionary Format:
        # {intersection:{direction1:{L:0, T:0, R:0
        #                                        },
        #                            direction2:{},
        #                            direction3:{}}}

        # print(volume_data)
        file = self.storage_dir + '\\' + scenario + '.csv'
        with open(file, 'w', newline='') as f:
            f.write('[Lanes]\nLane Group Data\n')
            writer = csv.DictWriter(f, fieldnames=movement_list)
            writer.writeheader()
            for intid in volume_data:
                payload = volume_data[intid]
                payload['RECORDNAME'] = 'Volume'
                payload['INTID'] = intid
                writer.writerow(payload)
        return file


# _______________LOS_______________

    def update_report(self, scenarios, report_table=None):
        # If no report_table is specified, default to 'synchronizer results.xlsx'
        if report_table is None:
            report_table = 'synchronizer results.xlsx'

        # Combine the storage directory with the report_table name to get the full file path
        report_table = self.storage_dir + '\\' + report_table

        # Create a new Excel workbook and activate the default sheet
        wb = xl.Workbook()
        ws = wb.active

        # Rename the default sheet to 'AM'
        ws.title = 'AM'

        # Loop through each scenario in the scenarios list
        for scenario in scenarios:
            # Get the LOS (Level of Service) data and hour from the scenario object
            data = scenario.los_data
            hour = scenario.hour

            # Retrieve or create the sheet based on the scenario's hour (e.g., 'AM', 'PM')
            sheet = get_sheet(wb, hour)

            # Get the traffic condition (e.g., EXISTING, NO-BUILD, BUILD)
            condition = scenario.condition

            # Determine the column to store the data based on the condition
            if condition == 'EXISTING':
                column = 5
            elif condition == 'NO-BUILD':
                column = 8
            elif condition == 'BUILD':
                column = 11
            else:
                column = sheet.max_column  # If an unrecognized condition, use the last column

            # Loop through each intersection in the LOS data
            for intersection in data:
                # Get the row in the sheet corresponding to this intersection
                row, method = get_row(sheet, intersection)
                ov_los = None
                ov_delay = None

                # Loop through each turning movement in the intersection's data
                for turn_move, values in data[intersection].items():
                    # Special handling for 'overall' turning movement (aggregated data)
                    if turn_move == 'overall':
                        ov_delay = values['delay']
                        ov_los = values['los']
                        continue

                    # Generate a name for the movement (e.g., 'Left Turn') based on the config
                    movement_name = label(turn_move, values.get('config', ''))
                    if movement_name:
                        # Initialize lists to store various values (e.g., v/c ratio, LOS, delay)
                        vc_ratios = list()
                        los_values = list()
                        delays = list()
                        app_los_values = list()
                        app_delays = list()
                        last_move = turn_move[:2]

                        # Process each direction for the movement (e.g., EB, WB)
                        for direction in movement_name:
                            search = turn_move[:2] + direction

                            # Ensure the search key exists before retrieving the data
                            if search not in data[intersection].keys():
                                continue

                            # Append the corresponding values for v/c ratio, LOS, and delay
                            vc_ratios.append(
                                data[intersection][search].get('vc_ratio', ''))
                            los_values.append(
                                data[intersection][search].get('ln_los', ''))
                            delays.append(data[intersection]
                                          [search].get('ln_delay', ''))
                            app_los_values.append(
                                data[intersection][search].get('app_los', ''))
                            app_delays.append(
                                data[intersection][search].get('app_delay', ''))

                        # Take the maximum values for each metric
                        vc = max(vc_ratios)
                        los = max(los_values)
                        delay = max(delays)
                        app_los = max(app_los_values)
                        app_delay = max(app_delays)

                        # If all values are empty, skip the current movement
                        if vc == '' and los == '' and delay == '':
                            continue

                        # Write the data into the sheet based on the method (direct, insert, append)
                        if method == 'direct':
                            sheet.cell(row, 1).value = intersection
                            sheet.cell(row, 3).value = turn_move[:2]
                            sheet.cell(row, 4).value = movement_name
                            sheet.cell(row, column).value = vc
                            sheet.cell(row, column + 1).value = los
                            sheet.cell(row, column + 2).value = delay
                            row += 1

                        elif method == 'insert':
                            sheet.insert_rows(row)
                            sheet.cell(row, 1).value = intersection
                            sheet.cell(row, 3).value = turn_move[:2]
                            sheet.cell(row, 4).value = movement_name
                            sheet.cell(row, column).value = vc
                            sheet.cell(row, column + 1).value = los
                            sheet.cell(row, column + 2).value = delay
                            row += 1

                        elif method == 'append':
                            row += 1
                            sheet.cell(row, 1).value = intersection
                            sheet.cell(row, 3).value = turn_move[:2]
                            sheet.cell(row, 4).value = movement_name
                            sheet.cell(row, column).value = vc
                            sheet.cell(row, column + 1).value = los
                            sheet.cell(row, column + 2).value = delay

                        # If there are no approach LOS and delay values, skip to the next move
                        if app_delay == '' and app_los == '':
                            continue

                        # Write the approach LOS and delay if the last movement is different
                        if last_move and turn_move != last_move:
                            if method == 'direct':
                                sheet.cell(row, 1).value = intersection
                                sheet.cell(row, 3).value = 'Approach'
                                sheet.cell(row, column + 1).value = app_los
                                sheet.cell(row, column + 2).value = app_delay
                                row += 1

                            elif method == 'insert':
                                sheet.insert_rows(row)
                                sheet.cell(row, 1).value = intersection
                                sheet.cell(row, 3).value = 'Approach'
                                sheet.cell(row, column + 1).value = app_los
                                sheet.cell(row, column + 2).value = app_delay
                                row += 1

                            elif method == 'append':
                                row += 1
                                sheet.cell(row, 1).value = intersection
                                sheet.cell(row, 3).value = 'Approach'
                                sheet.cell(row, column + 1).value = app_los
                                sheet.cell(row, column + 2).value = app_delay

                    # Write the overall LOS and delay if available
                    if ov_los and ov_delay:
                        if method == 'direct':
                            sheet.cell(row, 1).value = intersection
                            sheet.cell(row, 3).value = 'Overall'
                            sheet.cell(row, column + 1).value = ov_los
                            sheet.cell(row, column + 2).value = ov_delay
                            row += 1

                        elif method == 'insert':
                            sheet.insert_rows(row)
                            sheet.cell(row, 1).value = intersection
                            sheet.cell(row, 3).value = 'Overall'
                            sheet.cell(row, column + 1).value = ov_los
                            sheet.cell(row, column + 2).value = ov_delay
                            row += 1

                        elif method == 'append':
                            row += 1
                            sheet.cell(row, 1).value = intersection
                            sheet.cell(row, 3).value = 'Overall'
                            sheet.cell(row, column + 1).value = ov_los
                            sheet.cell(row, column + 2).value = ov_delay

        # Save the workbook to the specified file
        wb.save(report_table)

        # Return the path to the report file
        return report_table


class ProgressWindow:
    def __init__(self, master=None):
        self.master = master
        # build ui
        self.progress_window = tk.Toplevel(self.master)
        self.progress_window.geometry(center_window(400, 400, self.master))
        self.progress_window.columnconfigure(0, weight=1)
        self.progress_frame = ttk.Frame(self.progress_window)
        self.progress_frame.columnconfigure(0, weight=1)
        self.progress_frame.columnconfigure(1, weight=0)
        self.status_text_box = tk.Text(self.progress_frame)
        self.status_text_box.config(autoseparators='false')
        self.status_text_box.grid(column='0', row='0', sticky='nsew')
        self.scrollbar_3 = ttk.Scrollbar(
            self.progress_frame, command=self.status_text_box.yview)
        self.scrollbar_3.grid(column='1', row='0', sticky='nsew')
        self.status_text_box.configure(yscrollcommand=self.scrollbar_3.set)
        # self.button_3 = ttk.Button(self.progress_frame)
        # self.button_3.config(text='button_3')
        # self.button_3.grid(column='0', row='1', sticky='s')
        # self.progress_frame.config(height='200', width='200')
        self.progress_frame.grid(padx=10, pady=10, sticky='nsew')
        # self.progress_window.config(height='200', width='200')
        self.progress_window.title('Program Status')
        self.progress_window.after(6000, self.run)

    def run(self):
        time.sleep(2)
        # success = self.master.startup()
        # if success != 0:
        #     self.status_text_box.insert('end', 'Failed to start Synchro\n')
        #     return
        for scenario_obj in self.master.selected_scenarios:
            scenario = scenario_obj.name
            filename = scenario_obj.syn_file
            column = scenario_obj.model_data_column
            process_update = 'Processing: ' + scenario + '\n'
            self.status_text_box.insert('end', process_update)
            utdf_volumes = self.master.convert_utdf(
                scenario=scenario, column=column)
            self.status_text_box.insert(
                'end', 'Importing volumes to Synchro... \n')
            self.master.import_to_synchro(filename, utdf_volumes)
            self.status_text_box.insert('end', 'Import complete\n')

            if self.master.update_los:
                self.status_text_box.insert(
                    'end', 'Exporting LOS data from Synchro...\n')
                scenario_obj.los_results = self.master.export_from_synchro(
                    scenario)
                time.sleep(5)
                self.status_text_box.insert('end', 'Export complete\n')
                scenario_obj.los_data = standardize(scenario_obj.los_results)

        if self.master.update_los:
            self.status_text_box.insert(
                'end', 'Writing LOS data to excel file\n')
            output_results = self.master.update_report(
                self.master.selected_scenarios)
            self.status_text_box.insert(
                'end', 'Write complete\nThe program has finished\n')
            self.status_text_box.insert(
                'end', f'LOS results are saved at: {output_results}')


class Copier:
    def __init__(self, master=None):
        self.window = tk.Toplevel(master)
        self.window.columnconfigure(1, weight=1)

        self.old_dir_label = ttk.Label(self.window, text='Copy from:')
        self.old_dir_label.grid(row=0, column=0, sticky='e', padx=10)

        self.old_dir_entry = ttk.Entry(self.window)
        self.old_dir_entry.grid(row=0, column=1, sticky='ew')

        self.old_dir_button = ttk.Button(
            self.window, text='Browse', command=self.browse)
        self.old_dir_button.bind('<Button 1>', self.browse)
        self.old_dir_button.grid(row=0, column=2)

        self.new_dir_label = ttk.Label(self.window, text='Copy to:')
        self.new_dir_label.grid(row=1, column=0, sticky='e', padx=10)

        self.new_dir_entry = ttk.Entry(self.window)
        self.new_dir_entry.grid(row=1, column=1, sticky='ew')

        self.new_dir_button = ttk.Button(
            self.window, text='Browse', command=self.browse)
        self.new_dir_button.bind('<Button 1>', self.browse)
        self.new_dir_button.grid(row=1, column=2)

        self.new_date_label = ttk.Label(self.window, text='New date:')
        self.new_date_label.grid(row=2, column=0, sticky='e', padx=10)

        self.new_date_entry = ttk.Entry(self.window)
        self.new_date_entry.grid(row=2, column=1, sticky='ew')

        self.start = ttk.Button(
            self.window, text='Start', command=self.copy_files)
        self.start.grid(row=3, columnspan=3)

        # self.check_syn = ttk.Checkbutton(self.window, text='Synchro')
        # self.check_syn.grid(row=3)
        #
        # self.check_pdf = ttk.Checkbutton(self.window, text='Synchro')
        # self.check_pdf.grid(row=4)

    def browse(self, event):
        file = filedialog.askdirectory()
        if file is None:
            return
        row = event.widget.grid_info()['row']
        if row == 0:
            entry = self.old_dir_entry
        else:
            entry = self.new_dir_entry

        entry.delete(0, 'end')
        entry.insert('end', file)

    def copy_files(self):
        pattern = re.compile('[0-9]*')
        old_dir = self.old_dir_entry.get()

        for file in os.scandir(old_dir):
            print(file)
            if not file.name.endswith('syn'):
                continue
            new_date = self.new_date_entry.get()
            old_date = re.match(pattern, file.name).group(0)
            if old_date == '':
                new_file_name = new_date + file.name
            else:
                new_file_name = file.name.replace(old_date, new_date)
            new_path = self.new_dir_entry.get() + '\\' + new_file_name
            copy(file.path, new_path)
        self.window.destroy()

"""

"""
def write_to_excel(file_path, lane_groups, delay, vc_ratio, los):
    # Helper function to transform each sublist into a dictionary with lane directions
    def separate_characters(sublist):
        result_dict = {}
        for item in sublist:
            # Extract the prefix and 'L', 'T', 'R' characters
            rest_chars = ''.join([char for char in item if char not in 'LTR'])
            separated_chars = [char for char in item if char in 'LTR']

            if rest_chars in result_dict:
                # Append characters if the key exists
                result_dict[rest_chars].extend(separated_chars)
            else:
                # Create new entry for the key
                result_dict[rest_chars] = separated_chars

        return result_dict

    # Function to enumerate the lane groups and create a dictionary
    def enumerate_result_list(result):
        enumerated_dict = {}
        for index, item in enumerate(result, start=1):
            if isinstance(item, list):  # Ensure each item is processed as a list
                enumerated_dict[index] = separate_characters(
                    item)  # Transform each list into a dictionary
            else:
                # Handle already existing dictionaries
                enumerated_dict[index] = item
        return enumerated_dict

    # Helper function to write headers
    def write_headers(ws, start_col='C'):
        headers = ['V/c', 'LOS', 'Delay']
        for idx, header in enumerate(headers):
            col = chr(ord(start_col) + idx)  # Dynamic column calculation
            ws[f'{col}2'] = header  # Write the header to row 2

    # Get the file name without extension
    file_with_ext = os.path.basename(file_path)
    file_name = os.path.splitext(file_with_ext)[0]

    # Enumerate the lane groups
    intersection_data = enumerate_result_list(lane_groups)

    # Create a new Excel workbook and add a sheet
    wb = Workbook()
    ws = wb.active  # Get the active worksheet

    # Write the file name in cell A1
    ws['A1'] = file_name

    # Write the headers for V/C, LOS, and Delay in row 2 (starting at column C)
    write_headers(ws, 'C')
    write_headers(ws, 'F')
    write_headers(ws, 'I')

    # Define the order of keys to process
    key_order = ['EB', 'WB', 'NB', 'SB', 'NE', 'NW', 'SE', 'SW']

    # Keep track of the last used row
    current_row = 2  # Starting at row 2 (A2 will be filled first)

    # Iterate through the enumerated intersection data
    for intersection_id, data in intersection_data.items():
        # Write the intersection ID in column A
        ws[f'A{current_row}'] = intersection_id

        # Move to the next row after writing the intersection ID
        current_row += 1

        # Ensure that 'data' is a dictionary before attempting to access its keys
        if isinstance(data, dict):
            # For each key in the specified order:
            for key in key_order:
                # Write the key in column A
                ws[f'A{current_row}'] = key

                # Check if there's a corresponding value in the lane data
                if key in data and data[key]:  # If values exist
                    # Write the corresponding values in column B, starting from the current row
                    for idx, item in enumerate(data[key]):
                        # Write each value downwards in column B
                        ws[f'B{current_row + idx}'] = item
                    # Update the current_row to the next empty row after writing all values
                    current_row += len(data[key])
                else:
                    # If no values exist, write an empty cell in column B
                    # Explicitly write an empty string
                    ws[f'B{current_row}'] = ''
                    current_row += 1  # Move to the next row for the next key

    # Write V/C, LOS, and Delay data into their respective columns, ensuring empty entries are included
    # Start from row 3 (the row after the headers)

    # Write V/C Ratio
    current_row_vc = 3  # Start writing V/C values
    for idx, vc_list in enumerate(vc_ratio):
        if idx < len(intersection_data):  # Ensure we don't exceed the intersection data length
            for item in vc_list:
                # Write each value downwards in column C
                ws[f'C{current_row_vc}'] = item
                current_row_vc += 1  # Move to the next row
            # Fill empty cells if the length of vc_list is shorter than the max
            while current_row_vc < (3 + len(vc_ratio)):
                ws[f'C{current_row_vc}'] = ''  # Write an empty string
                current_row_vc += 1

    # Write LOS values
    current_row_los = 3  # Start writing LOS values
    for idx, los_list in enumerate(los):
        if idx < len(intersection_data):  # Ensure we don't exceed the intersection data length
            for item in los_list:
                # Write each value downwards in column D
                ws[f'D{current_row_los}'] = item
                current_row_los += 1  # Move to the next row
            # Fill empty cells if the length of los_list is shorter than the max
            while current_row_los < (3 + len(los)):
                ws[f'D{current_row_los}'] = ''  # Write an empty string
                current_row_los += 1

    # Write Delay values
    current_row_delay = 3  # Start writing Delay values
    for idx, delay_list in enumerate(delay):
        if idx < len(intersection_data):  # Ensure we don't exceed the intersection data length
            for item in delay_list:
                # Write each value downwards in column E
                ws[f'E{current_row_delay}'] = item
                current_row_delay += 1  # Move to the next row
            # Fill empty cells if the length of delay_list is shorter than the max
            while current_row_delay < (3 + len(delay)):
                ws[f'E{current_row_delay}'] = ''  # Write an empty string
                current_row_delay += 1

    # Save the workbook
    excel_file_path = f"{file_name}_results.xlsx"
    wb.save(excel_file_path)
    print(f"Intersection data written to {excel_file_path}")


def separate_characters(result):
    # Initialize a list to hold the dictionaries
    transformed_results = []

    # Iterate through each sublist in the result
    for sublist in result:
        # Initialize a dictionary for this sublist
        result_dict = {}

        # Process each string in the sublist
        for item in sublist:
            # Extract the characters that are not 'L', 'T', or 'R' for the prefix
            # Characters other than L, T, R
            rest_chars = ''.join([char for char in item if char not in 'LTR'])
            # Characters that are L, T, or R
            separated_chars = [char for char in item if char in 'LTR']

            # Use the prefix as the key in the dictionary
            if rest_chars in result_dict:
                # Append to the existing entry if the key already exists
                result_dict[rest_chars].extend(separated_chars)
            else:
                # Create a new entry if the key does not exist
                result_dict[rest_chars] = separated_chars

        # Append the dictionary to the list of transformed results
        transformed_results.append(result_dict)

    return transformed_results


def save_as_csv(excel_file_path, csv_file_path):
    workbook = load_workbook(filename=excel_file_path)
    sheet = workbook.active

    with open(csv_file_path, mode='w', newline="") as file:
        writer = csv.writer(file)

        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)


def write_direction_data_to_files(sheet, matched_results, relevant_columns, headers, output_start_row=4):
    """
    Writes Volume, PHF, and HeavyVehicles data for each intersection and direction-turn 
    from the specified column ranges in relevant_columns, and saves the results to 
    separate files named based on the header in row 1 of each column range.

    Args:
    - sheet: The active sheet from which data is being read.
    - matched_results: A dictionary containing intersections and their corresponding turn data.
    - relevant_columns: A list of starting columns (e.g., [6, 9, 12] for 'F', 'I', 'L') 
      from which Volume, PHF, and HeavyVehicles are read.
    - output_start_row: The row in the output sheet to start writing the data (default is 4).
    """
    for start_column in relevant_columns:
        # Define column positions relative to the starting column
        volume_col = start_column       # Volume is in start_column (e.g., F)
        phf_col = start_column + 1      # PHF is in start_column + 1 (e.g., G)
        # HeavyVehicles is in start_column + 2 (e.g., H)
        heavy_vehicles_col = start_column + 2

        # Get the header name from row 1 of the start_column (e.g., F1, I1, etc.)
        file_name_header = sheet.cell(row=1, column=start_column).value
        if not file_name_header:
            print(
                f"Skipping columns starting at {start_column} as no header was found in row 1.")
            continue

        # Create a new workbook for this specific column set
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "Results"
        output_sheet["A1"] = "[Lanes]"
        output_sheet["A2"] = "Lane Group Data"

        # Label cells with corresponding headers (A3-P3)
        for col, header in enumerate(headers, start=1):
            output_sheet.cell(row=3, column=col).value = header

        # Reset the output start row for each file
        output_start_row = 4

        # Iterate over each intersection and its direction-turn results
        for intersection_id, turns in matched_results.items():
            # Write Intersection ID and Labels in the output sheet
            output_sheet.cell(row=output_start_row, column=1).value = "Volume"
            output_sheet.cell(row=output_start_row + 1, column=1).value = "PHF"
            output_sheet.cell(row=output_start_row + 2,
                              column=1).value = "HeavyVehicles"
            output_sheet.cell(row=output_start_row,
                              column=2).value = intersection_id
            output_sheet.cell(row=output_start_row + 1,
                              column=2).value = intersection_id
            output_sheet.cell(row=output_start_row + 2,
                              column=2).value = intersection_id

            # Process each direction-turn within the intersection
            for direction_turn, info in turns.items():
                row_found = info['row']

                # Read data from the specified columns for the current row
                volume = sheet.cell(row=row_found, column=volume_col).value
                phf = sheet.cell(row=row_found, column=phf_col).value
                heavy_vehicles = sheet.cell(
                    row=row_found, column=heavy_vehicles_col).value

                # Write the data into the output sheet under the correct direction-turn column
                header_column = info['header_column']
                output_sheet.cell(row=output_start_row,
                                  column=header_column).value = volume
                output_sheet.cell(row=output_start_row + 1,
                                  column=header_column).value = phf
                output_sheet.cell(row=output_start_row + 2,
                                  column=header_column).value = heavy_vehicles

                # Debugging output
                print(f"Wrote to Results for intersection {intersection_id}, direction {direction_turn}: "
                      f"Volume: {volume}, PHF: {phf}, HeavyVehicles: {heavy_vehicles}")

            # Move to the next output row for the next intersection
            output_start_row += 3  # 3 rows for data + 1 row for separation

        # Save the output workbook to a file named by the header in row 1 of the start column
        output_file_path = f"{file_name_header}.xlsx"
        output_workbook.save(output_file_path)
        save_as_csv(output_file_path, f"{file_name_header}.csv")
        os.remove(f"{file_name_header}.xlsx")
        print(f"Output file saved as {file_name_header}.csv")

    return


""" STEP 1 """


def read_input_file(file_path):
    # Load the input workbook and select the active sheet
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    # Define headers for the output sheet
    headers = [
        "RECORDNAME", "INTID", "NBL", "NBT", "NBR",
        "SBL", "SBT", "SBR", "EBL", "EBT", "EBR",
        "WBL", "WBT", "WBR", "NWR", "NWL", "NWT", "NEL", "NET", "NER",
        "SEL", "SER", "SET", "SWL", "SWR", "SWT", "PED", "HOLD"
    ]

    consecutive_empty_cells = 0
    intersections = {}

    # First pass: Find all intersection IDs and their corresponding row numbers
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value is None:
            consecutive_empty_cells += 1
            if consecutive_empty_cells >= 25:
                break
        else:
            consecutive_empty_cells = 0
            if isinstance(cell_value, int):
                intersections[cell_value] = row

    print(f"Found intersections: {intersections}")

    directions = ["EB", "WB", "NB", "SB", "NW", "NE", "SW", "SE"]
    # output_start_row = 4  # Start writing from row 4

    # Dictionary to store results for each intersection
    intersection_results = {}

    # Second pass: Process each intersection ID and search for directions
    for intersection_id, row_with_int in intersections.items():
        found_directions = {}

        # Search column C for directions starting from the intersection row
        for search_row in range(row_with_int, sheet.max_row + 1):
            direction_value = sheet.cell(search_row, column=3).value
            if direction_value in directions and direction_value not in found_directions:
                found_directions[direction_value] = search_row
                if len(found_directions) == len(directions):
                    break

        # Dictionary to store combined direction-turn keys (e.g., EBL, WBT)
        direction_turn_results = {}

        # For each found direction, search column D for 'L', 'T', 'R'
        for direction, found_row in found_directions.items():
            # Default is None (not found)
            turn_values = {"L": None, "T": None, "R": None}
            for search_row in range(found_row, sheet.max_row + 1):
                turn_value = sheet.cell(search_row, column=4).value
                if turn_value in ["L", "T", "R"]:
                    # Store the row number for each turn type found
                    turn_values[turn_value] = search_row
                # Break when all turn values have been found
                if all(turn_values.values()):
                    break

            # Combine direction and turn type to form keys like "EBL", "NBT", etc.
            for turn_type, row_found in turn_values.items():
                if row_found is not None:  # Only store if the turn was found
                    combined_key = f"{direction}{turn_type}"
                    direction_turn_results[combined_key] = row_found

        # Store the results for the current intersection
        intersection_results[intersection_id] = direction_turn_results

        # Display the results for debugging
        print(f"Direction-turn results for intersection {intersection_id}: {direction_turn_results}")

    # Match direction-turn results with corresponding headers
    header_mapping = {header: idx + 1 for idx, header in enumerate(headers)}

    matched_results = {}

    for intersection_id, turn_results in intersection_results.items():
        matched_results[intersection_id] = {}
        for direction_turn, row in turn_results.items():
            if direction_turn in header_mapping:
                matched_results[intersection_id][direction_turn] = {
                    "row": row,
                    "header_column": header_mapping[direction_turn]
                }

    relevant_columns = [6, 9, 12, 15]  # F-H, I-K, L-N

    write_direction_data_to_files(
        sheet, matched_results, relevant_columns, headers=headers, output_start_row=4)

    # Return intersection results if needed elsewhere
    return intersection_results

"""
 ~ Data extraction functions ~

    * parse_minor_lane_mvmt(lines, start_line, end_line)
    * process_directions(twsc_summary_results)
    * parse_overall_data_v2(file_path)
    * parse_twsc_approach(df)
    * extract_data_to_csv(file_path, output_file)
    * parse_lane_configs(int_lane_groups, intersection_ids)
"""


def parse_minor_lane_mvmt(lines, start_line, end_line):
    """
        Parse the "Minor Lane/Major Mvmt" data between the start and end lines.
        This function extracts the delay, V/C ratio, and LOS from lines containing these terms.
        Helper function to the parse_overall_data function.
    """
    
    search_phrase = "Minor Lane/Major Mvmt"
    search_terms = [r'\bControl Delay\b', r'\bV/C Ratio\b', r'\bLOS\b', r'\bCapacity\b']

    # Initialize lists for each search term
    result = []
    delay_results = []
    vc_ratio_results = []
    los_results = []
    capacity_results = []

    # Search for "Minor Lane/Major Mvmt" in the provided line range
    for line_number in range(start_line, end_line):
        line = lines[line_number]
        if search_phrase in line:
            # Process the "Minor Lane/Major Mvmt" line to get the directions
            after_phrase = line.split(search_phrase)[1].strip()
            cleaned_value = ' '.join(after_phrase.split()).strip()
            result = (cleaned_value.split())
            # print(f"Result: {result}")
            # Now search for Delay, V/C Ratio, and LOS in lines below the current line
            for term in search_terms:
                term_results = []  # Temporary list to hold results for the current term
                for search_line_number in range(line_number + 1, end_line):
                    # Remove whitespace for accurate matching
                    line = lines[search_line_number].strip()
                    if re.search(term, line, re.IGNORECASE):
                        # Extract values after the term
                        # print(f"Found term: {term} in line: {line}")

                        if 'control delay' in term.lower() or 'v/c ratio' in term.lower() or 'capacity' in term.lower():
                            # For control delay or V/C ratio, we extract numbers (floats or '-')
                            numbers = re.findall(r'(\d+\.\d+|\d+|-)', line)
                            term_results.extend(
                                [float(num) if num != '-' else num for num in numbers])

                        elif 'los' in term.lower():
                            # For LOS, we extract single capital letters (A-F) or '-'
                            capital_letters = re.findall(r'\b[A-F]\b|-', line)
                            term_results.extend(capital_letters)

                # Add the term results to the corresponding results list
                if 'control delay' in term.lower():
                    delay_results.append(
                        term_results if term_results else ['-'])
                elif 'v/c ratio' in term.lower():
                    vc_ratio_results.append(
                        term_results if term_results else ['-'])
                elif 'los' in term.lower():
                    los_results.append(term_results if term_results else ['-'])
                elif 'capacity' in term.lower():
                    capacity_results.append(term_results if term_results else ['-'])
                    
    # Combine the results into tuples for easier reading
    merged_results = []
    for vc_list, los_list, delay_list, capacity_list in zip(vc_ratio_results, los_results, delay_results, capacity_results):
        merged_results = (list(zip(vc_list, los_list, delay_list, capacity_list))) 
    
    # Return the parsed results for integration with other parsing logic
    return result, merged_results


def process_directions(twsc_summary_results, lane_configs):
    processed_list = []
    original_key_list = []
    combined_mvmt_names = []
    # print("Processing Directions...\n")
    print(twsc_summary_results)
    print(f"\n{lane_configs}\n")
    for entry in twsc_summary_results:
        # Start with a dictionary containing just the ID
        processed_dict = {"ID": entry["ID"]}
        original_key_dict = {"ID": entry["ID"]}
        
        # processed_list_str = ''
        # original_list_str = ''
        # Retrieve the lane configuration for the current intersection
        intersection_id = int(entry["ID"])
        lane_config = next(
            (config for config in lane_configs if config["Intersection ID"] == intersection_id), None
        )
        
        combined_mvmt = []
        
        # print(lane_config, '\n')
        # int_label_text = ":"*5 + f' Intersection {intersection_id} ' + ':'*5
        # print('='*len(int_label_text))
        # print(int_label_text)
        # print('='*len(int_label_text))
        # print(" * (Original => Updated)")
        last_dir = None
        # Loop through the dictionary to process directional keys
        for key, value in entry.items():
            if key == "ID":
                continue  # Skip the ID key itself
            # Split the direction from the suffix
            direction = key[:2]  # The first two characters are the direction (EB, WB, NB, SB, NE, NW, SE, SW)
            suffix = key[2:]  # The remaining part is the suffix (Ln1, T, etc.)
            
            config_amount = len(lane_config[direction]) if (lane_config and direction in lane_config) else 1

            # Add suffixes to the original_key_dict
            if direction in original_key_dict:
                original_key_dict[direction].append(suffix)
            else:
                original_key_dict[direction] = [suffix]  # Initialize as a list
                
            # Handle Ln suffix by matching with the lane configuration
            if "Ln" in suffix and lane_config and direction in lane_config:
                try:
                    # Extract the index from Ln suffix and subtract 1 (Ln1 -> index 0)
                    lane_index = int(suffix[2:]) - 1
                    if 0 <= lane_index < len(lane_config[direction]):
                        # Replace suffix with the lane type from the configuration
                        suffix = lane_config[direction][lane_index]
                except (ValueError, IndexError):
                    pass  # If parsing or index retrieval fails, keep the original suffix
            
            # Determine storage format based on config_amount
            if direction not in processed_dict:
                processed_dict[direction] = [] if config_amount > 1 else ""  # Initialize as list or string

            if config_amount > 1:
                processed_dict[direction].append(suffix)
            else:
                processed_dict[direction] += suffix

            # if last_dir != direction:
            #     print(f" {direction}:|")
            # print(f"    '-- {key[2:]} => {suffix}")
            # last_dir = direction
        
        # print()
        
        # # Print summary for the intersection
        # print("\n........ Summary .........")
        # print(f"   {'_' * 17 if len(str(intersection_id)) == 1 else '_' * 18}")
        # print(f"  | Intersection #{intersection_id} |")
        # print(f"  '{'-' * 17 if len(str(intersection_id)) == 1 else '-' * 18}'")

        # print(f">>> Original lane values: {original_key_dict}")
        # print(f"\n>>> Lane Configurations: {lane_config}\n")
        # print(f">>> Updated lane values: {processed_dict}")
        # Format the combined movement names
        
        for direction, value in processed_dict.items():
            if direction != "ID":
                # Join lists into strings for combined movement names
                combined_mvmt.append(direction + ''.join(value) if isinstance(value, list) else direction + value)
        # print(f"   * {original_key_dict[direction]} -> {direction + ''.join(value) if isinstance(value, list) else direction + value}")
        combined_mvmt_names.append(combined_mvmt)
        # Append the processed dictionaries to their respective lists
        processed_list.append(processed_dict)
        original_key_list.append(original_key_dict)

        # print(f"\nUpdated identifiers: {', '.join(combined_mvmt)}")
        # print("\n" + ("*" * 30) + "\n")
        
    return processed_list, original_key_list, combined_mvmt_names


def parse_overall_data_v2(file_path):
    """
        Function to handle the parsing of the summary data
    """

    int_regex = r'^\d+:'  # Regex to match lines that start with an integer followed by a colon

    search_phrases = ["Minor Lane/Major Mvmt", "Intersection Summary", r'^Lane']

    # Lists to store results
    synchro_results = []
    hcm_results = []
    twsc_results = []
    awsc_results = []
    
    # List to store matching line numbers
    matching_lines = []

    # Open and read the file into memory
    with open(file_path, 'r') as file:
        lines = file.readlines()  # Read all lines

    # Find the line numbers containing the integer followed by a colon
    for line_number, line in enumerate(lines, start=1):
        if re.match(int_regex, line):  # Use re.match to check if the line starts with the regex
            matching_lines.append(line_number)
    
    intersection_index = 0
    
    # Now search from each line in matching_lines until the next line matching the regex
    for start_line in matching_lines:
        # Get the ID from the corresponding line in 'lines'
        # Accessing lines using start_line - 1
        id_match = re.match(int_regex, lines[start_line - 1])
        id_value = id_match.group(0).strip(
            ':') if id_match else None  # Get the ID before the colon
        
        # print(f"\nProcessing ID {id_value} at line {start_line}")
        
        # Process Synchro Results first
        found_phrase = False
        for line_number in range(start_line, len(lines)):
            line = lines[line_number]
            
            # Check for a new ID match before processing further
            new_id_match = re.match(int_regex, line)
            if new_id_match:
                new_id_value = new_id_match.group(0).strip(':')
                # print(f"Found new ID {new_id_value} at line {line_number}")
                id_value = new_id_value  # Update the ID to the new one found
            
            if "Intersection Summary" in line:
                # print(f"Found 'Intersection Summary' at line {line_number}")

                found_phrase = True  # Mark that we found the phrase

                # Set the end line to the next empty line starting from this line
                end_line = line_number + 1
                while end_line < len(lines) and lines[end_line].strip() != '':
                    if "HCM" in lines[end_line]:
                        break
                    end_line += 1  # Continue until we find an empty line
                
                if "HCM" in lines[end_line]:
                    # print(f"Found 'HCM' at line {end_line}, skipping Synchro block for this ID\n")
                    continue    
                
                # print(f"Synchro block ends at line {end_line}\n")

                # Initialize values to None
                vc_ratio_value = None
                los_value = None
                delay_value = None

                # Now process the following lines for the search terms until the next blank line
                for search_line_number in range(line_number + 1, end_line):
                    line = lines[search_line_number]
                    
                    # print(f"Processing Synchro data at line {search_line_number}: {line.strip()}")

                    
                    # Check for 'v/c ratio' and extract the next float
                    # if re.search(r'v/c ratio', line, re.IGNORECASE):
                    #     float_match = re.search(r'(\d+\.\d+|\d+)', line)
                    #     if float_match:
                    #         vc_ratio_value = float(float_match.group(0))
                    #         # print(f"Extracted v/c ratio: {vc_ratio_value}")

                    # Check for 'delay' and extract the next float
                    if re.search(r'delay', line, re.IGNORECASE):
                        float_match = re.search(r'(\d+\.\d+|\d+)', line)
                        if float_match:
                            delay_value = float(float_match.group(0))
                            # print(f"Extracted delay: {delay_value}")

                    # Check for 'LOS' and extract the next capital letter (A-F)
                    if re.search(r'LOS', line, re.IGNORECASE):
                        capital_match = re.search(r'\b[A-F]\b', line)
                        if capital_match:
                            los_value = capital_match.group(0)
                            # print(f"Extracted LOS: {los_value}")
                            
                # print(f"Final Synchro values for ID {id_value}: v/c ratio={vc_ratio_value}, delay={delay_value}, LOS={los_value}")

                # Store Synchro results only if ID is not already present
                if not (los_value is None and delay_value is None) and found_phrase:
                    if not any(result['ID'] == id_value for result in synchro_results):
                        synchro_results.append({
                                'ID': id_value,
                                'v/c ratio': vc_ratio_value if vc_ratio_value is not None else '-',
                                'los': los_value if los_value is not None else '-',
                                'delay': delay_value if delay_value is not None else '-',
                                'index': intersection_index
                            })
                        intersection_index += 1

                # Stop further processing of Synchro block and move on to HCM
                break  # Exit after processing this block for Synchro

            # Skip lines between the ID and the next search phrase
            if found_phrase:
                break  # Stop looking at this block and continue with HCM
                
        # Now search for HCM Results
        found_phrase = False  # Reset for HCM block
        for line_number in range(start_line, len(lines)):
            line = lines[line_number]
            
            # Check if a new ID appears before parsing the HCM block
            new_id_match = re.match(int_regex, lines[line_number - 1])
            if new_id_match:
                id_value = new_id_match.group(0).strip(':')
                # print(f"Updated ID to {id_value} for HCM parsing at line {line_number}")
            
            # Process only HCM-related lines
            for search_phrase in search_phrases:
                if search_phrase in line:
                    found_phrase = True
                    phrase_found = search_phrase  # Store which search phrase was found

                    # Set the end line to the next empty line starting from this line
                    end_line = line_number + 1
                    while end_line < len(lines) and lines[end_line].strip() != '':
                        end_line += 1  # Continue until we find an empty line

                    # Initialize values to None
                    vc_ratio_value = '-'
                    los_value = None
                    delay_value = None
                    found_hcm = False  # Flag to track if we found HCM lines

                    # Process HCM lines
                    for search_line_number in range(line_number + 1, end_line):
                        line = lines[search_line_number]

                        # Check for 'v/c ratio' and extract the next float
                        if re.search(r'v/c ratio', line, re.IGNORECASE):
                            float_match = re.search(r'(\d+\.\d+|\d+)', line)
                            if float_match:
                                vc_ratio_value = float(float_match.group(0))

                        # Check for 'delay' and extract the next float
                        if re.search(r'delay', line, re.IGNORECASE):
                            # Find the position of 'delay' in the line
                            delay_pos = line.lower().find('delay')
                            if delay_pos != -1:  # If 'delay' is found
                                # Get everything after 'delay'
                                after_delay = line[delay_pos +
                                                   len('delay'):].strip()
                                # Search for float in the remaining substring
                                float_match = re.search(
                                    r'(\d+\.\d+|\d+)', after_delay)
                                if float_match:
                                    delay_value = float(float_match.group(0))

                        # Check for 'LOS' and extract the next capital letter (A-F)
                        if re.search(r'LOS', line, re.IGNORECASE):
                            capital_match = re.search(r'\b[A-F]\b', line)
                            if capital_match:
                                los_value = capital_match.group(0)
                        
                        if re.search(r'capacity', line, re.IGNORECASE):
                            float_match = re.search(r'(\d+\.\d+|\d+)', line)
                            if float_match:
                                capacity_value = float(float_match.group(0))
                        
                        # Mark the line as HCM only if it starts with "HCM"
                        if line.startswith("HCM"):
                            found_hcm = True

                    # Store HCM results only if we processed HCM lines
                    if found_hcm:
                        # Check which search phrase was found and structure the result accordingly
                        if phrase_found == "Intersection Summary":
                            hcm_results.append({
                                'ID': id_value,
                                'v/c ratio': vc_ratio_value if vc_ratio_value is not None else '-',
                                'los': los_value if los_value is not None else '-',
                                'delay': delay_value if delay_value is not None else '-',
                                'capacity': capacity_value if capacity_value is not None else '-',
                                'index': intersection_index
                            })
                            intersection_index += 1
                        elif phrase_found == "Minor Lane/Major Mvmt":
                            # Now we parse with the other function
                            movement_results, merged_results = parse_minor_lane_mvmt(
                                lines, line_number, end_line)
                            # Create a dictionary where keys are from the movement results
                            hcm_entry = {'ID': id_value}
                            print(movement_results)
                            for i in range(len(movement_results)):
                                # Using the movement results as keys
                                hcm_entry[movement_results[i]] = merged_results[i]
                            twsc_results.append(hcm_entry)
                            # hcm_results.append(hcm_entry)
                        
                    break  # Exit after processing the HCM block

            # Skip lines between the ID and the next search phrase
            if found_phrase:
                break  # Stop looking at this block and move on to the next intersection
    
    # Print the results for debugging
    print("\nSynchro Signalized Summary Results (Intersection Summary):", synchro_results)
    print("\nHCM Signalized Summary Results (Intersection Summary):", hcm_results)
    print("\nTWSC Summary Results (Minor Lane/...):", twsc_results, '\n')

    return twsc_results, synchro_results, hcm_results


def parse_twsc_approach(df):
    """
        Parses the approach data for each direction of a TWSC intersection
        Returns a list of dictionaries, one for each TWSC intersection found in the dataframe
    """
    approach_data = []  # List to hold all parsed data
    intersection_id = None # Store the ID of the intersection we are currently looking at
    
    # Iterate through each row in the DataFrame
    for index, row in df.iterrows():
        line = str(row[0]).strip()  # Consider column 1 as the line to process
        # print(f"\nProcessing line {index}: {line}")
        
        # Check if the row in column 1 contains an integer (this is the intersection ID)
        if line.isdigit():
            intersection_id = int(line)
            # print(f"Found Intersection ID {intersection_id} at line {index}")
            continue  # Move to the next row
        
        # Check if the line starts with "approach" and contains at least one direction
        if line.lower() == "approach":
            # print(f"Found 'Approach' at line {index}: {line}")
            
            # Check if any of the specified directions are present in the line after "approach"
            present_directions = {direction: direction in row.values for direction in ["EB", "WB", "NB", "SB", 'NE', 'NW', 'SE', 'SW']}
            # print(f"Present directions: {present_directions}")
            
            # If no directions are found after "approach", skip this line
            if not any(present_directions.values()):
                print("No valid directions found, skipping line.")
                continue  # Skip the line if no directions are present

            approach_dict = {
                "Intersection ID": intersection_id,
                "EB": {"Approach Delay": None, "Approach LOS": '-'},
                "WB": {"Approach Delay": None, "Approach LOS": '-'},
                "NB": {"Approach Delay": None, "Approach LOS": '-'},
                "SB": {"Approach Delay": None, "Approach LOS": '-'},
                "NE": {"Approach Delay": None, "Approach LOS": '-'},
                "NW": {"Approach Delay": None, "Approach LOS": '-'},
                "SE": {"Approach Delay": None, "Approach LOS": '-'},
                "SW": {"Approach Delay": None, "Approach LOS": '-'},
            }
            
            # pd.set_option('display.max_columns', 100)  # Show all columns
            
            # Step 1: Record the positions (columns) of directions
            direction_columns = {}
            for direction in ["EB", "WB", "NB", "SB", 'NE', 'NW', 'SE', 'SW']:
                if present_directions[direction]:
                    direction_columns[direction] = row[row == direction].index[0]  # Find the column where the direction was found
                    # print(f"Direction {direction} found") # in column {direction_columns[direction]}.")
                    
            # Now check subsequent rows for "HCM Control Delay" and "HCM LOS"
            next_row_index = index + 1
            while next_row_index < len(df):
                next_row = df.iloc[next_row_index]  # Get the next row
                
                # Check for "HCM Control Delay"
                if "hcm control delay" in str(next_row.iloc[0]).lower():
                    # print(f"Found 'HCM Control Delay' at row {next_row_index}.")
                    
                    # Assign the delay values from the columns where directions were found
                    for direction, col in direction_columns.items():
                        delay_value = next_row[col]
                        if pd.notna(delay_value) and re.match(r'\b\d+\.\d+|\b\d+', str(delay_value)):  # Check for numeric values
                            approach_dict[direction]["Approach Delay"] = delay_value
                            # print(f"Setting {direction} Approach Delay: {delay_value}")
                        else:
                            approach_dict[direction]["Approach Delay"] = '-'  # Store '-' if no valid value
                            # print(f"No valid delay value for {direction}, setting to '-'.")

                # Check for "HCM LOS"
                elif "hcm los" in str(next_row.iloc[0]).lower():
                    # print(f"Found 'HCM LOS' at row {next_row_index}.")
                    
                    # Assign the LOS values (A-F) from the columns where directions were found
                    for direction, col in direction_columns.items():
                        los_value = str(next_row[col]).strip().upper()
                        if los_value in 'ABCDEF' and los_value != '':  # Check if the value is a valid LOS (A-F)
                            approach_dict[direction]["Approach LOS"] = los_value
                            # print(f"Setting {direction} Approach LOS: {los_value}")
                        else:
                            approach_dict[direction]["Approach LOS"] = '-'  # Store '-' if no valid LOS value
                            # print(f"No LOS value for {direction}, setting to '-'.")

                # Move to the next row
                next_row_index += 1

                # Exit the loop if an empty row is found
                if next_row.isna().all():  # Check if the row is entirely empty or NaN
                    break

            # Step 3: Remove directions with no valid data
            approach_dict = {
                k: v for k, v in approach_dict.items() if k == "Intersection ID" or (isinstance(v, dict) and (v["Approach Delay"] is not None or v["Approach LOS"] != '-'))
            }
            
            # If there's any valid data, add it to approach_data
            if approach_dict:
                approach_data.append(approach_dict)
                # print(f"Added approach data: {approach_dict}")

    # print(f"\nFinal approach data (TWSC Intersections):\n{approach_data}")
    return approach_data


def extract_data_to_csv(file_path, output_file):
    data = []  # To store the final data for CSV
    # skip_lines = 0  # Counter to track skipped lines
    intersection_count = 0
    intersection_ids = []
    collecting = False  # Flag to indicate if we're collecting data
    collecting_minor_lane = False  # Flag to track if we're collecting after "Minor Lane/Major Mvmt"
    lane_groups = []
    
    # Parsing the relevant data out of the text file
    with open(file_path, 'r') as file:
        for line in file:
            stripped_line = line.rstrip('\n')  # Remove the newline character

            # Step 1: Look for a line starting with a digit and a colon
            if re.match(r'^\d+:', stripped_line):
                # Extract the intersection count from the beginning of the line
                intersection_count = int(re.match(r'^(\d+):', stripped_line).group(1))
                data.append([intersection_count])
                intersection_ids.append(intersection_count)
                
                collecting = True  # Start collecting data
                collecting_minor_lane = False                
                continue  # Skip the current line
                
            # Step 2: Start collecting after finding "Lane Group", "Movement", or "Intersection"
            if collecting or collecting_minor_lane:
                # If the line is empty, stop collecting (except when we're in the middle of collecting after "Minor Lane/Major Mvmt")
                if stripped_line == "":
                    collecting = False
                    collecting_minor_lane = False
                    continue  # Skip the empty line

                # Split the line based on double tabs or multiple spaces
                new_row = re.split(r'\t\t|\s{2}\t|\s\t|\t', stripped_line.strip())
                new_row = [cell for cell in new_row if cell]  # Remove empty cells
                data.append(new_row)  # Append the new row to data
                
            # Look for "Lane Group", "Movement", or "Intersection" after finding the intersection line
            if not collecting and not collecting_minor_lane:
                if re.match(r'\s*Lane Group', stripped_line) or re.match(r'\s*Movement', stripped_line):
                    collecting = True  # Start collecting data when either is found
                    # Split the line based on double tabs or multiple spaces
                    new_row = re.split(r'\t\t|\s{2}\t|\s\t|\t', stripped_line.strip())
                    new_row = [cell.strip() for cell in new_row if cell]  # Remove empty cells after stripping
                    if new_row:  # Check if new_row is not empty after cleaning
                        data.append(new_row)  # Append the new row to data
                        lane_groups.append(new_row[1:])  # Append cleaned values to lane_groups
                        # print(new_row[1:])  # Print the cleaned values
                    # print(f"Collecting data for {intersection_count} under {stripped_line.strip()}")
                # If we find "Intersection" before "Lane Group" or "Movement"
                elif re.match(r'^\s*Intersection', stripped_line):
                    collecting = True  # Start collecting, ignoring blank lines
                    # Split the line based on double tabs or multiple spaces
                    new_row = re.split(r'\t\t|\s{2}\t|\s\t|\t', stripped_line.strip())
                    new_row = [cell for cell in new_row if cell]  # Remove empty cells
                    data.append(new_row)  # Append the new row to data
                    # print(f"Collecting data for {intersection_count} under Intersection")
                # If we find "Minor Lane/Major Mvmt", start collecting until the next blank line
                elif re.match(r'^\s*Minor Lane/Major Mvmt', stripped_line) or re.match(r'^Approach', stripped_line) or re.match(r'^Lane', stripped_line):
                    collecting_minor_lane = True  # Start collecting after "Minor Lane/Major Mvmt"
                    # Split the line based on double tabs or multiple spaces
                    new_row = re.split(r'\t\t|\s{2}\t|\s\t|\t', stripped_line.strip())
                    new_row = [cell for cell in new_row if cell]  # Remove empty cells
                    data.append(new_row)  # Append the new row to data
                    # print(f"Started collecting for {intersection_count} after Minor Lane/Major Mvmt")
                
    # pd.set_option('display.max_rows')  # Show all rows
    # pd.set_option('display.max_columns')  # Show all columns

    # print(f"Intersection ID's stored (length = {len(intersection_ids)}): \n{intersection_ids}\n")
    
    # Step 5: Create a DataFrame and save to CSV
    df = pd.DataFrame(data)
    
    df.to_csv(output_file, index=False)
    
    # print(f"\nDataframe from collected data:\n{df}\n")
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
    
    # Define the terms to search for
    terms_to_match = [
        "V/c ratio(x)",
        "LnGrp Delay(d), s/veh",
        "LnGrp LOS",
        "V/c ratio",
        "Control delay (s/veh)",
        "LOS",
        "Approach Delay (s/veh)",
        "Approach Delay, s/veh",
        "Approach LOS"
    ]
    
    # Initialize an empty dictionary to store row indices
    signalized_int_data = []
    all_intersection_configs = []
    row_indices = {}
    group_config_data = {}  # List of dictionaries
    
    current_id = None
    j = 0
    
    """
        Storing and grouping intersection data (signalized/unsignalized)
    """
    # Iterate through DataFrame rows
    for index, row in df.iterrows():
        # Check if the first column contains a new intersection ID that is in the intersection_ids list
        first_column_value = str(row[0]).strip()
        if first_column_value.isdigit():
            potential_id = int(first_column_value)
            if potential_id in intersection_ids:
                # If we have an existing intersection ID, save the collected data for it
                if current_id is not None and (group_config_data or row_indices):
                    # Check if row_indices has at least one term from terms_to_match
                    if any(term in row_indices.values() for term in terms_to_match):
                        intersection_data = {
                            "Intersection ID": current_id,
                            "Configurations": group_config_data,
                            **{term: line_num for line_num, term in row_indices.items()}
                        }
                        signalized_int_data.append(intersection_data)
                        # print(f"Saved data for Intersection ID {current_id}: {intersection_data}")
                    # Reset data structures for the next intersection
                    row_indices = {}
                    group_config_data = {}
                # Set the new intersection ID
                current_id = potential_id
                continue  # Move to the next iteration as we've identified a new intersection
        # Detect rows containing "Lane Group" or "Movement"
        if "Lane Configurations" in row.values:
            # Create an empty dictionary to hold the configurations
            config_dict = {}
    
            # Iterate over lane_groups[j] and row values simultaneously, skipping empty values
            for i, key in enumerate(lane_groups[j]):
                if i + 1 < len(row):  # Ensure there's a corresponding value in the row
                    value = str(row[i + 1]).strip()  # Get and clean the value in the row
                    if value != 'None' and value != '':  # Only add the key-value pair if the value is non-empty
                        config_dict[key] = value
        
            # Append the config_dict to the group_config_data list if it contains data
            if config_dict:
                group_config_data = config_dict
                all_intersection_configs.append(config_dict)

            # Move to the next set of lane groups
            j += 1
        # Check for the presence of "LOS" first, as it is case-sensitive
        if "LOS" in row.values:
            row_indices[index] = "LOS"
            continue
        # For other terms, check in a case-insensitive manner
        for term in terms_to_match:
            if any(str(cell).lower() == term.lower() for cell in row if term.lower() != "los"):
                row_indices[index] = term
                break  # Exit the inner loop if a term is found
        
    # Final addition for the last intersection data
    if current_id is not None and (group_config_data or row_indices):
        # Ensure the dictionary contains at least one of the terms_to_match
        if any(term in row_indices.values() for term in terms_to_match):
            intersection_data = {
                "Intersection ID": current_id,
                "Configurations": group_config_data,
                **{term: line_num for line_num, term in row_indices.items()}
            }
            signalized_int_data.append(intersection_data)

    # print(f"All intersection configurations: {all_intersection_configs}", f"\nlength = {len(all_intersection_configs)}")
    
    # Print the signalized intersection data for verification
    # print("Signalized Intersection Data:")
    # for entry in signalized_int_data:
    #     print(entry, '\n')
    
    lane_configurations, raw_lane_configs = parse_lane_configs(all_intersection_configs, intersection_ids)
    # print('\nLane Configurations collected...', lane_configurations, f"\nlength = {len(lane_configurations)}")
    # print("\nRaw Lane Configurations read...", raw_lane_configs, f"\nlength = {len(raw_lane_configs)}")
    
    # Initialize an empty list to store the combined dictionaries
    combined_list = []
    
    # Iterate through the list of signalized intersections
    for intersection in signalized_int_data:
        intersection_id = intersection.get("Intersection ID")
        combined_dict = {"Intersection ID": intersection_id}
    
        # Initialize approach data for all possible directions
        approach_data = {
            direction: {"Approach Delay": None, "Approach LOS": None}
            for direction in ['EB', 'WB', 'NB', 'SB', 'NE', 'NW', 'SE', 'SW']
        }
    
        # Access lane configurations for the current intersection
        lane_config = lane_configurations[int(intersection_id) - 1]
        raw_lane_config = raw_lane_configs[int(intersection_id) - 1]
    
        # Initialize dictionaries to store v/c ratio, LOS, and delay values
        directions = ['EB', 'WB', 'NB', 'SB', 'NE', 'NW', 'SE', 'SW']
        vc_ratio_values = {direction: [None, None, None] for direction in directions}
        los_values = {direction: [None, None, None] for direction in directions}
        delay_values = {direction: [None, None, None] for direction in directions}
    
        # Flags to ensure each type of data is added once
        vc_ratio_added = False
        ln_grp_los_added = False
        los_added = False
    
        # Process each term in the intersection data
        for term, row_index in intersection.items():
            if term not in terms_to_match:
                continue  # Skip terms not in the list
    
            term_lower = term.lower()
            row_data = df.iloc[row_index].replace("", "-").fillna("-").tolist()[1:]  # Exclude first column value
    
            # Process v/c ratio values
            if "v/c ratio(x)" == term_lower or "v/c ratio" == term_lower:
                for idx, direction in enumerate(directions):
                    start_index = idx * 3  # Adjust based on 3 values per direction
                    if start_index + 2 < len(row_data):
                        vc_ratio_values[direction] = row_data[start_index:start_index + 3]
                    else:
                        vc_ratio_values[direction] = ["-", "-", "-"]
                if not vc_ratio_added:
                    combined_dict[term] = [value for value in row_data if value != '-']
                    vc_ratio_added = True
    
            # Process LOS values
            if "lngrp los" == term_lower or "los" == term_lower:
                for idx, direction in enumerate(directions):
                    start_index = idx * 3
                    if start_index + 2 < len(row_data):
                        los_values[direction] = row_data[start_index:start_index + 3]
                    else:
                        los_values[direction] = ["-", "-", "-"]
                if "lngrp los" == term_lower and not ln_grp_los_added:
                    combined_dict[term] = [value for value in row_data if value != '-']
                    ln_grp_los_added = True
                elif "los" == term_lower and not los_added and not ln_grp_los_added:
                    combined_dict[term] = [value for value in row_data if value != '-']
                    los_added = True
    
            # Process delay values
            if "control delay (s/veh)" == term_lower or "lngrp delay(d), s/veh" == term_lower:
                for idx, direction in enumerate(directions):
                    start_index = idx * 3
                    if start_index + 2 < len(row_data):
                        delay_values[direction] = row_data[start_index:start_index + 3]
                    else:
                        delay_values[direction] = ["-", "-", "-"]
    
            # Process approach delay
            if "approach delay" in term_lower:
                filtered_row_data = [value for value in row_data if value != '-']
                for idx, direction in enumerate(directions):
                    if lane_config.get(direction) == '-':
                        filtered_row_data.insert(idx, '-')
                for idx, direction in enumerate(directions[:len(filtered_row_data)]):
                    approach_data[direction]["Approach Delay"] = filtered_row_data[idx]
    
            # Process approach LOS
            if term_lower == "approach los":
                filtered_row_data = [value for value in row_data if value != '-']
                for idx, direction in enumerate(directions):
                    if lane_config.get(direction) == '-':
                        filtered_row_data.insert(idx, '-')
                for idx, direction in enumerate(directions[:len(filtered_row_data)]):
                    approach_data[direction]["Approach LOS"] = filtered_row_data[idx]
    
            # Add other terms to combined_dict
            if term_lower not in ["v/c ratio", "los", "v/c ratio(x)", "lngrp los"]:
                combined_dict[term] = [value for value in row_data if value != '-']
    
        # Finalize combined_dict
        value_set = [vc_ratio_values, delay_values, los_values]
        
        """
        *** Preference algorithm
        """
        # Process each type of value set: vc_ratio, los, and delay
        for values, term in zip(value_set, list(combined_dict.keys())[1:]):
            for direction in directions:
                # Identify non-zero values based on the type of data
                if values == los_values:
                    # For LOS values, consider grades A-F (skipping '-' or None)
                    non_zero_values = [value if value not in ["-", None] else "Z" for value in values[direction]]
                    max_non_zero_value = max(non_zero_values)
                else:
                    # For vc_ratio and delay values, treat them as floats
                    non_zero_values = [float(value) if value not in ["-", "0", None] else float('inf') for value in values[direction]]
                    non_inf_values = [v for v in non_zero_values if v != float('inf')]
                    min_non_zero_value = min(non_zero_values)
                    max_non_zero_value = max(non_inf_values) if non_inf_values else float('-inf')
        
                # Determine if we need to update based on raw lane configuration
                needs_update = any(
                    value not in ["-", "0", None] and lane_config in ["0", None]
                    for value, lane_config in zip(values[direction], raw_lane_config[direction])
                )
                
                # Skip updating if all configurations are valid
                if not needs_update:
                    # print(f"Skipping updates for {direction} in {values} as all configurations are valid.")
                    continue
                
                # Apply the update logic based on the type of data
                for i, (value, lane_config) in enumerate(zip(values[direction], raw_lane_config[direction])):
                    if value in ["-", "0", None] and lane_config not in ["0", None]:
                        if values == los_values:
                            values[direction][i] = max_non_zero_value  # Higher letter for LOS
                        else:
                            values[direction][i] = str(max_non_zero_value)
                
                # Replace the lowest non-zero value with '-'
                if min_non_zero_value != float('inf') and values != los_values:
                    min_index = non_zero_values.index(min_non_zero_value)
                    values[direction][min_index] = '-'
                
                # Debug print
                # print(f"Updated {direction} for {term}: {values[direction]}")
            
            # Filter out '-' values and add to combined_dict
            filtered_values = {direction: [value for value in values[direction] if value != '-' and value != 'Z'] for direction in directions}
            combined_dict[term] = [value for sublist in filtered_values.values() for value in sublist]
        
            # Debug print for combined_dict update
            # print(f"Added to '{term}': {combined_dict[term]}")
            
        # Merge approach data into combined_dict and append to final lists
        combined_dict.update(approach_data)
        combined_list.append(combined_dict)
        # id_combined_list.append((intersection_id, combined_dict))
        
        # print(f"\nCombined data for signalized Intersection ID {intersection_id}: \n{combined_dict}")
    
    # Remove empty lane configurations
    for direction in ['EB', 'WB', 'NB', 'SB', 'NE', 'NW', 'SE', 'SW']:
        for config in lane_configurations:
            if config.get(direction) == '-':
                del config[direction]
    
    # print()
    # idx = 0
    # for i in combined_list:
    #     idx += 1
    #     print(f"Signalized Intersection #{idx}: \n{i}\n")
    # print(f"ID combined list: {id_combined_list}")
    
    # print(f"Lane Configurations (length {len(lane_configurations)}): \n{lane_configurations}")
    
    twsc_overall, synchro_overall, hcm_overall = parse_overall_data_v2(file_path)
    twsc_intersections = parse_twsc_approach(df)
    twsc_intersection_directions, original_twsc_directions, combined_mvmt_names = process_directions(twsc_overall, lane_configurations)
    
    # print(original_twsc_directions)
    # print(f"\nCombined Movement Names: {combined_mvmt_names}")
    # print("\nTWSC Summary Directions: ", twsc_intersection_directions)
    # print()
    print(twsc_intersections)
    # for i in twsc_intersections: 
    #     id_combined_list.append(i)
    
    # idx = 0
    # for i in id_combined_list:
    #     idx += 1
    #     print(f"\nCombined list (Result Set #{idx}): {i}")
    # print()
    # print(f"\nCombined list with TWSC data (length = {len(combined_list)}):\n{combined_list}\n")
    combined_list.extend(twsc_intersections)
    
    # Create an empty DataFrame to hold all intersections' data
    final_df = pd.DataFrame()
    
    general_terms = {
        'v/c': ['V/c ratio', 'V/c ratio(x)', 'LnGrp v/c'],
        'delay': ['Control delay (s/veh)', 'LnGrp Delay(d), s/veh'],
        'los': ['LOS', 'LnGrp LOS']
    }
    
    # Sort combined_list by Intersection ID for ordered processing
    combined_list_sorted = sorted(combined_list, key=lambda x: int(x.get("Intersection ID", "ID")))
    
    # print(f"Combined list for '{file_path}': {combined_list_sorted}")
    
    # Iterate over each item in the sorted combined_list
    # for idx, item in enumerate(combined_list_sorted, 1):
    #     print(f"Intersection #{idx} (ID: {item.get('Intersection ID')}):")
        
    #     # Iterate over the keys and values of each dictionary
    #     for key, value in item.items():
    #         # If the value is a list, print it in a readable format
    #         if isinstance(value, list):
    #             value_str = ', '.join(map(str, value))
    #             print(f"  {key}: [{value_str}]")
    #         else:
    #             print(f"  {key}: {value}")
        
    #     print("\n" + "-"*50)  # Add a separator line between intersections
    
    # Combine both lists and sort by the "index" key
    combined_overall_data = sorted(synchro_overall + hcm_overall, key=lambda x: x.get('index', 0))
    overall_idx = 0
    
    # print(f"\nOverall Data = {combined_overall_data}", '\n')
    
    # Process each intersection in the sorted list
    for data_dict in combined_list_sorted:
        intersection_id = data_dict.get("Intersection ID")
        
        # Control printing of Intersection ID only once per direction set
        intersection_id_printed = False
        
        # Find the matching lane configuration for this intersection
        lane_config = next((config for config in lane_configurations if config.get("Intersection ID") == intersection_id), None)
        
        # Check if the intersection has TWSC data
        twsc_summary_result = next((twsc for twsc in twsc_overall if twsc.get("ID") == str(intersection_id)), None)
        twsc_summary_directions = next((twsc_dir for twsc_dir in twsc_intersection_directions if twsc_dir.get("ID") == str(intersection_id)), None)
        # original_twsc = next((og_dir for og_dir in original_twsc_directions if og_dir.get("ID") == str(intersection_id)), None)
        
        # Prefer TWSC summary if available
        if twsc_summary_result and twsc_summary_directions:
            lane_config = None
        # Skip intersections without lane config or TWSC summary
        if not lane_config and not twsc_summary_result:
            print(f"No lane configuration or TWSC summary found for Intersection ID: {intersection_id}")
            continue
        
        # print(data_dict, '\n')
        
        # Prepare data for the intersection's DataFrame
        intersection_data = []
    
        # Separate indexing for v/c, LOS, and Delay values
        j = 0
        
        # Retrieve all entries from combined_overall_data for this intersection
        overall_data = [item for item in combined_overall_data if item['ID'] == str(intersection_id)]
        # Default values for the overall data row (to be updated if data is found)
        overall_los = '-'
        overall_delay = '-'

        # Processing if lane configuration is available
        if lane_config:
            for direction, lanes in lane_config.items():
                # Skip the "Intersection ID" key in lane_config
                if direction == "Intersection ID":
                    continue
                
                # Retrieve approach delay and LOS for the current direction
                approach_delay = data_dict.get(direction, {}).get("Approach Delay", '-')
                approach_los = data_dict.get(direction, {}).get("Approach LOS", '-')
                
                if approach_delay is None:
                    approach_delay = '-'
                if approach_los is None:
                    approach_los = '-'
                # Loop through each lane in the direction
                for i, lane in enumerate(lanes):
                    # Print the Intersection ID only once at the start of the set
                    intersection_id_str = str(intersection_id) if not intersection_id_printed else ''
                    intersection_id_printed = True
                    direction_value = direction if i == 0 else ''  # Only print direction once
                    
                    # Get v/c, LOS, and Delay values based on general terms dictionary
                    vc_value = los_value = delay_value ='-'
    
                    # Check and get v/c value from general terms
                    for term in general_terms['v/c']:
                        if term in data_dict:
                            vc_value = data_dict[term][j] if j < len(data_dict[term]) else '-'
                            break
    
                    # Check and get LOS value from general terms
                    for term in general_terms['los']:
                        if term in data_dict:
                            los_value = data_dict[term][j] if j < len(data_dict[term]) else '-'
                            break
    
                    # Check and get Delay value from general terms
                    for term in general_terms['delay']:
                        if term in data_dict:
                            delay_value = data_dict[term][j] if j < len(data_dict[term]) else '-'
                            break
                    
                    if vc_value and los_value and delay_value != '-':
                        # Append the row for this lane
                        intersection_data.append([intersection_id_str, direction_value, lane, vc_value, los_value, delay_value])
    
                    # Increment the indices for v/c, LOS, and Delay values
                    j += 1
                
                if approach_delay != 0:
                    # Add an overall row for this direction
                    intersection_data.append(['', f"{direction} Overall", '', '-', f'{approach_los}', f'{approach_delay}'])

            if overall_idx == 0:
                overall_los = overall_data[0].get("los")
                overall_delay = overall_data[0].get("delay")
                overall_idx += 1
            else:
                try:
                    # Attempt to access the second element in the overall_data list
                    overall_los = overall_data[1].get("los")
                    overall_delay = overall_data[1].get("delay")
                    overall_idx = 1  # Indicate that the second item was used
                except IndexError:
                    # If the second element does not exist, fall back to the first element
                    if overall_data:
                        overall_los = overall_data[0].get("los")
                        overall_delay = overall_data[0].get("delay")
                        overall_idx = 0  # Indicate that the first item was used
                    else:
                        # If overall_data is empty, set default values
                        overall_los = '-'
                        overall_delay = '-'
                        overall_idx = None  # No data available
            if overall_delay != 0 or '-':
                # Add an overall row for this intersection, including data from Synchro and HCM
                intersection_data.append(['', "Overall", '', '-', overall_los, overall_delay])
                
        # Processing if TWSC summary data is available
        if twsc_summary_result:
            print(f"\nIntersection {intersection_id}")

            print("\nDirections: ", twsc_summary_directions, '\n')
            print("Values:", twsc_summary_result)

            # Iterate through TWSC summary directions
            for direction, movement_values in twsc_summary_result.items():
                # Skip the "ID" key in TWSC summary
                if direction == "ID":
                    continue
                
                print(direction[:2], movement_values)
                
                # Retrieve approach delay and LOS for the current direction
                # approach_delay = data_dict.get(direction[:2], {}).get("Approach Delay", '-')
                # approach_los = data_dict.get(direction[:2], {}).get("Approach LOS", '-')
                
                # Find the lane configuration in the TWSC summary for this direction
                if direction in twsc_summary_result:
                    if movement_values[3] == '-':
                        continue
                    lane_data = twsc_summary_result[direction]
                else:
                    lane_data = ('-', '-', '-', '-')  # Default or placeholder value
                
                
                # Unpack v/c, LOS, and Delay values from TWSC data
                vc_value, los_value, delay_value, capacity_value = (
                    lane_data if isinstance(lane_data, tuple) else ('-', '-', '-', '-')
                )
                
                # print(f"Capacity: {capacity_value}")
                
                # Add an entry for the TWSC summary direction
                intersection_id_str = str(intersection_id) if not intersection_id_printed else ''
                direction_value = twsc_summary_directions[direction[:2]]
                print(direction_value)
                # Append the row for this direction (from TWSC summary)
                # if isinstance(direction)
                # for d in direction_value:
                # If the direction value is a list (like ['L', 'T']), process each lane in the list
                if isinstance(direction_value, list):
                    for dir_val in direction_value:
                        key = direction[:2] + dir_val
                        lane_data = twsc_summary_result[key]
                        print(lane_data)
                        # For each lane, append a row with the lane and corresponding values
                        if lane_data != ('-', '-', '-', '-'):
                            intersection_data.append([
                                intersection_id_str,  # Intersection ID (if not printed)
                                direction[:2],        # Direction value (e.g., 'SW')
                                dir_val,              # Lane value (e.g., 'L', 'T')
                                vc_value,            # V/c value
                                los_value,           # LOS value
                                delay_value          # Delay value
                            ])
                else:
                    # If not a list, process the single direction normally
                    intersection_data.append([
                        intersection_id_str,  # Intersection ID (if not printed)
                        direction[:2],      # Direction value (e.g., 'WB')
                        direction_value,         # Lane data (since it's not a list here)
                        vc_value,            # V/c value
                        los_value,           # LOS value
                        delay_value          # Delay value
                    ])
                intersection_id_printed = True
                # print(intersection_data)
                
                # Add the "Overall" row for the direction only once after processing all lanes
                # if direction[:2] not in overall_row_added:
                #     intersection_data.append([
                #         '', 
                #         f"{direction[:2]} Overall", 
                #         '', 
                #         '-', 
                #         f'{approach_los}', 
                #         f'{approach_delay}'
                #     ])
                #     overall_row_added[direction[:2]] = True
        
        if intersection_data != []:
            # Add a blank row to separate intersections
            intersection_data.append([''] * 6)
        
            # Create a DataFrame for the current intersection's data
            intersection_df = pd.DataFrame(intersection_data, columns=['Intersection ID', 'Direction', 'Lane', 'V/c', 'LOS', 'Delay'])
    
            # Append it to the final DataFrame
            final_df = pd.concat([final_df, intersection_df], ignore_index=True)
    
    # Write the final DataFrame to a CSV file
    file_name, _ = os.path.splitext(file_path)
    final_df.to_csv(f"{file_name}-filtered.csv", index=False)


    """
    *** Output for testing
    """
    # i = 0
    # # Initialize the intersection ID from id_combined_list
    # for item in combined_list_sorted:
        
    #     # Determine the intersection ID and the data dictionary based on whether the item is a tuple or dictionary
        
    #     intersection_id = item.get("Intersection ID")
    #     data_dict = item
        
    #     # Print each term and its data in a readable format, excluding direction data (EB, WB, NB, SB)
    #     print(f"Intersection {intersection_id}:")
    #     for term, data in data_dict.items():
    #         if term not in ['EB', 'WB', 'NB', 'SB', 'NE', 'NW', 'SE', 'SW']:  # Only print non-directional data here
    #             # Join data if it's a list, otherwise convert it to a string
    #             if isinstance(data, list):
    #                 data_str = ", ".join(map(str, data))
    #             else:
    #                 data_str = str(data)
    #             print(f"  {term}: {data_str}")
    #     print()
    #     # Print Approach Delay and LOS for each direction (EB, WB, NB, SB)
    #     for direction in ['EB', 'WB', 'NB', 'SB', 'NE', 'NW', 'SE', 'SW']:
    #         # Retrieve approach delay and LOS for the current direction
    #         # print(f"Getting approach data for '{direction}'...")
    #         # print(data_dict.get(direction, {}))
    #         approach_delay = data_dict.get(direction, {}).get("Approach Delay", '-')
    #         approach_los = data_dict.get(direction, {}).get("Approach LOS", '-')
    #         # Only delete if the direction exists in data_dict and conditions are met
    #         print(f"  {direction}: Approach Delay = {approach_delay}, Approach LOS = {approach_los}")
    #     print()
    #     # Find the matching lane configuration for this intersection ID in group_config_data
    #     lane_config = next((config for config in lane_configurations if config.get("Intersection ID") == intersection_id), None)
    #     raw_config = next((raw for raw in raw_lane_configs if raw.get("Intersection ID") == intersection_id), None)
        
    #     # Print lane configurations for the current intersection if available
    #     if lane_config:
    #         lane_config_str = ", ".join(
    #             f"{key}: {value}" for key, value in lane_config.items() if key != "Intersection ID" and value != '-'
    #         )
    #         print(f"  Lane Configurations: {lane_config_str}")
    #     else:
    #         print(f"  No lane configurations found for Intersection ID: {intersection_id}")
        
    #     # Print raw direction configurations for the current intersection if available
    #     if raw_config:
    #         raw_config_str = ", ".join(
    #             f"{key}: {value}" for key, value in raw_config.items() if key != "Intersection ID" and value != [None, None, None]
    #         )
    #         print(f"  Raw Direction Configurations: {raw_config_str}")
    #     else:
    #         print(f"  No raw direction configurations found for Intersection ID: {intersection_id}")

    #     i+=1
    #     # Add a blank line for readability between intersections
    #     print("\n" + "_" * 40 + "\n")
        
    print(f"Total number of useable datasets found: {len(combined_list_sorted)}")
    print("_" * 40 + "\n")


def parse_lane_configs(int_lane_groups, intersection_ids):
    parsed_list = []  # This will store the parsed dictionaries for each group
    raw_data_list = []
    
    for idx, lane_dict in enumerate(int_lane_groups):
        
        intersection_id = intersection_ids[idx]

        # Skip if the intersection ID is already in parsed_list
        if any(parsed_dict.get("Intersection ID") == intersection_id for parsed_dict in parsed_list):
            continue
        
        parsed_dict = {
            "Intersection ID" : intersection_id,
            # Initialize with three None values for L, T, R
            'EB': [None, None, None],
            'WB': [None, None, None],
            'NB': [None, None, None],
            'SB': [None, None, None],
            'NE': [None, None, None],
            'NW': [None, None, None],
            'SE': [None, None, None],
            'SW': [None, None, None]
        }
        
        # Initialize the raw data dictionary
        raw_data_dict = {
            "Intersection ID" : intersection_id,
            'EB': [None, None, None],
            'WB': [None, None, None],
            'NB': [None, None, None],
            'SB': [None, None, None],
            'NE': [None, None, None],
            'NW': [None, None, None],
            'SE': [None, None, None],
            'SW': [None, None, None]
        }

        for direction, value in lane_dict.items():
            if value is None or value == '':
                value = '-'
                continue

            # Process each direction and suffix (L, T, R)
            suffixes = {
                'L': 0,  # Index 0 for Left
                'T': 1,  # Index 1 for Through
                'R': 2   # Index 2 for Right
            }

            for suffix, idx in suffixes.items():
                # Parse the value for numbers and special characters < and >
                if direction.endswith(suffix) :                    
                    # Store the raw value directly in raw_data_dict in the correct position
                    direction_prefix = direction[:-1]
                    if direction_prefix in raw_data_dict:
                        raw_data_dict[direction_prefix][idx] = value  # Store unparsed raw value
                        
                    parsed_value = ''

                    if '<' in value:
                        parsed_value += 'L'  # Leading left if < is present
                    number_part = ''.join(
                        filter(str.isdigit, value))  # Get the number
                    if number_part:
                        # Repeat based on the number
                        parsed_value += suffix * int(number_part)
                        
                    else:
                        parsed_value += suffix

                    if '>' in value:
                        parsed_value += 'R'  # Trailing right if > is present

                    # Store the parsed value in the respective direction and suffix position
                    # Get the prefix like EB, WB, etc.
                    direction_prefix = direction[:-1]
                    if direction_prefix in parsed_dict:
                        parsed_dict[direction_prefix][idx] = parsed_value or None
                    
        # Remove None values from each list in the parsed_dict
        for key in list(parsed_dict.keys()):
            if key != "Intersection ID":  # Don't touch the Intersection ID key
                parsed_dict[key] = [value for value in parsed_dict[key] if value is not None]
                # If the list is empty (no valid values), set it to '-'
                if not parsed_dict[key]:
                    parsed_dict[key] = '-'
                    
        # Clean up the raw_data_dict in the same way
        for key in raw_data_dict:
            if key != "Intersection ID":
                raw_data_dict[key] = [value for value in raw_data_dict[key]]
                if not raw_data_dict[key]:
                    raw_data_dict[key] = '-'
        
        # Debugging output
        # print(f"\nParsed Lane Config (Intersection #{intersection_id}):\n{parsed_dict} \nRaw Lane Config (Intersection #{intersection_id}):\n{raw_data_dict}")
        
        # Append the parsed_dict for this lane group to the final list
        parsed_list.append(parsed_dict)
        raw_data_list.append(raw_data_dict)

    return parsed_list, raw_data_list

if __name__ == "__main__":
    # read_input_file("test-input.xlsx")
    test_report_1 = "test/Test Report 1.txt"
    test_report_2 = "test/Test Report 2.txt"
    test_report_3 = "test/Test Report 3.txt"
    test_report_4 = "test/Test Report 4.txt"
    test_twsc = "test/TEST TWSC.txt"
    test_awsc = "test/TEST AWSC.txt"
    
    test_report_1_csv = "test-report-1.csv"
    test_report_2_csv = "test-report-2.csv"
    test_report_3_csv = "test-report-3.csv"
    test_report_4_csv = "test-report-4.csv"
    test_twsc_csv = "test-twsc.csv"
    test_awsc_csv = "test-awsc.csv"
    
    # parse_overall_data_v2(file)  # Gets the data for overall
    
    # Testing with Test Report 1.txt
    print('\n' + "*"*35 + "\n| Results for 'Test Report 1.txt' |\n" + "*"*35 +'\n')
    extract_data_to_csv(test_report_1, test_report_1_csv)
    
    # Testing with Test Report 2.
    print('\n' + "*"*35 + "\n| Results for 'Test Report 2.txt' |\n" + "*"*35 +'\n')
    extract_data_to_csv(test_report_2, test_report_2_csv)
    
    # Testing with Test Report 3.txt
    print('\n' + "*"*35 + "\n| Results for 'Test Report 3.txt' |\n" + "*"*35 +'\n')
    extract_data_to_csv(test_report_3, test_report_3_csv)
    
    # Testing with Test Report 3.txt
    print('\n' + "*"*35 + "\n| Results for 'Test Report 4.txt' |\n" + "*"*35 +'\n')
    extract_data_to_csv(test_report_4, test_report_4_csv)
    
    # print('\n' + "*"*35 + "\n| Results for 'TEST TWSC.txt' |\n" + "*"*35 +'\n')
    # extract_data_to_csv(test_twsc, test_twsc_csv)
    
    print('\n' + "*"*35 + "\n| Results for 'TEST AWSC.txt' |\n" + "*"*35 +'\n')
    extract_data_to_csv(test_awsc, test_awsc_csv)
    
    # lane_groups = separate_characters(movement)
    # print(f"\nLane groups:\n{lane_groups}")
    # write_to_excel(file, movement, delay, vc, los)
