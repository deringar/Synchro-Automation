# -*- coding: utf-8 -*-
"""
Created on Tue Oct  1 13:29:29 2024

@author: alex.dering
"""

from openpyxl import load_workbook, Workbook

def load_input_workbook(file_path):
    """Load the input Excel file and return the active sheet."""
    workbook = load_workbook(filename=file_path)
    return workbook.active

def create_output_workbook():
    """Create a new workbook and set up the basic structure for the output."""
    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "Results"
    output_sheet["A1"] = "[Lanes]"
    output_sheet["A2"] = "Lane Group Data"
    
    headers = [
        "RECORDNAME", "INTID", "NBL", "NBT", "NBR", 
        "SBL", "SBT", "SBR", "EBL", "EBT", "EBR", 
        "WBL", "WBT", "WBR", "PED", "HOLD"
    ]
    
    # Label the header row in the output file
    for col, header in enumerate(headers, start=1):
        output_sheet.cell(row=3, column=col).value = header
    
    return output_workbook, output_sheet

def find_intersections(sheet):
    """Find all intersections (integer values) in column A and return a dictionary of intersection IDs with their corresponding row."""
    intersections = {}
    consecutive_empty_cells = 0

    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value is None:
            consecutive_empty_cells += 1
            if consecutive_empty_cells >= 50:
                break
        else:
            consecutive_empty_cells = 0
            if isinstance(cell_value, int):
                intersections[cell_value] = row

    return intersections

def find_directions(sheet, start_row):
    """Search column C for the directions 'EB', 'WB', 'NB', 'SB' starting from the given row."""
    directions = ["EB", "WB", "NB", "SB"]
    found_directions = {}

    for row in range(start_row, sheet.max_row + 1):
        direction_value = sheet.cell(row=row, column=3).value
        if direction_value in directions and direction_value not in found_directions:
            found_directions[direction_value] = row
            if len(found_directions) == len(directions):
                break

    return found_directions

def find_turns(sheet, direction_row):
    """Search column D starting from the direction row for 'L', 'T', 'R'. Return a dictionary of found turns and their rows."""
    turn_values = {"L": None, "T": None, "R": None}

    for row in range(direction_row, sheet.max_row + 1):
        turn_value = sheet.cell(row=row, column=4).value
        if turn_value in ["L", "T", "R"]:
            turn_values[turn_value] = row
        if all(turn_values.values()):
            break

    return turn_values

def process_intersections(sheet, intersections):
    """Process each intersection to find corresponding directions and turns. Return a dictionary of intersection results."""
    intersection_results = {}

    for intersection_id, row_with_int in intersections.items():
        found_directions = find_directions(sheet, row_with_int)

        direction_turn_results = {}
        for direction, direction_row in found_directions.items():
            turn_values = find_turns(sheet, direction_row)

            # Combine direction and turn types (e.g., 'EBL', 'WBT') and store their corresponding rows
            for turn_type, row_found in turn_values.items():
                if row_found is not None:
                    combined_key = f"{direction}{turn_type}"
                    direction_turn_results[combined_key] = row_found

        intersection_results[intersection_id] = direction_turn_results

    return intersection_results

def write_output(output_sheet, intersections, intersection_results):
    """Write intersection IDs and results to the output workbook."""
    output_start_row = 4  # Start writing from row 4

    for intersection_id in intersections.keys():
        # Write the intersection ID into cells B4, B5, and B6 for each intersection
        output_sheet[f"B{output_start_row}"] = intersection_id
        output_sheet[f"B{output_start_row + 1}"] = intersection_id
        output_sheet[f"B{output_start_row + 2}"] = intersection_id

        # Write "Volume", "PHF", and "HeavyVehicles" into column A
        output_sheet[f"A{output_start_row}"] = "Volume"
        output_sheet[f"A{output_start_row + 1}"] = "PHF"
        output_sheet[f"A{output_start_row + 2}"] = "HeavyVehicles"

        # Skip 3 rows for each intersection block
        output_start_row += 4

def save_output_workbook(output_workbook, file_path="Results.xlsx"):
    """Save the output workbook to a given file path."""
    output_workbook.save(file_path)
    print(f"Output file saved as {file_path}")

def read_input_file(file_path):
    """Main function to read input, process intersections, and generate an output Excel file."""
    # Load the input sheet
    sheet = load_input_workbook(file_path)

    # Create the output workbook and sheet
    output_workbook, output_sheet = create_output_workbook()

    # Step 1: Find all intersections
    intersections = find_intersections(sheet)
    print(f"Found intersections: {intersections}")

    # Step 2: Process intersections and find directions and turns
    intersection_results = process_intersections(sheet, intersections)

    # Step 3: Write the results to the output sheet
    write_output(output_sheet, intersections, intersection_results)

    # Step 4: Save the output workbook
    save_output_workbook(output_workbook)

    return intersection_results

def main():
    """Main function to run the script."""
    file_path = input("Enter the path to the input Excel file: ")
    read_input_file(file_path)

if __name__ == "__main__":
    main()
